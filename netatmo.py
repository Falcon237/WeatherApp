import requests
import json
import datetime
import csv
import time
import os
from datetime import timedelta
from typing import Dict, List, Any, Optional
import logging

# --- NEUE IMPORTE FÜR DIE GUI ---
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
import threading
from queue import Queue

# Für Excel-Formatierung (optional - Fallback auf CSV wenn nicht installiert)
try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    Workbook = None
    load_workbook = None
    Font = None
    Alignment = None

# ----------------------------------
# VEREINFACHTE, FUNKTIONIERENDE VERSION
# ----------------------------------

# Netatmo API: Sensor-Typen auf API-Parameter mappen
SENSOR_API_MAP = {
    "Temperature": "temperature",
    "Humidity": "humidity",
    "CO2": "co2",
    "Noise": "noise",
    "Pressure": "pressure",
    "Rain": "sum_rain_1",
    "WindStrength": "windstrength",
    "WindAngle": "windangle",
    "GustStrength": "guststrength",
    "GustAngle": "gustangle",
}

# Einheiten für die Sensoren
SENSOR_UNITS = {
    "Temperature": "\u00b0C",
    "Humidity": "%",
    "CO2": "ppm",
    "Noise": "dB",
    "Pressure": "mbar",
    "Rain": "mm",
    "WindStrength": "km/h",
    "WindAngle": "\u00b0",
    "GustStrength": "km/h",
    "GustAngle": "\u00b0",
}


class NetatmoDataDownloader:
    def __init__(self, client_id: str, client_secret: str, access_token: str, log_queue: Optional[Queue] = None):
        self.client_id = client_id
        self.client_secret = client_secret
        self.access_token = access_token

        # API URLs
        self.data_url = "https://api.netatmo.com/api/getstationsdata"
        self.measure_url = "https://api.netatmo.com/api/getmeasure"

        # Rate limiting settings
        self.request_delay = 2.0
        self.max_retries = 5
        self.retry_delay = 10.0

        # Setup logging
        self.log_queue = log_queue
        self.setup_logging()

        # Headers
        self.headers = {
            "Authorization": f"Bearer {self.access_token}"
        }

        # Data storage
        self.all_data = []

    def log_message(self, level: str, message: str):
        """Sendet Log-Nachrichten an die Konsole, Datei und optional an die GUI."""
        log_entry = f"{datetime.datetime.now().strftime('%H:%M:%S')} | {level.upper()} | {message}"

        # An die GUI senden, falls eine Queue vorhanden ist
        if self.log_queue:
            self.log_queue.put(log_entry)

        # An den Standard-Logger senden
        if level == "info":
            self.logger.info(message)
        elif level == "warning":
            self.logger.warning(message)
        elif level == "error":
            self.logger.error(message)
        elif level == "debug":
            self.logger.debug(message)
        else:
            self.logger.info(message)

    def setup_logging(self):
        """Setup detailliertes Logging"""
        file_handler = logging.FileHandler(
            'netatmo_download.log', encoding='utf-8')
        file_handler.setLevel(logging.DEBUG)
        formatter = logging.Formatter(
            '%(asctime)s | %(levelname)s | %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
        file_handler.setFormatter(formatter)

        self.logger = logging.getLogger('NetatmoDownloader')
        self.logger.setLevel(logging.DEBUG)
        if not self.logger.handlers:
            self.logger.addHandler(file_handler)
            self.logger.propagate = False

    def make_request(self, url: str, params: Optional[Dict] = None,
                     description: str = "API request") -> Optional[requests.Response]:
        """Robuste API-Anfrage mit Retry-Mechanismus"""
        for attempt in range(self.max_retries):
            try:
                self.log_message(
                    "debug", f"🌐 {description} (Versuch {attempt + 1}/{self.max_retries})")

                if attempt > 0:
                    delay = self.retry_delay * (2 ** (attempt - 1))
                    self.log_message(
                        "info", f"⏳ Warte {delay:.1f}s vor Wiederholung...")
                    time.sleep(delay)
                else:
                    time.sleep(self.request_delay)

                response = requests.get(
                    url, headers=self.headers, params=params, timeout=45)

                if response.status_code == 200:
                    self.log_message("debug", f"✅ {description} erfolgreich")
                    return response
                elif response.status_code == 429:
                    self.log_message(
                        "warning", f"⚠️ Rate limit erreicht (429), warte 90s...")
                    time.sleep(90)
                    continue
                elif response.status_code == 403:
                    self.log_message(
                        "error", f"❌ Authentifizierung fehlgeschlagen (403). Prüfen Sie den Access Token.")
                    return None
                else:
                    self.log_message(
                        "warning", f"⚠️ Unerwarteter Status {response.status_code}: {response.text[:200]}")

            except requests.exceptions.RequestException as e:
                self.log_message(
                    "warning", f"📡 Request-Fehler bei {description}: {str(e)}")
            except Exception as e:
                self.log_message(
                    "error", f"💥 Unerwarteter Fehler bei {description}: {str(e)}")

        self.log_message(
            "error", f"❌ {description} nach {self.max_retries} Versuchen fehlgeschlagen")
        return None

    def get_stations_data(self) -> Optional[Dict]:
        """Hole Stationsdaten"""
        self.log_message("info", "🏠 Hole Netatmo-Stationsdaten...")
        response = self.make_request(
            self.data_url, description="Stationsdaten abrufen")
        if response is None:
            return None
        try:
            data = response.json()
            if not data.get("body", {}).get("devices", []):
                self.log_message(
                    "error", "❌ Keine Geräte in der Antwort gefunden")
                return None
            self.log_message(
                "info", f"✅ {len(data['body']['devices'])} Gerät(e) gefunden")
            return data
        except json.JSONDecodeError as e:
            self.log_message("error", f"❌ JSON-Parsing Fehler: {str(e)}")
            return None

    def extract_modules(self, devices: List[Dict]) -> Dict[str, Dict]:
        """Extrahiere alle Module und ihre verfügbaren Sensoren"""
        modules = {}
        for device in devices:
            device_id = device.get("_id")
            if not device_id:
                continue

            data_types = device.get("data_type", [])
            if data_types:
                name = device.get("module_name") or device.get(
                    "station_name", "Hauptstation")
                modules[device_id] = {"name": name, "data_types": data_types}
                self.log_message(
                    "info", f"📱 Hauptmodul: {name} ({', '.join(data_types)})")

            for module in device.get("modules", []):
                module_id = module["_id"]
                module_data_types = module.get("data_type", [])
                if module_data_types:
                    module_name = module.get(
                        "module_name", f"Modul_{module_id}")
                    modules[module_id] = {
                        "name": module_name, "data_types": module_data_types}
                    self.log_message(
                        "info", f"📡 Modul: {module_name} ({', '.join(module_data_types)})")

        self.log_message("info", f"✅ {len(modules)} Module gefunden.")
        return modules

    def get_temperature_data(self, device_id: str, module_id: str, module_name: str,
                             start_date: datetime.datetime, end_date: datetime.datetime) -> List[Dict]:
        all_points = []
        current_start = start_date

        while current_start < end_date:
            params = {
                "device_id": device_id,
                "module_id": module_id,
                "scale": "max",
                "type": "temperature",
                "date_begin": int(current_start.timestamp()),
                "date_end": int(end_date.timestamp()),
                "limit": "1024",
                "optimize": "false",
                "real_time": "false"
            }
            self.log_message(
                "debug", f"📊 Temperatur-Request: {current_start.strftime('%d.%m.%Y')} bis {end_date.strftime('%d.%m.%Y')}")

            response = self.make_request(
                self.measure_url, params, f"Temperaturdaten für {module_name}")
            if response is None:
                break

            try:
                body = response.json().get("body")
                if not body:
                    self.log_message(
                        "warning", f"⚠️ Keine Daten für {module_name} ab {current_start.strftime('%d.%m.%Y')}")
                    break

                if isinstance(body, dict):
                    items = list(body.items())
                elif isinstance(body, list):
                    items = [(entry.get("beg_time"), entry.get("value"))
                             for entry in body]
                else:
                    break

                batch = []
                last_ts = None
                for timestamp, values in items:
                    try:
                        dt = datetime.datetime.fromtimestamp(int(timestamp))
                        temp = values[0] if values and len(
                            values) > 0 else None
                        if temp is not None:
                            batch.append({
                                "Module": module_name,
                                "Datum": dt.strftime('%d.%m.%Y'),
                                "Uhrzeit": dt.strftime('%H:%M'),
                                "Temperatur": round(float(temp), 1),
                                "Sortierung": dt
                            })
                            last_ts = int(timestamp)
                    except (ValueError, IndexError, TypeError) as e:
                        self.log_message("debug", f"⚠️ Datenpunkt-Fehler: {e}")
                        continue

                all_points.extend(batch)
                self.log_message(
                    "info", f"✅ {len(batch)} Punkte erhalten (ab {current_start.strftime('%d.%m.%Y')}, gesamt: {len(all_points)})")

                # Weniger als 1024 Punkte → alle Daten im Bereich abgerufen
                if len(batch) < 1024 or last_ts is None:
                    break
                # Nächste Seite ab der Sekunde nach dem letzten Timestamp
                current_start = datetime.datetime.fromtimestamp(last_ts + 1)
                time.sleep(self.request_delay)

            except json.JSONDecodeError as e:
                self.log_message(
                    "error", f"❌ JSON-Fehler für {module_name}: {str(e)}")
                break

        self.log_message(
            "info", f"✅ Gesamt {len(all_points)} Temperatur-Datenpunkte für {module_name}")
        return all_points

    def get_sensor_data(self, device_id: str, module_id: str, module_name: str,
                        sensor_type: str, start_date: datetime.datetime,
                        end_date: datetime.datetime) -> List[Dict]:
        """Holt Daten für einen beliebigen Sensortyp (Temperatur, Luftfeuchtigkeit, CO2, ...)"""
        api_type = SENSOR_API_MAP.get(sensor_type, sensor_type.lower())
        einheit = SENSOR_UNITS.get(sensor_type, "")

        all_points = []
        current_start = start_date

        while current_start < end_date:
            params = {
                "device_id": device_id,
                "module_id": module_id,
                "scale": "max",
                "type": api_type,
                "date_begin": int(current_start.timestamp()),
                "date_end": int(end_date.timestamp()),
                "limit": "1024",
                "optimize": "false",
                "real_time": "false"
            }
            self.log_message(
                "debug", f"📊 {sensor_type}-Request: {current_start.strftime('%d.%m.%Y')} bis {end_date.strftime('%d.%m.%Y')}")

            response = self.make_request(
                self.measure_url, params, f"{sensor_type}-Daten für {module_name}")
            if response is None:
                break

            try:
                body = response.json().get("body")
                if not body:
                    self.log_message(
                        "warning", f"⚠️ Keine {sensor_type}-Daten für {module_name} ab {current_start.strftime('%d.%m.%Y')}")
                    break

                if isinstance(body, dict):
                    items = list(body.items())
                elif isinstance(body, list):
                    items = [(entry.get("beg_time"), entry.get("value"))
                             for entry in body]
                else:
                    break

                batch = []
                last_ts = None
                for timestamp, values in items:
                    try:
                        dt = datetime.datetime.fromtimestamp(int(timestamp))
                        val = values[0] if values and len(values) > 0 else None
                        if val is not None:
                            batch.append({
                                "Module": module_name,
                                "Datum": dt.strftime('%d.%m.%Y'),
                                "Uhrzeit": dt.strftime('%H:%M'),
                                "Messwert": sensor_type,
                                "Wert": round(float(val), 2),
                                "Einheit": einheit,
                                "Sortierung": dt
                            })
                            last_ts = int(timestamp)
                    except (ValueError, IndexError, TypeError) as e:
                        self.log_message("debug", f"⚠️ Datenpunkt-Fehler: {e}")
                        continue

                all_points.extend(batch)
                self.log_message(
                    "info", f"✅ {len(batch)} {sensor_type}-Werte (ab {current_start.strftime('%d.%m.%Y')}, gesamt: {len(all_points)})")

                # Weniger als 1024 Punkte → alle Daten im Bereich abgerufen
                if len(batch) < 1024 or last_ts is None:
                    break
                # Nächste Seite ab der Sekunde nach dem letzten Timestamp
                current_start = datetime.datetime.fromtimestamp(last_ts + 1)
                time.sleep(self.request_delay)

            except json.JSONDecodeError as e:
                self.log_message(
                    "error", f"❌ JSON-Fehler für {module_name}: {str(e)}")
                break

        self.log_message(
            "info", f"✅ Gesamt {len(all_points)} {sensor_type}-Werte für {module_name}")
        return all_points

    def download_selected_data(self, selected_modules: Dict[str, str], start_date: datetime.datetime,
                               end_date: datetime.datetime, interval_days: int = 7) -> bool:
        """Hauptfunktion zum Download der Daten für ausgewählte Module mit LÜCKENLOSEN Intervallen."""
        self.all_data = []
        self.log_message(
            "info", f"🚀 Starte Download für {len(selected_modules)} Modul(e)...")
        self.log_message(
            "info", f"📅 Zeitraum: {start_date.strftime('%d.%m.%Y')} bis {end_date.strftime('%d.%m.%Y')}")

        stations_data = self.get_stations_data()
        if not stations_data:
            return False

        device_id = stations_data.get("body", {}).get(
            "devices", [])[0].get("_id")

        # KORREKTUR: Lückenlose Intervalle
        current_date = start_date
        total_intervals = 0
        successful_intervals = 0

        # PHASE 1: Lückenloser Download
        while current_date < end_date:
            # WICHTIG: Nächstes Intervall beginnt GENAU wo das vorherige aufhört
            next_date = current_date + timedelta(days=interval_days)
            if next_date > end_date:
                next_date = end_date

            total_intervals += 1
            self.log_message(
                "info", f"📥 Intervall {total_intervals}: {current_date.strftime('%d.%m.%Y')} - {next_date.strftime('%d.%m.%Y')}")

            interval_success = True
            for module_id, module_name in selected_modules.items():
                data_points = self.get_temperature_data(
                    device_id, module_id, module_name, current_date, next_date)
                if data_points:
                    # Entferne Duplikate am Intervall-Übergang
                    filtered_points = self.remove_overlapping_points(
                        data_points, current_date)
                    self.all_data.extend(filtered_points)
                    self.log_message(
                        "debug", f"✅ {len(filtered_points)} eindeutige Datenpunkte hinzugefügt")
                else:
                    interval_success = False

            if interval_success:
                successful_intervals += 1

            # KORREKTUR: Kein Überlappung - nächstes Intervall startet am Ende des aktuellen
            current_date = next_date

            if current_date < end_date:
                self.log_message("debug", "⏳ Pause zwischen Intervallen...")
                time.sleep(3)

        self.log_message(
            "info", f"🎉 Phase 1 abgeschlossen! {len(self.all_data)} Datenpunkte gesammelt.")
        self.log_message(
            "info", f"📊 Erfolgreiche Intervalle: {successful_intervals}/{total_intervals}")

        # PHASE 2: Aggressivere Lücken-Nachladung
        if self.all_data:
            self.log_message("info", "🔍 Analysiere Daten-Lücken...")
            missing_periods = self.find_data_gaps(start_date, end_date)

            if missing_periods:
                self.log_message(
                    "warning", f"⚠️ {len(missing_periods)} Lücken gefunden - starte Nachladung...")

                for i, (gap_start, gap_end) in enumerate(missing_periods, 1):
                    days_missing = (gap_end.date() - gap_start.date()).days + 1
                    self.log_message(
                        "info", f"🔄 Lücke {i}/{len(missing_periods)}: {gap_start.strftime('%d.%m.%Y')} - {gap_end.strftime('%d.%m.%Y')} ({days_missing} Tage)")
                    # Lücken jeder Größe in 3-Tage-Stücke aufteilen und nachladen
                    sub_start = gap_start
                    while sub_start < gap_end:
                        sub_end = min(sub_start + timedelta(days=3), gap_end)
                        for module_id, module_name in selected_modules.items():
                            gap_data = self.get_temperature_data(
                                device_id, module_id, module_name, sub_start, sub_end)
                            if gap_data:
                                self.all_data.extend(gap_data)
                                self.log_message(
                                    "info", f"✅ {len(gap_data)} Punkte nachgeladen ({sub_start.strftime('%d.%m.%Y')} - {sub_end.strftime('%d.%m.%Y')})")
                        sub_start = sub_end
                        time.sleep(1)

                self.log_message(
                    "info", f"🎉 Nachladung abgeschlossen! Gesamt: {len(self.all_data)} Datenpunkte")
            else:
                self.log_message(
                    "info", "✅ Keine Lücken gefunden - perfekte Datenabdeckung!")

        return True

    def remove_overlapping_points(self, data_points: List[Dict], interval_start: datetime.datetime) -> List[Dict]:
        """Entfernt überlappende Datenpunkte zwischen Intervallen.
        Schlüssel: (Timestamp, Sensortyp) – gleiche Timestamps für verschiedene
        Sensoren werden NICHT als Duplikate gewertet."""
        if not data_points or not self.all_data:
            return data_points

        existing_keys = {
            (p['Sortierung'], p.get('Messwert', ''))
            for p in self.all_data
        }

        return [
            p for p in data_points
            if (p['Sortierung'], p.get('Messwert', '')) not in existing_keys
        ]

    def find_data_gaps(self, start_date: datetime.datetime, end_date: datetime.datetime) -> List[tuple]:
        """Findet Lücken in den heruntergeladenen Daten"""
        if not self.all_data:
            return []

        try:
            # Verfügbare Tage
            available_dates = {row['Sortierung'].date()
                               for row in self.all_data}

            # Alle erwarteten Tage
            expected_dates = set()
            _day = start_date.date()
            while _day <= end_date.date():
                expected_dates.add(_day)
                _day += timedelta(days=1)

            # Fehlende Tage
            missing_dates = sorted(expected_dates - available_dates)

            if not missing_dates:
                return []

            # Gruppiere aufeinanderfolgende fehlende Tage zu Perioden
            gaps = []
            current_gap_start = missing_dates[0]
            current_gap_end = missing_dates[0]

            for i in range(1, len(missing_dates)):
                if missing_dates[i] == current_gap_end + timedelta(days=1):
                    # Aufeinanderfolgender Tag
                    current_gap_end = missing_dates[i]
                else:
                    # Lücke beendet, speichere sie
                    if (current_gap_end - current_gap_start).days >= 0:  # Mindestens 1 Tag
                        gaps.append((
                            datetime.datetime.combine(
                                current_gap_start, datetime.time.min),
                            datetime.datetime.combine(
                                current_gap_end, datetime.time.max)
                        ))

                    # Neue Lücke beginnt
                    current_gap_start = missing_dates[i]
                    current_gap_end = missing_dates[i]

            # Letzte Lücke hinzufügen
            if (current_gap_end - current_gap_start).days >= 0:
                gaps.append((
                    datetime.datetime.combine(
                        current_gap_start, datetime.time.min),
                    datetime.datetime.combine(
                        current_gap_end, datetime.time.max)
                ))

            # Log der gefundenen Lücken
            for gap_start, gap_end in gaps:
                days_missing = (gap_end.date() - gap_start.date()).days + 1
                self.log_message(
                    "warning", f"❌ Lücke: {gap_start.strftime('%d.%m.%Y')} - {gap_end.strftime('%d.%m.%Y')} ({days_missing} Tage)")

            return gaps

        except Exception as e:
            self.log_message("error", f"❌ Fehler bei Lückenanalyse: {e}")
            return []

    def export_to_excel_multi_sensor(self, filename: str) -> bool:
        """Excel-Export für Multi-Sensor-Daten mit Charts"""
        if not self.all_data:
            self.log_message(
                "error", "❌ Keine Daten zum Exportieren verfügbar.")
            return False

        try:
            self.log_message(
                "info", f"📊 Exportiere {len(self.all_data)} Multi-Sensor-Datenpunkte nach {filename}...")

            # SCHRITT 0: Datei-Zugriff prüfen (wie vorher)
            if os.path.exists(filename):
                try:
                    with open(filename, 'r+b'):
                        pass
                    self.log_message(
                        "info", "📁 Datei existiert und ist nicht gesperrt - überschreibe...")
                except PermissionError:
                    import time
                    timestamp = datetime.datetime.now().strftime("_%H%M%S")
                    base_name = filename.rsplit('.', 1)[0]
                    extension = filename.rsplit(
                        '.', 1)[1] if '.' in filename else 'xlsx'
                    filename = f"{base_name}{timestamp}.{extension}"
                    self.log_message(
                        "warning", f"⚠️ Original-Datei ist geöffnet - verwende: {filename}")

            # Daten sortieren
            data = sorted(self.all_data, key=lambda r: (
                r['Module'], r['Messwert'], r['Sortierung']))
            if not data:
                self.log_message("warning", "⚠️ Keine Daten vorhanden.")
                return False

            export_cols = ['Module', 'Datum', 'Uhrzeit',
                           'Messwert', 'Wert', 'Einheit']

            # Wenn openpyxl nicht verfügbar -> direkt CSV-Export
            if not OPENPYXL_AVAILABLE:
                csv_filename = filename.replace('.xlsx', '.csv') if filename.endswith(
                    '.xlsx') else filename + '.csv'
                with open(csv_filename, 'w', newline='', encoding='utf-8') as _f:
                    _w = csv.DictWriter(
                        _f, fieldnames=export_cols, extrasaction='ignore')
                    _w.writeheader()
                    _w.writerows(data)
                self.log_message(
                    "warning", "⚠️ openpyxl nicht installiert - Daten als CSV gespeichert: " + csv_filename)
                self.log_message(
                    "info", "💡 Für Excel-Export: conda create -n netatmo python openpyxl requests -y")
                return True

            # SCHRITT 1: Excel-Export mit mehreren Tabs
            self.log_message("info", "💾 Schreibe Multi-Sensor Excel-Datei...")

            max_retries = 3
            for attempt in range(max_retries):
                try:
                    wb_write = Workbook()
                    # Tab 1: Alle Rohdaten
                    ws_all = wb_write.active
                    ws_all.title = 'Alle Sensordaten'
                    ws_all.append(export_cols)
                    for row in data:
                        ws_all.append([row.get(c, '') for c in export_cols])

                        # Tab 2: Nach Sensortyp getrennte Tabs
                    sensor_types = list({r['Messwert'] for r in data})
                    sensor_cols = ['Module', 'Datum',
                                   'Uhrzeit', 'Wert', 'Einheit']
                    for sensor_type in sensor_types:
                        sensor_data = [
                            r for r in data if r['Messwert'] == sensor_type]
                        tab_name = sensor_type[:31]
                        ws_sensor = wb_write.create_sheet(title=tab_name)
                        ws_sensor.append(sensor_cols)
                        for row in sensor_data:
                            ws_sensor.append([row.get(c, '')
                                             for c in sensor_cols])

                    # Tab 3: Tägliche Zusammenfassung
                    self.log_message(
                        "info", "📈 Erstelle tägliche Sensor-Übersicht...")
                    daily_stats = self.create_daily_sensor_stats(data)
                    if daily_stats:
                        ws_daily = wb_write.create_sheet(
                            title='Täglich Übersicht')
                        daily_headers = list(daily_stats[0].keys())
                        ws_daily.append(daily_headers)
                        for row in daily_stats:
                            ws_daily.append([row[h] for h in daily_headers])

                    wb_write.save(filename)
                    wb_write.close()

                    self.log_message(
                        "info", f"✅ Excel-Basis-Datei erfolgreich erstellt (Versuch {attempt + 1})")
                    break

                except PermissionError as e:
                    if attempt < max_retries - 1:
                        timestamp = datetime.datetime.now().strftime("_%H%M%S")
                        base_name = filename.rsplit('.', 1)[0]
                        extension = filename.rsplit(
                            '.', 1)[1] if '.' in filename else 'xlsx'
                        filename = f"{base_name}{timestamp}.{extension}"
                        self.log_message(
                            "warning", f"⚠️ Zugriff verweigert - Versuch {attempt + 2} mit: {filename}")
                        time.sleep(1)
                    else:
                        self.log_message(
                            "error", f"❌ Datei kann nicht erstellt werden nach {max_retries} Versuchen")
                        return False

            # SCHRITT 2: Formatierung mit openpyxl
            self.log_message(
                "info", "🎨 Formatiere Multi-Sensor Excel-Datei...")

            try:
                wb = load_workbook(filename)

                # Formatiere alle Worksheets
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    ws.auto_filter.ref = ws.dimensions

                    # Header formatieren
                    header_font = Font(bold=True)
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center')

                    # Spaltenbreiten anpassen
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 20)
                        ws.column_dimensions[column_letter].width = adjusted_width

                # SCHRITT 3: Charts für numerische Sensoren erstellen
                self.log_message("info", "📊 Erstelle Multi-Sensor Charts...")
                self.create_multi_sensor_charts(wb, data)

                # Speichern
                for save_attempt in range(3):
                    try:
                        wb.save(filename)
                        wb.close()
                        break
                    except PermissionError:
                        if save_attempt < 2:
                            self.log_message(
                                "warning", f"⚠️ Speichern fehlgeschlagen - Versuch {save_attempt + 2}")
                            time.sleep(2)
                        else:
                            wb.close()
                            self.log_message(
                                "error", "❌ Konnte formatierte Datei nicht speichern - Basis-Datei verfügbar")

            except Exception as format_error:
                self.log_message(
                    "warning", f"⚠️ Formatierung fehlgeschlagen: {str(format_error)}")
                self.log_message(
                    "info", "📁 Basis-Excel-Datei ist trotzdem verfügbar!")

            self.log_message(
                "info", f"✅ Multi-Sensor Excel-Datei erfolgreich erstellt: {filename}")
            self.log_message(
                "info", "📈 Mit Auto-Filter, Formatierung und Charts für alle Sensoren!")

            # Statistik
            min_datum = min(r['Sortierung'] for r in data).strftime('%d.%m.%Y')
            max_datum = max(r['Sortierung'] for r in data).strftime('%d.%m.%Y')
            sensor_count = len({r['Messwert'] for r in data})
            module_count = len({r['Module'] for r in data})
            self.log_message(
                "info", f"📊 {sensor_count} Sensortypen, {module_count} Module, Datenbereich: {min_datum} bis {max_datum}")

            return True

        except Exception as e:
            self.log_message(
                "error", f"❌ Fehler beim Multi-Sensor Excel-Export: {str(e)}")
            import traceback
            self.log_message("error", f"❌ Traceback: {traceback.format_exc()}")

            # Fallback: CSV-Export
            try:
                csv_filename = filename.replace('.xlsx', '.csv')
                _cols = ['Module', 'Datum', 'Uhrzeit',
                         'Messwert', 'Wert', 'Einheit']
                with open(csv_filename, 'w', newline='', encoding='utf-8') as _f:
                    _w = csv.DictWriter(
                        _f, fieldnames=_cols, extrasaction='ignore')
                    _w.writeheader()
                    _w.writerows(self.all_data)
                self.log_message(
                    "info", f"💾 Fallback: Daten als CSV gespeichert: {csv_filename}")
                return True
            except:
                self.log_message("error", "❌ Auch CSV-Export fehlgeschlagen")
                return False

    def create_daily_sensor_stats(self, data: List[Dict]) -> List[Dict]:
        """Erstellt tägliche Statistiken für alle Sensoren"""
        try:
            groups: Dict[tuple, Dict] = {}
            for row in data:
                datum = row['Sortierung'].date()
                key = (datum, row['Module'], row['Messwert'])
                if key not in groups:
                    groups[key] = {'values': [],
                                   'einheit': row.get('Einheit', '')}
                try:
                    groups[key]['values'].append(float(row['Wert']))
                except (ValueError, TypeError):
                    pass

            daily_stats = []
            for (datum, modul, messwert), grp in sorted(groups.items()):
                values = grp['values']
                if values:
                    daily_stats.append({
                        'Datum': datum.strftime('%d.%m.%Y'),
                        'Modul': modul,
                        'Messwert': messwert,
                        'Min': round(min(values), 2),
                        'Max': round(max(values), 2),
                        'Durchschnitt': round(sum(values) / len(values), 2),
                        'Einheit': grp['einheit'],
                        'Messungen': len(values),
                    })

            return daily_stats

        except Exception as e:
            self.log_message(
                "error", f"❌ Fehler bei täglichen Sensor-Statistiken: {str(e)}")
            return []

    def create_multi_sensor_charts(self, workbook, data):
        """Erstellt Charts für verschiedene Sensortypen"""
        try:
            from openpyxl.chart import LineChart, Reference

            # Nur numerische Sensoren für Charts
            numeric_sensors = ['Temperature', 'Humidity',
                               'CO2', 'Noise', 'Pressure', 'Rain']

            # Für jeden Sensor-Typ und jedes Modul ein Chart
            for sensor_type in {r['Messwert'] for r in data}:
                if sensor_type not in numeric_sensors:
                    continue

                for module in {r['Module'] for r in data}:
                    module_sensor_data = [
                        r for r in data if r['Messwert'] == sensor_type and r['Module'] == module]

                    if len(module_sensor_data) < 2:
                        continue

                    # Tägliche Werte für Chart (manuell aggregieren)
                    day_groups: Dict[object, list] = {}
                    for r in module_sensor_data:
                        day = r['Sortierung'].date()
                        try:
                            day_groups.setdefault(
                                day, []).append(float(r['Wert']))
                        except (ValueError, TypeError):
                            pass

                    daily_data = sorted([
                        {'Sortierung': day,
                         'Min': round(min(vals), 2),
                         'Max': round(max(vals), 2),
                         'Durchschnitt': round(sum(vals) / len(vals), 2)}
                        for day, vals in day_groups.items() if vals
                    ], key=lambda x: x['Sortierung'])

                    if len(daily_data) < 2:
                        continue

                    # Neues Worksheet für Chart
                    chart_ws_name = f"Chart {module} {sensor_type}"[:31]
                    chart_ws = workbook.create_sheet(chart_ws_name)

                    # Daten ins Worksheet
                    chart_ws.cell(row=1, column=1, value="Tag des Jahres")
                    chart_ws.cell(row=1, column=2, value=f"Min {sensor_type}")
                    chart_ws.cell(row=1, column=3, value=f"Max {sensor_type}")
                    chart_ws.cell(row=1, column=4, value=f"Ø {sensor_type}")

                    for i, row in enumerate(daily_data, start=2):
                        day_of_year = row['Sortierung'].timetuple().tm_yday
                        chart_ws.cell(row=i, column=1, value=day_of_year)
                        chart_ws.cell(row=i, column=2, value=row['Min'])
                        chart_ws.cell(row=i, column=3, value=row['Max'])
                        chart_ws.cell(row=i, column=4,
                                      value=row['Durchschnitt'])

                    # Chart erstellen
                    chart = LineChart()
                    chart.title = f"{sensor_type} - {module}"
                    chart.x_axis.title = "Tag des Jahres"

                    einheit = module_sensor_data[0].get('Einheit', '')
                    chart.y_axis.title = f"{sensor_type} ({einheit})"
                    chart.width = 15
                    chart.height = 10

                    # Daten hinzufügen
                    data_rows = len(daily_data)
                    min_data = Reference(
                        chart_ws, min_col=2, min_row=1, max_col=2, max_row=data_rows + 1)
                    max_data = Reference(
                        chart_ws, min_col=3, min_row=1, max_col=3, max_row=data_rows + 1)
                    avg_data = Reference(
                        chart_ws, min_col=4, min_row=1, max_col=4, max_row=data_rows + 1)

                    chart.add_data(min_data, titles_from_data=True)
                    chart.add_data(max_data, titles_from_data=True)
                    chart.add_data(avg_data, titles_from_data=True)

                    # X-Achse
                    days = Reference(chart_ws, min_col=1,
                                     min_row=2, max_row=data_rows + 1)
                    chart.set_categories(days)

                    # Farben setzen
                    if len(chart.series) >= 3:
                        # Min (Blau)
                        chart.series[0].graphicalProperties.line.solidFill = "0066CC"
                        # Max (Rot)
                        chart.series[1].graphicalProperties.line.solidFill = "CC0000"
                        # Avg (Grün)
                        chart.series[2].graphicalProperties.line.solidFill = "00AA00"

                    # Chart einfügen
                    chart_ws.add_chart(chart, "F2")

                    self.log_message(
                        "debug", f"✅ Chart für {module} - {sensor_type} erstellt")

            self.log_message(
                "info", "✅ Multi-Sensor Charts erfolgreich erstellt!")

        except ImportError:
            self.log_message(
                "warning", "⚠️ openpyxl.chart nicht verfügbar - Charts übersprungen")
        except Exception as e:
            self.log_message(
                "error", f"❌ Fehler beim Erstellen der Multi-Sensor Charts: {str(e)}")
            import traceback
            self.log_message(
                "debug", f"❌ Multi-Chart Traceback: {traceback.format_exc()}")

        except Exception as e:
            self.log_message("error", f"❌ Fehler bei Lücken-Analyse: {str(e)}")
            return []

    def export_to_excel(self, filename: str) -> bool:
        """EINFACHER UND FUNKTIONIERENDER Excel-Export MIT CHARTS und FEHLERBEHANDLUNG"""
        if not self.all_data:
            self.log_message(
                "error", "❌ Keine Daten zum Exportieren verfügbar.")
            return False

        try:
            self.log_message(
                "info", f"📊 Exportiere {len(self.all_data)} Datenpunkte nach {filename}...")

            # SCHRITT 0: Prüfe ob Datei bereits existiert und geöffnet ist
            if os.path.exists(filename):
                try:
                    # Versuche die Datei zu öffnen um zu prüfen ob sie gesperrt ist
                    with open(filename, 'r+b'):
                        pass
                    self.log_message(
                        "info", "📁 Datei existiert und ist nicht gesperrt - überschreibe...")
                except PermissionError:
                    # Datei ist geöffnet - erstelle alternative
                    import time
                    timestamp = datetime.datetime.now().strftime("_%H%M%S")
                    base_name = filename.rsplit('.', 1)[0]
                    extension = filename.rsplit(
                        '.', 1)[1] if '.' in filename else 'xlsx'
                    filename = f"{base_name}{timestamp}.{extension}"
                    self.log_message(
                        "warning", f"⚠️ Original-Datei ist geöffnet - verwende: {filename}")

            # Daten sortieren
            data = sorted(self.all_data, key=lambda r: (
                r['Module'], r['Sortierung']))
            if not data:
                self.log_message("warning", "⚠️ Keine Daten vorhanden.")
                return False

            export_cols_e = ['Module', 'Datum', 'Uhrzeit', 'Temperatur']

            # Wenn openpyxl nicht verfügbar -> direkt CSV-Export
            if not OPENPYXL_AVAILABLE:
                csv_filename = filename.replace('.xlsx', '.csv') if filename.endswith(
                    '.xlsx') else filename + '.csv'
                with open(csv_filename, 'w', newline='', encoding='utf-8') as _f:
                    _w = csv.DictWriter(
                        _f, fieldnames=export_cols_e, extrasaction='ignore')
                    _w.writeheader()
                    _w.writerows(data)
                self.log_message(
                    "warning", "⚠️ openpyxl nicht installiert - Daten als CSV gespeichert: " + csv_filename)
                self.log_message(
                    "info", "💡 Für Excel-Export: conda create -n netatmo python openpyxl requests -y")
                return True

            # SCHRITT 1: Sicherer Excel-Export mit Retry-Mechanismus
            self.log_message("info", "💾 Schreibe Excel-Datei...")

            max_retries = 3
            for attempt in range(max_retries):
                try:
                    wb_write_e = Workbook()
                    # Tab 1: Alle Rohdaten
                    ws_all_e = wb_write_e.active
                    ws_all_e.title = 'Alle Daten'
                    ws_all_e.append(export_cols_e)
                    for row in data:
                        ws_all_e.append([row.get(c, '')
                                        for c in export_cols_e])

                    # Tab 2: Tägliche Min/Max-Werte erstellen
                    self.log_message(
                        "info", "📈 Erstelle tägliche Min/Max-Übersicht...")
                    daily_stats_e = self.create_daily_stats(data)
                    ws_daily_e = wb_write_e.create_sheet(
                        title='Täglich Min-Max')
                    if daily_stats_e:
                        daily_headers_e = list(daily_stats_e[0].keys())
                        ws_daily_e.append(daily_headers_e)
                        for row in daily_stats_e:
                            ws_daily_e.append([row[h]
                                              for h in daily_headers_e])

                    wb_write_e.save(filename)
                    wb_write_e.close()

                    self.log_message(
                        "info", f"✅ Excel-Basis-Datei erfolgreich erstellt (Versuch {attempt + 1})")
                    break

                except PermissionError as e:
                    if attempt < max_retries - 1:
                        # Neuen Dateinamen mit Timestamp erstellen
                        timestamp = datetime.datetime.now().strftime("_%H%M%S")
                        base_name = filename.rsplit('.', 1)[0]
                        extension = filename.rsplit(
                            '.', 1)[1] if '.' in filename else 'xlsx'
                        filename = f"{base_name}{timestamp}.{extension}"
                        self.log_message(
                            "warning", f"⚠️ Zugriff verweigert - Versuch {attempt + 2} mit: {filename}")
                        time.sleep(1)
                    else:
                        self.log_message(
                            "error", f"❌ Datei kann nicht erstellt werden nach {max_retries} Versuchen")
                        self.log_message("error", "💡 LÖSUNGEN:")
                        self.log_message(
                            "error", "   1. Schließen Sie die Excel-Datei falls geöffnet")
                        self.log_message(
                            "error", "   2. Wählen Sie einen anderen Speicherort (nicht OneDrive)")
                        self.log_message(
                            "error", "   3. Speichern Sie auf dem Desktop oder C:\\Temp\\")
                        return False

            # SCHRITT 2: Nachbearbeitung mit openpyxl
            self.log_message("info", "🎨 Formatiere Excel-Datei...")

            try:
                wb = load_workbook(filename)

                # --- Formatierung Tab 1: Alle Daten ---
                ws1 = wb['Alle Daten']
                ws1.auto_filter.ref = ws1.dimensions

                # Spaltenbreiten Tab 1
                ws1.column_dimensions['A'].width = 15  # Module
                ws1.column_dimensions['B'].width = 12  # Datum
                ws1.column_dimensions['C'].width = 10  # Uhrzeit
                ws1.column_dimensions['D'].width = 15  # Temperatur

                # Header formatieren Tab 1
                header_font = Font(bold=True)
                for cell in ws1[1]:
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')

                # Temperatur-Spalte rechtsbündig Tab 1
                for row in range(2, ws1.max_row + 1):
                    ws1.cell(row=row, column=4).alignment = Alignment(
                        horizontal='right')

                # --- Formatierung Tab 2: Täglich Min-Max ---
                ws2 = wb['Täglich Min-Max']
                ws2.auto_filter.ref = ws2.dimensions

                # Spaltenbreiten Tab 2 - ERWEITERT für neue Spalte
                ws2.column_dimensions['A'].width = 12  # Datum
                ws2.column_dimensions['B'].width = 15  # Modul
                ws2.column_dimensions['C'].width = 15  # Min Temperatur
                ws2.column_dimensions['D'].width = 15  # Max Temperatur
                ws2.column_dimensions['E'].width = 15  # Durchschnitt
                ws2.column_dimensions['F'].width = 12  # Messungen (NEU)

                # Header formatieren Tab 2
                for cell in ws2[1]:
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')

                # Zahlen rechtsbündig Tab 2 - ERWEITERT
                for row in range(2, ws2.max_row + 1):
                    for col in [3, 4, 5, 6]:  # Min, Max, Durchschnitt, Messungen
                        ws2.cell(row=row, column=col).alignment = Alignment(
                            horizontal='right')

                # SCHRITT 3: Charts erstellen
                self.log_message("info", "📊 Erstelle Charts...")
                self.create_temperature_charts(wb, ws2)

                # Speichern mit Retry
                for save_attempt in range(3):
                    try:
                        wb.save(filename)
                        wb.close()
                        break
                    except PermissionError:
                        if save_attempt < 2:
                            self.log_message(
                                "warning", f"⚠️ Speichern fehlgeschlagen - Versuch {save_attempt + 2}")
                            time.sleep(2)
                        else:
                            wb.close()
                            self.log_message(
                                "error", "❌ Konnte formatierte Datei nicht speichern - Basis-Datei verfügbar")

            except Exception as format_error:
                self.log_message(
                    "warning", f"⚠️ Formatierung fehlgeschlagen: {str(format_error)}")
                self.log_message(
                    "info", "📁 Basis-Excel-Datei ist trotzdem verfügbar!")

            self.log_message(
                "info", f"✅ Excel-Datei erfolgreich erstellt: {filename}")
            self.log_message(
                "info", "📈 Mit Auto-Filter, Formatierung und Charts!")

            # Statistik
            min_datum = min(r['Sortierung'] for r in data).strftime('%d.%m.%Y')
            max_datum = max(r['Sortierung'] for r in data).strftime('%d.%m.%Y')
            self.log_message(
                "info", f"📊 Datenbereich: {min_datum} bis {max_datum}")

            return True

        except Exception as e:
            self.log_message("error", f"❌ Fehler beim Excel-Export: {str(e)}")
            import traceback
            self.log_message("error", f"❌ Traceback: {traceback.format_exc()}")

            # Fallback: CSV-Export
            try:
                csv_filename = filename.replace('.xlsx', '.csv')
                _cols_e = ['Module', 'Datum', 'Uhrzeit', 'Temperatur']
                with open(csv_filename, 'w', newline='', encoding='utf-8') as _f:
                    _w = csv.DictWriter(
                        _f, fieldnames=_cols_e, extrasaction='ignore')
                    _w.writeheader()
                    _w.writerows(self.all_data)
                self.log_message(
                    "info", f"💾 Fallback: Daten als CSV gespeichert: {csv_filename}")
                return True
            except:
                self.log_message("error", "❌ Auch CSV-Export fehlgeschlagen")
                return False

    def create_daily_stats(self, data: List[Dict]) -> List[Dict]:
        """Erstellt tägliche Min/Max/Durchschnitt-Statistiken mit Lücken-Analyse"""
        try:
            self.log_message(
                "info", f"🔍 Analysiere Rohdaten: {len(data)} Datenpunkte")

            # Analysiere verfügbare Tage
            unique_dates = sorted({row['Sortierung'].date() for row in data})
            self.log_message(
                "info", f"📅 Verfügbare Tage in Rohdaten: {len(unique_dates)}")
            self.log_message(
                "debug", f"📅 Erster Tag: {unique_dates[0]}, Letzter Tag: {unique_dates[-1]}")

            # Prüfe auf Lücken
            if len(unique_dates) > 1:
                _d = unique_dates[0]
                full_date_range = []
                while _d <= unique_dates[-1]:
                    full_date_range.append(_d)
                    _d += timedelta(days=1)
                missing_dates = [
                    d for d in full_date_range if d not in set(unique_dates)]

                if missing_dates:
                    self.log_message(
                        "warning", f"⚠️ {len(missing_dates)} Tage ohne Daten gefunden!")
                    self.log_message(
                        "debug", f"❌ Fehlende Tage: {missing_dates[:10]}...")
                else:
                    self.log_message(
                        "info", "✅ Keine Lücken in den Daten gefunden")

            # Nach Datum und Modul gruppieren
            groups: Dict[tuple, list] = {}
            for row in data:
                key = (row['Sortierung'].date(), row['Module'])
                groups.setdefault(key, []).append(float(row['Temperatur']))

            daily_stats = []
            for (datum, modul), temps in sorted(groups.items()):
                daily_stats.append({
                    'Datum': datum.strftime('%d.%m.%Y'),
                    'Modul': modul,
                    'Min Temperatur': round(min(temps), 1),
                    'Max Temperatur': round(max(temps), 1),
                    'Durchschnitt': round(sum(temps) / len(temps), 1),
                    'Messungen': len(temps),
                })

            self.log_message(
                "info", f"📊 Tägliche Statistiken erstellt: {len(daily_stats)} Tag-Modul-Kombinationen")

            return daily_stats

        except Exception as e:
            self.log_message(
                "error", f"❌ Fehler bei täglichen Statistiken: {str(e)}")
            import traceback
            self.log_message("error", f"❌ Traceback: {traceback.format_exc()}")
            return []

    def create_temperature_charts(self, workbook, worksheet):
        """Erstellt Min/Max-Temperatur-Charts für jedes Modul"""
        try:
            from openpyxl.chart import LineChart, Reference
            from openpyxl.chart.marker import DataPoint

            # Alle Module finden
            modules = set()
            for row in range(2, worksheet.max_row + 1):
                module_name = worksheet.cell(row=row, column=2).value
                if module_name:
                    modules.add(module_name)

            self.log_message(
                "info", f"📊 Erstelle Charts für {len(modules)} Module: {', '.join(modules)}")

            for module in sorted(modules):
                self.log_message(
                    "info", f"📈 Erstelle Chart für Modul: {module}")

                # Daten für dieses Modul sammeln - CHRONOLOGISCH SORTIERT
                module_data = []
                for row in range(2, worksheet.max_row + 1):
                    if worksheet.cell(row=row, column=2).value == module:
                        datum_text = worksheet.cell(row=row, column=1).value
                        min_temp = worksheet.cell(row=row, column=3).value
                        max_temp = worksheet.cell(row=row, column=4).value

                        # Datum für Sortierung konvertieren
                        try:
                            datum_obj = datetime.datetime.strptime(
                                datum_text, '%d.%m.%Y')
                            module_data.append(
                                (datum_obj, datum_text, min_temp, max_temp))
                        except:
                            continue

                if len(module_data) < 2:
                    self.log_message(
                        "warning", f"⚠️ Zu wenig Daten für Chart: {module}")
                    continue

                # WICHTIG: Chronologisch sortieren!
                module_data.sort(key=lambda x: x[0])

                # Neues Worksheet für dieses Modul-Chart
                chart_ws_name = f"Chart {module}"[:31]
                chart_ws = workbook.create_sheet(chart_ws_name)

                # Header
                chart_ws.cell(row=1, column=1, value="Datum")
                chart_ws.cell(row=1, column=2, value="Min °C")
                chart_ws.cell(row=1, column=3, value="Max °C")

                # Daten chronologisch ins Chart-Worksheet kopieren
                for i, (datum_obj, datum_text, min_temp, max_temp) in enumerate(module_data, start=2):
                    chart_ws.cell(row=i, column=1, value=datum_text)
                    chart_ws.cell(row=i, column=2, value=min_temp)
                    chart_ws.cell(row=i, column=3, value=max_temp)

                # Chart erstellen
                chart = LineChart()
                chart.title = f"Temperaturverlauf - {module}"
                chart.style = 10
                chart.x_axis.title = "Datum"
                chart.y_axis.title = "Temperatur (°C)"
                chart.width = 20
                chart.height = 12

                # Daten für Chart
                data_rows = len(module_data)

                # Min-Temperatur Serie (Blau)
                min_data = Reference(
                    chart_ws, min_col=2, min_row=1, max_col=2, max_row=data_rows + 1)
                chart.add_data(min_data, titles_from_data=True)

                # Max-Temperatur Serie (Rot)
                max_data = Reference(
                    chart_ws, min_col=3, min_row=1, max_col=3, max_row=data_rows + 1)
                chart.add_data(max_data, titles_from_data=True)

                # X-Achse (Datum)
                dates = Reference(chart_ws, min_col=1,
                                  min_row=2, max_row=data_rows + 1)
                chart.set_categories(dates)

                # Serien-Farben setzen
                if len(chart.series) >= 2:
                    # Min-Serie (Blau)
                    chart.series[0].graphicalProperties.line.solidFill = "0066CC"
                    chart.series[0].graphicalProperties.line.width = 25000

                    # Max-Serie (Rot)
                    chart.series[1].graphicalProperties.line.solidFill = "CC0000"
                    chart.series[1].graphicalProperties.line.width = 25000

                # Chart ins Worksheet einfügen
                chart_ws.add_chart(chart, "E2")

                # Spaltenbreiten im Chart-Worksheet
                chart_ws.column_dimensions['A'].width = 12
                chart_ws.column_dimensions['B'].width = 10
                chart_ws.column_dimensions['C'].width = 10

                self.log_message(
                    "info", f"✅ Chart für {module} erstellt ({len(module_data)} Datenpunkte)")

            self.log_message(
                "info", f"✅ {len(modules)} Charts erfolgreich erstellt!")

        except ImportError:
            self.log_message(
                "warning", "⚠️ openpyxl.chart nicht verfügbar - Charts übersprungen")
        except Exception as e:
            self.log_message(
                "error", f"❌ Fehler beim Erstellen der Charts: {str(e)}")
            import traceback
            self.log_message(
                "error", f"❌ Chart Traceback: {traceback.format_exc()}")

# ----------------------------------
# LOGIN
# ----------------------------------


class LoginWindow:
    """Login-Fenster: OAuth2 Browser-Flow ODER manuelles Token."""

    NETATMO_AUTH_URL = "https://api.netatmo.com/oauth2/authorize"
    NETATMO_TOKEN_URL = "https://api.netatmo.com/oauth2/token"
    REDIRECT_PORT = 9876
    REDIRECT_URI = f"http://localhost:{REDIRECT_PORT}/callback"
    DEFAULT_CLIENT_ID = os.environ.get("NETATMO_CLIENT_ID", "")
    DEFAULT_CLIENT_SECRET = os.environ.get("NETATMO_CLIENT_SECRET", "")

    def __init__(self, root, on_success):
        self.root = root
        self.on_success = on_success

        root.title("Netatmo – Anmeldung")
        root.geometry("500x540")
        root.resizable(True, True)

        outer = ttk.Frame(root, padding=20)
        outer.pack(fill=tk.BOTH, expand=True)
        outer.columnconfigure(1, weight=1)

        ttk.Label(outer, text="Netatmo Anmeldung",
                  font=("Segoe UI", 13, "bold")).grid(
            row=0, column=0, columnspan=2, pady=(0, 14))

        for i, (lbl, default) in enumerate([
                ("Client ID:",     self.DEFAULT_CLIENT_ID),
                ("Client Secret:", self.DEFAULT_CLIENT_SECRET)]):
            ttk.Label(outer, text=lbl).grid(
                row=i + 1, column=0, sticky="w", padx=(0, 10), pady=4)
            e = ttk.Entry(outer, width=40)
            e.insert(0, default)
            e.grid(row=i + 1, column=1, sticky="ew", pady=4)
            setattr(self, "entry_client_id" if i ==
                    0 else "entry_client_secret", e)

        ttk.Separator(outer, orient="horizontal").grid(
            row=3, column=0, columnspan=2, sticky="ew", pady=(10, 6))
        ttk.Label(outer,
                  text="Redirect URI – einmalig in Netatmo App eintragen:",
                  font=("Segoe UI", 9)).grid(
            row=4, column=0, columnspan=2, sticky="w")
        uri_row = ttk.Frame(outer)
        uri_row.grid(row=5, column=0, columnspan=2, sticky="ew", pady=(2, 8))
        uri_row.columnconfigure(0, weight=1)
        uri_e = ttk.Entry(uri_row, font=("Consolas", 9))
        uri_e.insert(0, self.REDIRECT_URI)
        uri_e.config(state="readonly")
        uri_e.grid(row=0, column=0, sticky="ew")
        ttk.Button(uri_row, text="Kopieren", width=10,
                   command=lambda: (root.clipboard_clear(),
                                    root.clipboard_append(self.REDIRECT_URI))
                   ).grid(row=0, column=1, padx=(6, 0))

        self.status_var = tk.StringVar()
        self.status_label = ttk.Label(
            outer, textvariable=self.status_var, foreground="red", wraplength=460)
        self.status_label.grid(row=6, column=0, columnspan=2, pady=(0, 4))

        btn_row = ttk.Frame(outer)
        btn_row.grid(row=7, column=0, columnspan=2)
        self.login_btn = ttk.Button(
            btn_row, text="Im Browser anmelden",
            command=self._start_oauth, width=22)
        self.login_btn.pack(side=tk.LEFT, padx=6)
        ttk.Button(btn_row, text="Beenden",
                   command=root.destroy, width=14).pack(side=tk.LEFT, padx=6)

        # ── Fallback: Token manuell ────────────────────────────────────────
        ttk.Separator(outer, orient="horizontal").grid(
            row=8, column=0, columnspan=2, sticky="ew", pady=(16, 8))
        ttk.Label(outer,
                  text="Alternative – Token direkt einfügen:\n"
                       "developer.netatmo.com → App → Token generieren"
                       " → Scope: read_station",
                  font=("Segoe UI", 9), foreground="#555").grid(
            row=9, column=0, columnspan=2, sticky="w", pady=(0, 4))

        tok_row = ttk.Frame(outer)
        tok_row.grid(row=10, column=0, columnspan=2, sticky="ew")
        tok_row.columnconfigure(0, weight=1)
        self.entry_token = ttk.Entry(tok_row, font=("Consolas", 9))
        self.entry_token.grid(row=0, column=0, sticky="ew")
        ttk.Button(tok_row, text="Token verwenden", width=16,
                   command=self._use_manual_token).grid(
            row=0, column=1, padx=(6, 0))

        self.token_status_var = tk.StringVar()
        ttk.Label(outer, textvariable=self.token_status_var,
                  foreground="red", wraplength=460).grid(
            row=11, column=0, columnspan=2, pady=(4, 0))

    # ── Browser OAuth ──────────────────────────────────────────────────────

    def _start_oauth(self):
        import urllib.parse
        import secrets
        import webbrowser

        client_id = self.entry_client_id.get().strip()
        client_secret = self.entry_client_secret.get().strip()
        if not client_id or not client_secret:
            self._set_status("Bitte Client ID und Client Secret ausfüllen.")
            return

        self.login_btn.config(state=tk.DISABLED)
        self._set_status("Browser öffnet sich …", "blue")
        self._oauth_state = secrets.token_urlsafe(16)

        auth_url = (
            self.NETATMO_AUTH_URL
            + "?response_type=code"
            + "&client_id=" + urllib.parse.quote(client_id, safe="")
            + "&redirect_uri=" + urllib.parse.quote(self.REDIRECT_URI, safe="")
            + "&scope=read_station"
            + "&state=" + self._oauth_state
        )
        # Server ZUERST starten, dann erst Browser öffnen – kein Race Condition
        self._auth_url_pending = auth_url
        threading.Thread(target=self._wait_for_callback,
                         args=(client_id, client_secret), daemon=True).start()

    def _wait_for_callback(self, client_id, client_secret):
        import http.server
        import urllib.parse
        import webbrowser

        result = {"code": None, "state": None, "error": None}

        class Handler(http.server.BaseHTTPRequestHandler):
            def do_GET(self):
                p = urllib.parse.parse_qs(
                    urllib.parse.urlparse(self.path).query)
                result["code"] = p.get("code",  [None])[0]
                result["state"] = p.get("state", [None])[0]
                result["error"] = p.get("error", [None])[0]
                self.send_response(200)
                self.send_header("Content-type", "text/html; charset=utf-8")
                self.end_headers()
                ok = bool(result["code"])
                self.wfile.write((
                    "<html><body style='font-family:sans-serif;"
                    "text-align:center;margin-top:80px'>"
                    + ("<h2>&#10003; Erfolgreich! Fenster schließen.</h2>"
                       if ok else "<h2>&#10007; Fehlgeschlagen.</h2>")
                    + "</body></html>").encode("utf-8"))

            def log_message(self, *_): pass

        try:
            srv = http.server.HTTPServer(
                ("localhost", self.REDIRECT_PORT), Handler)
            srv.timeout = 180
            # Server ist gebunden – jetzt erst Browser öffnen
            webbrowser.open(self._auth_url_pending)
            srv.handle_request()
            srv.server_close()
        except OSError as e:
            self.root.after(0, lambda: (
                self._set_status(
                    f"Port {self.REDIRECT_PORT} belegt – neu starten."),
                self.login_btn.config(state=tk.NORMAL)))
            return

        if result["error"]:
            self.root.after(0, lambda: (
                self._set_status(f"Netatmo: {result['error']}"),
                self.login_btn.config(state=tk.NORMAL)))
            return
        if not result["code"] or result["state"] != self._oauth_state:
            self.root.after(0, lambda: (
                self._set_status("Abgebrochen oder ungültige Antwort."),
                self.login_btn.config(state=tk.NORMAL)))
            return

        self.root.after(0, lambda: self._set_status(
            "Code erhalten – hole Token …", "blue"))

        code = result["code"]
        CT = {"Content-Type": "application/x-www-form-urlencoded;charset=UTF-8"}

        # Laut Netatmo-Docs: redirect_uri ist OPTIONAL wenn in der App registriert,
        # PFLICHT wenn NICHT registriert. Bei leerem Redirect-URI-Feld in der App
        # kann das Mitsenden der URI zu invalid_client führen.
        attempts = [
            # 1) Standard: alles inkl. redirect_uri und scope
            dict(headers=CT, data={
                "grant_type": "authorization_code",
                "client_id": client_id, "client_secret": client_secret,
                "code": code, "redirect_uri": self.REDIRECT_URI,
                "scope": "read_station"}),
            # 2) Ohne redirect_uri (App hat kein Redirect URI registriert)
            dict(headers=CT, data={
                "grant_type": "authorization_code",
                "client_id": client_id, "client_secret": client_secret,
                "code": code, "scope": "read_station"}),
            # 3) Ohne scope
            dict(headers=CT, data={
                "grant_type": "authorization_code",
                "client_id": client_id, "client_secret": client_secret,
                "code": code, "redirect_uri": self.REDIRECT_URI}),
            # 4) Ohne redirect_uri und ohne scope
            dict(headers=CT, data={
                "grant_type": "authorization_code",
                "client_id": client_id, "client_secret": client_secret,
                "code": code}),
            # 5) Basic Auth statt Body-Credentials
            dict(headers=CT, auth=(client_id, client_secret), data={
                "grant_type": "authorization_code",
                "code": code, "redirect_uri": self.REDIRECT_URI,
                "scope": "read_station"}),
        ]
        last_err = "Unbekannt"
        for i, attempt in enumerate(attempts, 1):
            try:
                resp = requests.post(self.NETATMO_TOKEN_URL,
                                     timeout=15, **attempt)
                if resp.status_code == 200:
                    d = resp.json()
                    creds = {"client_id": client_id,
                             "client_secret": client_secret,
                             "access_token": d["access_token"],
                             "refresh_token": d.get("refresh_token", "")}
                    self.root.after(0, lambda c=creds: self._on_ok(c))
                    return
                try:
                    b = resp.json()
                    last_err = (f"[{i}] HTTP {resp.status_code} – "
                                f"{b.get('error', '?')}: "
                                f"{b.get('error_description', '')}")
                except Exception:
                    last_err = f"[{i}] HTTP {resp.status_code} – {resp.text[:150]}"
            except Exception as exc:
                last_err = f"[{i}] Netzwerk: {exc}"

        hint = ("\n\n→ Nutzen Sie den Token Generator auf:\n"
                "  developer.netatmo.com → App → Token generator\n"
                "  Scope: read_station → Token unten einfügen")
        self.root.after(0, lambda m=last_err + hint: (
            self._set_status(m),
            self.login_btn.config(state=tk.NORMAL)))

    # ── Manuelles Token ────────────────────────────────────────────────────

    def _use_manual_token(self):
        token = self.entry_token.get().strip()
        if not token:
            self.token_status_var.set("Bitte Access Token einfügen.")
            return
        client_id = self.entry_client_id.get().strip()
        client_secret = self.entry_client_secret.get().strip()
        self.token_status_var.set("Prüfe Token …")
        self.root.update()
        threading.Thread(target=self._verify_token_worker,
                         args=(token, client_id, client_secret),
                         daemon=True).start()

    def _verify_token_worker(self, token, client_id, client_secret):
        try:
            resp = requests.get(
                "https://api.netatmo.com/api/getstationsdata",
                headers={"Authorization": f"Bearer {token}"},
                timeout=10)
            if resp.status_code == 200:
                creds = {"client_id": client_id,
                         "client_secret": client_secret,
                         "access_token": token,
                         "refresh_token": ""}
                self.root.after(0, lambda c=creds: self._on_ok(c))
            else:
                try:
                    msg = resp.json().get("error", {}).get(
                        "message", resp.text[:150])
                except Exception:
                    msg = resp.text[:150]
                self.root.after(0, lambda m=msg: self.token_status_var.set(
                    f"Ungültig: {m}"))
        except Exception as exc:
            self.root.after(0, lambda e=exc: self.token_status_var.set(
                f"Fehler: {e}"))

    # ── Helpers ───────────────────────────────────────────────────────────

    def _set_status(self, msg, color="red"):
        self.status_var.set(msg)
        self.status_label.config(foreground=color)

    def _on_ok(self, creds):
        for w in self.root.winfo_children():
            w.destroy()
        self.on_success(creds)


# ----------------------------------
# GUI-KLASSE
# ----------------------------------


class NetatmoGUI:
    def __init__(self, root, client_id: str, client_secret: str, access_token: str):
        self.root = root
        self.root.title("Netatmo Daten-Downloader v3.0 - Alle Sensoren")
        self.root.geometry("900x700")

        self.CLIENT_ID = client_id
        self.CLIENT_SECRET = client_secret
        self.ACCESS_TOKEN = access_token

        self.log_queue = Queue()
        self.downloader = NetatmoDataDownloader(
            self.CLIENT_ID, self.CLIENT_SECRET, self.ACCESS_TOKEN, self.log_queue)
        self.all_modules = {}

        self.create_widgets()
        self.process_log_queue()

        # Lade Module beim Start
        self.load_modules_async()

    def create_widgets(self):
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # --- Einstellungs-Frame ---
        settings_frame = ttk.LabelFrame(
            main_frame, text="1. Einstellungen", padding="10")
        settings_frame.pack(fill=tk.X, pady=5)
        settings_frame.grid_columnconfigure(1, weight=1)
        settings_frame.grid_columnconfigure(3, weight=1)

        # Datum
        ttk.Label(settings_frame, text="Startdatum (JJJJ-MM-TT):").grid(row=0,
                                                                        column=0, padx=5, pady=5, sticky="w")
        self.start_date_entry = ttk.Entry(settings_frame)
        self.start_date_entry.insert(0, "2025-01-01")
        self.start_date_entry.grid(
            row=0, column=1, padx=5, pady=5, sticky="ew")

        ttk.Label(settings_frame, text="Enddatum (JJJJ-MM-TT):").grid(row=0,
                                                                      column=2, padx=5, pady=5, sticky="w")
        self.end_date_entry = ttk.Entry(settings_frame)
        self.end_date_entry.insert(
            0, datetime.datetime.now().strftime("%Y-%m-%d"))
        self.end_date_entry.grid(row=0, column=3, padx=5, pady=5, sticky="ew")

        # --- Auswahl Frame: Module (links) + Sensoren (rechts) ---
        selection_frame = ttk.Frame(main_frame)
        selection_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        selection_frame.columnconfigure(0, weight=1)
        selection_frame.columnconfigure(1, weight=1)

        module_frame = ttk.LabelFrame(
            selection_frame, text="2. Module auswählen", padding="10")
        module_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 3))

        self.module_listbox = tk.Listbox(
            module_frame, selectmode=tk.MULTIPLE, height=6, exportselection=False)
        self.module_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        mod_sb = ttk.Scrollbar(
            module_frame, orient=tk.VERTICAL, command=self.module_listbox.yview)
        mod_sb.pack(side=tk.RIGHT, fill=tk.Y)
        self.module_listbox.config(yscrollcommand=mod_sb.set)
        self.module_listbox.insert(0, "Lade Module...")
        self.module_listbox.bind("<<ListboxSelect>>", self.on_module_select)

        sensor_frame = ttk.LabelFrame(
            selection_frame, text="3. Sensoren auswählen", padding="10")
        sensor_frame.grid(row=0, column=1, sticky="nsew", padx=(3, 0))

        self.sensor_listbox = tk.Listbox(
            sensor_frame, selectmode=tk.MULTIPLE, height=6, exportselection=False)
        self.sensor_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sen_sb = ttk.Scrollbar(
            sensor_frame, orient=tk.VERTICAL, command=self.sensor_listbox.yview)
        sen_sb.pack(side=tk.RIGHT, fill=tk.Y)
        self.sensor_listbox.config(yscrollcommand=sen_sb.set)

        # --- Download-Button ---
        self.download_button = ttk.Button(
            main_frame, text="4. Download starten und als Excel speichern", command=self.start_download_async)
        self.download_button.pack(fill=tk.X, pady=10)

        # --- Progress Bar ---
        progress_frame = ttk.Frame(main_frame)
        progress_frame.pack(fill=tk.X, pady=(0, 5))
        self.progress_label = ttk.Label(progress_frame, text="Bereit")
        self.progress_label.pack(anchor="w")
        self.progress = ttk.Progressbar(
            progress_frame, mode='determinate', maximum=100)
        self.progress.pack(fill=tk.X)

        # --- Log-Fenster ---
        log_frame = ttk.LabelFrame(
            main_frame, text="Status & Logs", padding="10")
        log_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        self.log_text = scrolledtext.ScrolledText(
            log_frame, wrap=tk.WORD, state=tk.DISABLED, height=10)
        self.log_text.pack(fill=tk.BOTH, expand=True)

    def log_to_gui(self, message):
        """Schreibt eine Nachricht in das Log-Textfeld im GUI."""
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)

    def process_log_queue(self):
        """Verarbeitet Nachrichten aus der Log-Queue und zeigt sie an."""
        while not self.log_queue.empty():
            message = self.log_queue.get_nowait()
            self.log_to_gui(message)
        self.root.after(100, self.process_log_queue)

    def load_modules_async(self):
        """Startet das Laden der Module in einem separaten Thread."""
        self.download_button.config(state=tk.DISABLED)
        self.progress_label.config(text="Lade Module von Netatmo API...")
        self.progress.config(value=0)
        threading.Thread(target=self.load_modules_worker, daemon=True).start()

    def load_modules_worker(self):
        """Holt die Moduldaten (läuft im Worker-Thread)."""
        stations_data = self.downloader.get_stations_data()
        if stations_data:
            devices = stations_data.get("body", {}).get("devices", [])
            self.all_modules = self.downloader.extract_modules(devices)

            # GUI im Main-Thread aktualisieren
            self.root.after(0, self.update_module_listbox)
        else:
            self.root.after(0, lambda: messagebox.showerror(
                "Fehler", "Konnte Module nicht laden. Prüfen Sie die API-Credentials und die Logs."))
            self.root.after(
                0, lambda: self.download_button.config(state=tk.NORMAL))
            self.root.after(0, lambda: self.progress_label.config(
                text="Fehler beim Laden"))

    def update_module_listbox(self):
        """Aktualisiert die Listbox mit den geladenen Modulen."""
        self.progress.config(value=0)
        self.progress_label.config(text="Module geladen")
        self.module_listbox.delete(0, tk.END)
        if self.all_modules:
            for module_info in self.all_modules.values():
                self.module_listbox.insert(tk.END, module_info["name"])
            self.module_listbox.select_set(0, tk.END)
            self.download_button.config(state=tk.NORMAL)
            self.root.after(1, self.on_module_select)
        else:
            self.module_listbox.insert(
                tk.END, "Keine Module gefunden.")

    def on_module_select(self, event=None):
        """Aktualisiert die Sensor-Liste basierend auf den ausgewählten Modulen."""
        selected_indices = self.module_listbox.curselection()
        selected_names = {self.module_listbox.get(i) for i in selected_indices}

        available_sensors: set = set()
        for module_info in self.all_modules.values():
            # Wenn nichts selektiert, alle Module anzeigen
            if not selected_names or module_info["name"] in selected_names:
                available_sensors.update(module_info["data_types"])

        current = set(self.sensor_listbox.get(0, tk.END))
        if available_sensors != current:
            self.sensor_listbox.delete(0, tk.END)
            for sensor in sorted(available_sensors):
                self.sensor_listbox.insert(tk.END, sensor)
            self.sensor_listbox.select_set(0, tk.END)

    def update_progress(self, current: int, total: int, message: str):
        """Aktualisiert Progress Bar und Status-Label."""
        if total > 0:
            pct = int(100 * current / total)
            self.progress.config(value=pct)
        self.progress_label.config(text=message)

    def start_download_async(self):
        """Startet den gesamten Download- und Exportprozess mit Progress Bar."""
        # 1. Eingaben validieren
        try:
            start_date = datetime.datetime.strptime(
                self.start_date_entry.get(), "%Y-%m-%d")
            end_date = datetime.datetime.strptime(
                self.end_date_entry.get(), "%Y-%m-%d")
            interval_days = 7

        except ValueError:
            messagebox.showerror(
                "Fehler", "Ungültiges Datumsformat.")
            return

        # 2. Module validieren
        selected_module_indices = self.module_listbox.curselection()
        if not selected_module_indices:
            messagebox.showerror(
                "Fehler", "Bitte mindestens ein Modul auswählen.")
            return

        # 3. Sensoren validieren
        selected_sensor_indices = self.sensor_listbox.curselection()
        if not selected_sensor_indices:
            messagebox.showerror(
                "Fehler", "Bitte mindestens einen Sensor auswählen.")
            return

        # 4. Auswahl zusammenstellen
        selected_module_names = [self.module_listbox.get(
            i) for i in selected_module_indices]
        selected_sensors = [self.sensor_listbox.get(
            i) for i in selected_sensor_indices]

        # Module-IDs finden
        selected_modules = {}
        for module_id, module_info in self.all_modules.items():
            if module_info["name"] in selected_module_names:
                # Nur Sensoren hinzufügen die auch ausgewählt wurden
                available_sensors = [
                    s for s in module_info["data_types"] if s in selected_sensors]
                if available_sensors:
                    selected_modules[module_id] = {
                        "name": module_info["name"],
                        "sensors": available_sensors
                    }

        if not selected_modules:
            messagebox.showerror(
                "Fehler", "Keine gültigen Modul-Sensor-Kombinationen gefunden.")
            return

        # 5. "Speichern unter"-Dialog anzeigen
        default_filename = f"Netatmo_Daten_{start_date.strftime('%Y%m%d')}_bis_{end_date.strftime('%Y%m%d')}.xlsx"
        filename = filedialog.asksaveasfilename(
            title="Excel-Datei speichern unter...",
            defaultextension=".xlsx",
            filetypes=[("Excel-Dateien", "*.xlsx"), ("Alle Dateien", "*.*")]
        )

        if not filename:
            self.log_to_gui("INFO | Export abgebrochen.")
            return

        if filename.endswith('/') or filename.endswith('\\'):
            filename = filename + default_filename

        # 6. Progress Bar vorbereiten
        total_days = (end_date - start_date).days
        self.total_intervals = (total_days // interval_days) + \
            (1 if total_days % interval_days > 0 else 0)
        self.completed_intervals = 0
        self.start_time = time.time()
        self.update_progress(0, self.total_intervals, "Starte Download...")

        # 7. Download im Thread starten
        self.download_button.config(state=tk.DISABLED)
        thread = threading.Thread(
            target=self.download_worker_with_progress,
            args=(selected_modules, start_date,
                  end_date, interval_days, filename),
            daemon=True
        )
        thread.start()

    def download_worker_with_progress(self, selected_modules, start_date, end_date, interval_days, filename):
        """Führt den Download mit Progress-Updates durch."""
        try:
            # Download mit Progress-Callback
            success = self.download_selected_data_with_progress(
                selected_modules, start_date, end_date, interval_days)

            if success and self.downloader.all_data:
                self.root.after(0, lambda: self.update_progress(
                    self.total_intervals, self.total_intervals, "Erstelle Excel..."))
                export_success = self.downloader.export_to_excel_multi_sensor(
                    filename)
                if export_success:
                    self.root.after(0, lambda: messagebox.showinfo(
                        "Erfolg! 🎉", f"Export erfolgreich!\n\nDatei: {filename}\nDatensätze: {len(self.downloader.all_data)}\n\nDie Excel-Datei hat Auto-Filter und Charts."))
                else:
                    self.root.after(0, lambda: messagebox.showerror(
                        "Fehler", "Download erfolgreich, aber Excel-Export fehlgeschlagen."))
            elif success and not self.downloader.all_data:
                self.downloader.log_message(
                    "warning", "Download war erfolgreich, aber es wurden keine Daten gefunden.")
                self.root.after(0, lambda: messagebox.showwarning(
                    "Warnung", "Download erfolgreich, aber keine Daten gefunden."))
            else:
                self.root.after(0, lambda: messagebox.showerror(
                    "Fehler", "Download fehlgeschlagen. Prüfen Sie die Logs."))

        except Exception as e:
            error_msg = f"Unerwarteter Fehler: {str(e)}"
            self.downloader.log_message("error", error_msg)
            self.root.after(
                0, lambda: messagebox.showerror("Fehler", error_msg))

        # GUI am Ende wieder aktivieren
        self.root.after(
            0, lambda: self.download_button.config(state=tk.NORMAL))
        self.root.after(0, lambda: self.update_progress(
            self.total_intervals, self.total_intervals, "Abgeschlossen"))

    def download_selected_data_with_progress(self, selected_modules: Dict, start_date: datetime.datetime,
                                             end_date: datetime.datetime, interval_days: int = 7) -> bool:
        """Download mit Progress-Updates für GUI"""
        self.downloader.all_data = []

        stations_data = self.downloader.get_stations_data()
        if not stations_data:
            return False

        device_id = stations_data.get("body", {}).get(
            "devices", [])[0].get("_id")

        current_date = start_date
        interval_count = 0

        # PHASE 1: Hauptdownload mit Progress
        while current_date < end_date:
            next_date = current_date + timedelta(days=interval_days)
            if next_date > end_date:
                next_date = end_date

            interval_count += 1

            # GUI Progress Update
            message = f"Intervall {interval_count}/{self.total_intervals}"
            self.root.after(0, lambda m=message: self.update_progress(
                interval_count-1, self.total_intervals, m))

            for module_id, module_info in selected_modules.items():
                module_name = module_info["name"]
                for sensor_type in module_info["sensors"]:
                    data_points = self.downloader.get_sensor_data(
                        device_id, module_id, module_name, sensor_type, current_date, next_date)
                    if data_points:
                        filtered_points = self.downloader.remove_overlapping_points(
                            data_points, current_date)
                        self.downloader.all_data.extend(filtered_points)

            current_date = next_date

            # GUI Progress Update
            self.completed_intervals = interval_count
            self.root.after(0, lambda: self.update_progress(self.completed_intervals, self.total_intervals,
                            f"Abgeschlossen: {self.completed_intervals}/{self.total_intervals}"))

            if current_date < end_date:
                time.sleep(1)  # Reduzierte Pause

        # PHASE 2: Lücken nachaden (vereinfacht für Progress)
        if self.downloader.all_data:
            message = "Prüfe Lücken..."
            self.root.after(0, lambda: self.update_progress(
                self.total_intervals, self.total_intervals, message))

            missing_periods = self.downloader.find_data_gaps(
                start_date, end_date)
            if missing_periods:
                self.downloader.log_message(
                    "warning", f"⚠️ {len(missing_periods)} Lücken – starte Nachladung...")
                for gap_start, gap_end in missing_periods:
                    sub_start = gap_start
                    while sub_start < gap_end:
                        sub_end = min(sub_start + timedelta(days=3), gap_end)
                        for module_id, module_info in selected_modules.items():
                            module_name = module_info["name"]
                            for sensor_type in module_info["sensors"]:
                                gap_data = self.downloader.get_sensor_data(
                                    device_id, module_id, module_name, sensor_type, sub_start, sub_end)
                                if gap_data:
                                    self.downloader.all_data.extend(gap_data)
                                time.sleep(0.5)
                        sub_start = sub_end

        return True


def main():
    """Hauptfunktion: Startet die GUI"""
    print("=" * 60)
    print("🌡️  NETATMO DATEN-DOWNLOADER v3.0 - ALLE SENSOREN")
    print("=" * 60)
    print("✨ NEUE FEATURES:")
    print("✅ Alle Sensoren auswählbar (Temperatur, Luftfeuchtigkeit, CO2, etc.)")
    print("✅ Echte Progress Bar mit Zeitschätzung")
    print("✅ Lückenlose Datenabdeckung")
    print("✅ Auto-Filter und Charts")
    print("=" * 60)
    print("Logs werden in 'netatmo_download.log' und im Fenster angezeigt.")
    print("Starte grafische Benutzeroberfläche...")

    root = tk.Tk()

    def on_login_success(creds):
        root.geometry("900x700")
        root.resizable(True, True)
        NetatmoGUI(root, creds["client_id"],
                   creds["client_secret"], creds["access_token"])

    LoginWindow(root, on_success=on_login_success)
    root.mainloop()


if __name__ == "__main__":
    main()
