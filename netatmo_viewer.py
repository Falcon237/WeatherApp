#!/usr/bin/env python3
"""
Netatmo Daten-Viewer
Liest Netatmo Export-Dateien (Excel/CSV) und zeigt interaktive Charts im Browser.
"""

try:
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox
    _HAS_GUI = True
except Exception:
    _HAS_GUI = False
import datetime
import csv
import os
import json
import gzip
import pickle
import sqlite3
import tempfile
import threading
import webbrowser
from collections import defaultdict

import password_lock

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

MODULE_COLORS = [
    '#4dc9f6', '#f67019', '#f53794', '#537bc4',
    '#acc236', '#166a8f', '#00a950', '#8549ba',
    '#e8c83a', '#58595b',
]

# Cache-Format-Version, damit bei künftigen Schema-Änderungen alte Caches
# erkannt und ignoriert werden können.
CACHE_VERSION = 1
ARCHIVE_SCHEMA_VERSION = 1


def app_data_folder() -> str:
    """Persistenter Datenordner (Windows: %APPDATA%, sonst ~)."""
    base = os.environ.get('APPDATA') or os.path.expanduser('~')
    folder = os.path.join(base, 'NetatmoViewer')
    os.makedirs(folder, exist_ok=True)
    return folder


def cache_path() -> str:
    """Pfad zur persistenten Cache-Datei."""
    folder = app_data_folder()
    return os.path.join(folder, 'cache.pkl.gz')


def archive_path() -> str:
    """Pfad zum dauerhaften SQLite-Archiv."""
    folder = app_data_folder()
    return os.path.join(folder, 'netatmo_archive.sqlite')


def load_cache():
    """Lädt gecachte Daten. Gibt (data, meta) oder (None, None) zurück."""
    path = cache_path()
    if not os.path.isfile(path):
        return None, None
    try:
        with gzip.open(path, 'rb') as f:
            obj = pickle.load(f)
        if not isinstance(obj, dict) or obj.get('version') != CACHE_VERSION:
            return None, None
        return obj.get('data', []), obj.get('meta', {})
    except (OSError, pickle.UnpicklingError, EOFError, ValueError):
        return None, None


def save_cache(data: list, meta: dict) -> str:
    """Speichert Daten + Meta-Info persistent. Gibt den Pfad zurück."""
    path = cache_path()
    obj = {'version': CACHE_VERSION, 'data': data, 'meta': meta}
    with gzip.open(path, 'wb', compresslevel=6) as f:
        pickle.dump(obj, f, protocol=pickle.HIGHEST_PROTOCOL)
    return path


def clear_cache() -> bool:
    """Löscht die Cache-Datei. True wenn etwas gelöscht wurde."""
    path = cache_path()
    if os.path.isfile(path):
        try:
            os.remove(path)
            return True
        except OSError:
            return False
    return False


def _connect_archive() -> sqlite3.Connection:
    path = archive_path()
    conn = sqlite3.connect(path)
    conn.execute('PRAGMA journal_mode=WAL')
    conn.execute('PRAGMA synchronous=NORMAL')
    conn.execute(
        '''CREATE TABLE IF NOT EXISTS measurements (
               ts INTEGER NOT NULL,
               date TEXT NOT NULL,
               module TEXT NOT NULL,
               sensor TEXT NOT NULL,
               value REAL NOT NULL,
               unit TEXT NOT NULL,
               source TEXT NOT NULL,
               imported_at TEXT NOT NULL,
               PRIMARY KEY (ts, module, sensor)
           )''')
    conn.execute(
        'CREATE INDEX IF NOT EXISTS idx_measurements_date '
        'ON measurements(date)')
    conn.execute(
        'CREATE INDEX IF NOT EXISTS idx_measurements_module_sensor '
        'ON measurements(module, sensor)')
    conn.execute(
        '''CREATE TABLE IF NOT EXISTS archive_meta (
               key TEXT PRIMARY KEY,
               value TEXT NOT NULL
           )''')
    conn.execute(
        'INSERT OR REPLACE INTO archive_meta(key, value) VALUES (?, ?)',
        ('schema_version', str(ARCHIVE_SCHEMA_VERSION)))
    return conn


def save_archive(data: list, source: str = '') -> tuple:
    """Speichert Messpunkte dauerhaft in SQLite.
    Gibt (pfad, added_count, duplicate_count) zurück.
    """
    path = archive_path()
    if not data:
        return path, 0, 0

    imported_at = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    rows = []
    for d in data:
        ts = int(d['ts'])
        date = str(d.get('date') or datetime.datetime.fromtimestamp(
            ts / 1000).date().isoformat())
        rows.append((
            ts,
            date,
            str(d.get('module', '') or ''),
            str(d.get('sensor', '') or ''),
            float(d['value']),
            str(d.get('unit', '') or ''),
            source,
            imported_at,
        ))

    conn = _connect_archive()
    try:
        before = conn.total_changes
        conn.executemany(
            '''INSERT OR IGNORE INTO measurements
               (ts, date, module, sensor, value, unit, source, imported_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
            rows)
        conn.commit()
        added = conn.total_changes - before
        return path, added, len(rows) - added
    finally:
        conn.close()


def load_archive():
    """Lädt Messpunkte aus dem SQLite-Archiv. Gibt (data, meta) zurück."""
    path = archive_path()
    if not os.path.isfile(path):
        return None, None
    try:
        conn = _connect_archive()
        try:
            rows = conn.execute(
                '''SELECT ts, date, module, sensor, value, unit
                 FROM measurements
                 ORDER BY ts''').fetchall()
            sources = [r[0] for r in conn.execute(
                '''SELECT DISTINCT source
                 FROM measurements
                 WHERE source <> ''
                 ORDER BY source''').fetchall()]
        finally:
            conn.close()
    except (OSError, sqlite3.Error, ValueError):
        return None, None

    data = [{
        'ts': int(ts),
        'date': str(date),
        'module': str(module),
        'sensor': str(sensor),
        'value': float(value),
        'unit': str(unit),
    } for ts, date, module, sensor, value, unit in rows]
    meta = {'sources': sources, 'count': len(data), 'path': path}
    return data, meta


def merge_data(existing: list, new: list) -> tuple:
    """Fügt 'new' zu 'existing' hinzu; dedupliziert per (ts, module, sensor).
    Gibt (merged_list, added_count, duplicate_count) zurück.
    """
    seen = {(d['ts'], d['module'], d['sensor']) for d in existing}
    merged = list(existing)
    added = 0
    dupes = 0
    for d in new:
        key = (d['ts'], d['module'], d['sensor'])
        if key in seen:
            dupes += 1
            continue
        seen.add(key)
        merged.append(d)
        added += 1
    merged.sort(key=lambda x: x['ts'])
    return merged, added, dupes

# ─────────────────────────────────────────────
#  Daten laden
# ─────────────────────────────────────────────


def load_data(filepath: str):
    ext = os.path.splitext(filepath)[1].lower()
    rows = []
    if ext in ('.xlsx', '.xls'):
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl nicht installiert.\n"
                "Bitte CSV-Datei verwenden oder:\n"
                "pip install openpyxl"
            )
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws_name = 'Alle Sensordaten' if 'Alle Sensordaten' in wb.sheetnames else wb.sheetnames[
            0]
        ws = wb[ws_name]
        headers = None
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = [str(c) if c else '' for c in row]
                continue
            if all(v is None for v in row):
                continue
            rows.append(dict(zip(headers, row)))
        wb.close()
    elif ext == '.csv':
        with open(filepath, 'r', encoding='utf-8-sig') as f:
            rows = list(csv.DictReader(f))
    else:
        raise ValueError(f"Nicht unterstütztes Format: {ext}")

    result = []
    for row in rows:
        try:
            datum = str(row.get('Datum',   '') or '').strip()
            uhrzeit = str(row.get('Uhrzeit', '') or '').strip()
            wert_raw = row.get('Wert', None)
            if not datum or wert_raw is None or str(wert_raw).strip() == '':
                continue
            value = float(wert_raw)
            dt_str = f"{datum} {uhrzeit}" if uhrzeit else datum
            dt = None
            for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d',
                        '%d.%m.%Y %H:%M:%S', '%d.%m.%Y %H:%M', '%d.%m.%Y'):
                try:
                    dt = datetime.datetime.strptime(dt_str.strip(), fmt)
                    break
                except ValueError:
                    continue
            if dt is None:
                continue
            result.append({
                'ts':     int(dt.timestamp() * 1000),
                'date':   dt.date().isoformat(),
                'module': str(row.get('Module',   '') or '').strip(),
                'sensor': str(row.get('Messwert', '') or '').strip(),
                'value':  value,
                'unit':   str(row.get('Einheit',  '') or '').strip(),
            })
        except (ValueError, TypeError, KeyError):
            continue
    return result


def aggregate(data, mode: str):
    if mode == 'raw':
        return data
    groups = defaultdict(list)
    for d in data:
        dt = datetime.datetime.fromtimestamp(d['ts'] / 1000)
        bucket = (
            dt.replace(minute=0, second=0, microsecond=0)
            if mode == 'hourly'
            else datetime.datetime.combine(dt.date(), datetime.time())
        )
        groups[(d['module'], d['sensor'], bucket)].append(d['value'])

    units = {(d['module'], d['sensor']): d['unit'] for d in data}
    result = []
    for (module, sensor, dt), vals in groups.items():
        result.append({
            'ts':     int(dt.timestamp() * 1000),
            'date':   dt.date().isoformat(),
            'module': module,
            'sensor': sensor,
            'value':  round(sum(vals) / len(vals), 2),
            'unit':   units.get((module, sensor), ''),
        })
    return sorted(result, key=lambda x: x['ts'])


def prepare_chart_payload(data):
    sensors = sorted({d['sensor'] for d in data})
    modules = sorted({d['module'] for d in data})
    units = {d['sensor']: d['unit'] for d in data}
    colors = {m: MODULE_COLORS[i % len(MODULE_COLORS)]
              for i, m in enumerate(modules)}

    chart_data = {}
    for sensor in sensors:
        chart_data[sensor] = {}
        for module in modules:
            pts = sorted(
                [[d['ts'], d['value']] for d in data
                 if d['sensor'] == sensor and d['module'] == module],
                key=lambda x: x[0]
            )
            if pts:
                chart_data[sensor][module] = pts

    payload = {'sensors': sensors, 'modules': modules,
               'units': units, 'colors': colors, 'data': chart_data}

    public_data, locked_data, locked_modules = password_lock.split_locked_modules(
        chart_data, modules)
    if locked_modules:
        password = password_lock.get_viewer_password()
        payload['data'] = public_data
        payload['locked_modules'] = locked_modules
        if password and locked_data:
            payload['locked_enc'] = password_lock.encrypt_json(
                locked_data, password)

    return payload


# ─────────────────────────────────────────────
#  HTML generieren
# ─────────────────────────────────────────────

HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="de">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Netatmo Daten-Viewer</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/chartjs-adapter-date-fns@3.0.0/dist/chartjs-adapter-date-fns.bundle.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/hammerjs@2.0.8/hammer.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-zoom@2.0.1/dist/chartjs-plugin-zoom.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-datalabels@2.2.0/dist/chartjs-plugin-datalabels.min.js"></script>
<style>
  :root {
    --bg: #0d1117; --card: #161b22; --border: #30363d;
    --text: #e6edf3; --muted: #8b949e; --accent: #58a6ff;
    --green: #3fb950; --orange: #d29922; --red: #f85149;
  }
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { background: var(--bg); color: var(--text); font-family: -apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif; font-size: 14px; }
  header { background: linear-gradient(135deg,#1f2a3c,#0d1117); padding: 18px 24px; border-bottom: 1px solid var(--border); display:flex; align-items:center; gap:12px; }
  header h1 { font-size:20px; font-weight:600; color:var(--accent); }
  header .sub { color:var(--muted); font-size:12px; }
  .layout { display:flex; height:calc(100vh - 60px); }

  /* Sidebar */
  .sidebar { width:260px; min-width:220px; background:var(--card); border-right:1px solid var(--border); padding:16px; overflow-y:auto; flex-shrink:0; }
  .sidebar h3 { font-size:11px; text-transform:uppercase; letter-spacing:.8px; color:var(--muted); margin-bottom:8px; margin-top:16px; }
  .sidebar h3:first-child { margin-top:0; }
  .date-row { display:flex; flex-direction:column; gap:4px; margin-bottom:6px; }
  .date-row label { font-size:12px; color:var(--muted); }
  .date-row input[type=date] { background:#0d1117; border:1px solid var(--border); color:var(--text); padding:5px 8px; border-radius:6px; width:100%; font-size:13px; }
  .date-row input[type=date]:focus { outline:none; border-color:var(--accent); }
  .chips { display:flex; flex-wrap:wrap; gap:5px; margin-bottom:8px; }
  .chip { padding:3px 10px; border-radius:20px; border:1.5px solid; cursor:pointer; font-size:12px; font-weight:500; transition:.15s; user-select:none; }
  .chip.off { opacity:.35; }
  .agg-row { display:flex; gap:4px; margin-bottom:4px; }
  .agg-btn { flex:1; padding:5px; background:#0d1117; border:1px solid var(--border); color:var(--muted); border-radius:6px; cursor:pointer; font-size:12px; transition:.15s; }
  .agg-btn.active { background:var(--accent); border-color:var(--accent); color:#fff; }
  .btn-apply { width:100%; margin-top:16px; padding:9px; background:var(--accent); border:none; color:#fff; border-radius:8px; cursor:pointer; font-weight:600; font-size:13px; transition:.15s; }
  .btn-apply:hover { opacity:.85; }
  .btn-reset { width:100%; margin-top:6px; padding:7px; background:transparent; border:1px solid var(--border); color:var(--muted); border-radius:8px; cursor:pointer; font-size:12px; }

  /* Main */
  .main { flex:1; overflow-y:auto; padding:16px; display:flex; flex-direction:column; gap:16px; }

  /* Charts */
  .chart-card { background:var(--card); border:1px solid var(--border); border-radius:12px; padding:16px; }
  .chart-header { display:flex; justify-content:space-between; align-items:center; margin-bottom:12px; }
  .chart-header h2 { font-size:15px; font-weight:600; }
  .chart-header .unit-badge { background:#0d1117; border:1px solid var(--border); border-radius:6px; padding:2px 8px; font-size:11px; color:var(--muted); }
  .chart-wrap { position:relative; height:260px; }
  .chart-wrap.tall { height:340px; }
  .chart-hint { font-size:11px; color:var(--muted); text-align:right; margin-top:6px; }
  .no-data { color:var(--muted); font-size:13px; padding:20px; text-align:center; }

  /* Presets */
  .preset-row { display:flex; gap:4px; margin-bottom:6px; flex-wrap:wrap; }
  .preset-btn { flex:1; min-width:42px; padding:4px; background:#0d1117; border:1px solid var(--border); color:var(--muted); border-radius:6px; cursor:pointer; font-size:11px; }
  .preset-btn:hover { color:var(--accent); border-color:var(--accent); }
  .preset-btn.active-range { color:var(--accent); border-color:var(--accent); background:#1c2a3a; font-weight:600; }

  /* Tabs */
  .tabs { display:flex; gap:4px; border-bottom:1px solid var(--border); margin-bottom:4px; }
  .tab { padding:8px 14px; background:transparent; border:none; color:var(--muted); cursor:pointer; font-size:13px; font-weight:500; border-bottom:2px solid transparent; transition:.15s; }
  .tab:hover { color:var(--text); }
  .tab.active { color:var(--accent); border-bottom-color:var(--accent); }
  .view { display:flex; flex-direction:column; gap:16px; }
  .view.hidden { display:none; }

  /* Records / Tables */
  .records-grid { display:grid; grid-template-columns:repeat(auto-fit, minmax(360px, 1fr)); gap:12px; }
  .rec-card { background:var(--card); border:1px solid var(--border); border-radius:12px; padding:14px; }
  .rec-card h3 { font-size:14px; font-weight:600; margin-bottom:10px; color:var(--accent); display:flex; justify-content:space-between; align-items:center; }
  .rec-card h3 .badge { font-size:11px; color:var(--muted); font-weight:400; }
  .rec-table { width:100%; border-collapse:collapse; font-size:12px; }
  .rec-table th { text-align:left; color:var(--muted); font-weight:500; padding:4px 6px; border-bottom:1px solid var(--border); font-size:11px; text-transform:uppercase; }
  .rec-table td { padding:5px 6px; border-bottom:1px solid #1f242c; }
  .rec-table td.val { font-weight:600; color:var(--accent); text-align:right; white-space:nowrap; }
  .rec-table td.max { color:var(--red); }
  .rec-table td.min { color:var(--accent); }
  .rec-table tr:hover { background:#1a1f27; }

  /* Trend panel */
  .trend-controls { display:flex; flex-wrap:wrap; gap:8px; align-items:center; background:var(--card); border:1px solid var(--border); border-radius:10px; padding:10px 14px; }
  .trend-controls label { color:var(--muted); font-size:12px; }
  .trend-controls select { background:#0d1117; border:1px solid var(--border); color:var(--text); padding:4px 8px; border-radius:6px; font-size:12px; }
  .trend-info { display:flex; flex-wrap:wrap; gap:8px; margin-top:10px; }
  .trend-info .pill { background:#0d1117; border:1px solid var(--border); border-radius:14px; padding:4px 10px; font-size:11px; color:var(--muted); }
  .trend-info .pill b { color:var(--text); font-weight:600; }
  .trend-info .pill.up b { color:var(--red); }
  .trend-info .pill.down b { color:var(--green); }

  /* Year chips for YoY selector */
  .year-chips { display:flex; flex-wrap:wrap; gap:6px; align-items:center; margin:8px 0 4px; }
  .year-chips .label { color:var(--muted); font-size:12px; margin-right:4px; }
  .year-chip { background:#0d1117; border:1px solid var(--border); color:var(--muted); padding:3px 10px; border-radius:14px; cursor:pointer; font-size:11px; font-weight:600; transition:.15s; user-select:none; display:inline-flex; align-items:center; gap:6px; }
  .year-chip:hover { color:var(--text); }
  .year-chip.active { background:var(--accent); border-color:var(--accent); color:#fff; }
  .year-chip .swatch { width:10px; height:10px; border-radius:50%; display:inline-block; border:1px solid rgba(255,255,255,.25); }
  .year-chips .link-btn { background:transparent; border:none; color:var(--accent); cursor:pointer; font-size:11px; padding:3px 6px; }
  .year-chips .link-btn:hover { text-decoration:underline; }

  /* Heatmap */
  .heatmap-card { background:var(--card); border:1px solid var(--border); border-radius:12px; padding:14px; overflow-x:auto; }
  .heatmap { border-collapse:collapse; font-size:11px; margin-top:10px; }
  .heatmap th, .heatmap td { padding:4px 6px; text-align:center; border:1px solid var(--bg); min-width:42px; }
  .heatmap th { color:var(--muted); font-weight:500; background:transparent; }
  .heatmap td.year-label { color:var(--muted); font-weight:600; background:transparent; min-width:50px; }
  .heatmap td.cell { color:#fff; font-weight:500; cursor:default; }
  .heatmap td.cell.empty { background:#0d1117; color:var(--muted); }
  .heatmap-legend { display:flex; align-items:center; gap:8px; margin-top:8px; font-size:11px; color:var(--muted); }
  .legend-bar { width:160px; height:10px; border-radius:3px; background:linear-gradient(to right, #1e3a8a, #58a6ff, #f0e442, #f59e0b, #dc2626); }
  .legend-bar.diverging { background:linear-gradient(to right, #1e3a8a, #58a6ff, #f0f6fc, #f59e0b, #b91c1c); }
  /* Season heatmap highlights */
  .heatmap td.cell.hot-record  { box-shadow: inset 0 0 0 2px #ffd700; font-weight:700; }
  .heatmap td.cell.cold-record { box-shadow: inset 0 0 0 2px #00e0ff; font-weight:700; }
  .heatmap tr.year-warmest td.year-label::after { content:' 🔥'; }
  .heatmap tr.year-coldest td.year-label::after { content:' 🥶'; }
  .heatmap tr.year-warmest td.year-label { color:#f85149; }
  .heatmap tr.year-coldest td.year-label { color:#58a6ff; }

  /* Threshold slider */
  .thr-slider-row { display:flex; align-items:center; gap:12px; margin:10px 0 6px; flex-wrap:wrap; }
  .thr-slider-row label { color:var(--muted); font-size:12px; min-width:120px; }
  .thr-slider-row input[type=range] { flex:1; min-width:200px; accent-color:var(--accent); height:6px; cursor:pointer; }
  .thr-slider-row .thr-val { background:var(--accent); color:#fff; padding:3px 10px; border-radius:14px; font-weight:600; font-size:12px; min-width:60px; text-align:center; }
  .thr-slider-row .thr-info { color:var(--muted); font-size:11px; }
  .thr-slider-row.compare-row label { display:flex; align-items:center; gap:6px; min-width:120px; }
  .thr-slider-row.compare-row input[type=checkbox] { accent-color:var(--accent); }
  .thr-slider-row .thr-val.compare { background:#30363d; color:var(--text); }
  .insight-pills { display:flex; flex-wrap:wrap; gap:8px; margin:8px 0 10px; }
  .insight-pill { background:#0d1117; border:1px solid var(--border); border-radius:14px; padding:4px 10px; font-size:11px; color:var(--muted); }
  .insight-pill b { color:var(--text); }
  .insight-pill.hot b { color:var(--red); }
  .insight-pill.trop b { color:#bc8cff; }

  /* Rekord-Rangliste */
  .rank-stats { display:grid; grid-template-columns:repeat(auto-fit, minmax(130px,1fr)); gap:10px; margin-bottom:12px; }
  .rank-stat { background:#0d1117; border:1px solid var(--border); border-radius:10px; padding:10px 14px; }
  .rank-stat .label { font-size:11px; color:var(--muted); margin-bottom:4px; }
  .rank-stat .value { font-size:22px; font-weight:700; color:var(--text); }
  .rank-stat .value .unit { font-size:13px; font-weight:500; color:var(--muted); margin-left:2px; }
  .rank-legend { display:flex; flex-wrap:wrap; gap:16px; margin-bottom:10px; font-size:12px; color:var(--muted); }
  .rank-legend .swatch { width:10px; height:10px; border-radius:3px; display:inline-block; margin-right:5px; vertical-align:middle; }
  .rank-year-badges { display:flex; flex-wrap:wrap; gap:6px; margin-top:12px; }
  .rank-year-badge { background:#0d1117; border:1px solid var(--border); border-radius:8px; padding:6px 10px; font-size:11px; color:var(--muted); text-align:center; min-width:44px; }
  .rank-year-badge .y { font-weight:600; color:var(--text); display:block; }
  .rank-year-badge.current { border-color:var(--red); background:#2a1512; }
  .rank-year-badge.current .y { color:var(--red); }

  /* Innenraum-Passwortsperre */
  .lock-panel { background:#0d1117; border:1px solid var(--border); border-radius:8px; padding:10px; margin:8px 0 12px; }
  .lock-row { display:flex; gap:6px; }
  .lock-row input[type=password] { flex:1; min-width:0; background:var(--bg); border:1px solid var(--border); color:var(--text); border-radius:6px; padding:6px 8px; font-size:12px; }
  .lock-btn { background:var(--accent); border:none; color:#fff; border-radius:6px; padding:6px 10px; font-size:12px; cursor:pointer; white-space:nowrap; }
  .lock-btn:hover { opacity:.85; }
  .lock-msg { font-size:11px; color:var(--muted); margin-top:6px; }
  .lock-msg.ok  { color:var(--green); }
  .lock-msg.err { color:var(--red); }

  /* ── Indoor Komfort-Cockpit ── */
  .cockpit-grid { display:grid; grid-template-columns:repeat(auto-fill,minmax(155px,1fr)); gap:12px; margin-bottom:16px; }
  .cockpit-card { background:var(--card); border:2px solid var(--border); border-radius:12px; padding:14px 12px; text-align:center; position:relative; overflow:hidden; transition:.2s; }
  .cockpit-card::before { content:''; position:absolute; top:0; left:0; right:0; height:4px; }
  .cockpit-card.ck-ok::before  { background:var(--green); }
  .cockpit-card.ck-warn::before { background:var(--orange); }
  .cockpit-card.ck-bad::before  { background:var(--red); }
  .cockpit-card.ck-na::before   { background:var(--muted); }
  .cockpit-card.ck-ok  { border-color:#3fb95040; }
  .cockpit-card.ck-warn { border-color:#d2992240; }
  .cockpit-card.ck-bad  { border-color:#f8514940; }
  .ck-icon  { font-size:28px; line-height:1.2; margin-bottom:5px; }
  .ck-label { font-size:10px; color:var(--muted); text-transform:uppercase; letter-spacing:.7px; margin-bottom:6px; }
  .ck-val   { font-size:21px; font-weight:700; margin-bottom:7px; }
  .cockpit-card.ck-ok  .ck-val { color:var(--green); }
  .cockpit-card.ck-warn .ck-val { color:var(--orange); }
  .cockpit-card.ck-bad  .ck-val { color:var(--red); }
  .cockpit-card.ck-na  .ck-val { color:var(--muted); }
  .ck-bar-bg   { background:#0d1117; border-radius:4px; height:5px; margin-bottom:6px; overflow:hidden; }
  .ck-bar-fill { height:100%; border-radius:4px; transition:width .6s ease; }
  .cockpit-card.ck-ok  .ck-bar-fill { background:var(--green); }
  .cockpit-card.ck-warn .ck-bar-fill { background:var(--orange); }
  .cockpit-card.ck-bad  .ck-bar-fill { background:var(--red); }
  .ck-status { font-size:11px; font-weight:600; }
  .cockpit-card.ck-ok  .ck-status { color:var(--green); }
  .cockpit-card.ck-warn .ck-status { color:var(--orange); }
  .cockpit-card.ck-bad  .ck-status { color:var(--red); }
  .cockpit-card.ck-na  .ck-status { color:var(--muted); }
  .comfort-info { display:flex; flex-wrap:wrap; gap:8px; margin-bottom:16px; }
  .comfort-pill { background:#0d1117; border:1px solid var(--border); border-radius:14px; padding:5px 12px; font-size:12px; color:var(--muted); }
  .comfort-pill.good { border-color:var(--green); color:var(--green); }
  .comfort-pill.warn { border-color:var(--orange); color:var(--orange); }
  .comfort-pill.bad  { border-color:var(--red); color:var(--red); }
  .comfort-pill b { color:var(--text); }
</style>
</head>
<body>
<header>
  <div>
    <h1>🌡 Netatmo Daten-Viewer</h1>
    <div class="sub" id="fileInfo">Lade…</div>
  </div>
</header>
<div class="layout">
  <div class="sidebar">
    <h3>Zeitraum</h3>
    <div class="preset-row">
      <button class="preset-btn" onclick="setPreset(7)">7T</button>
      <button class="preset-btn" onclick="setPreset(30)">30T</button>
      <button class="preset-btn" onclick="setPreset(90)">90T</button>
      <button class="preset-btn" onclick="setPreset(365)">1J</button>
      <button class="preset-btn" onclick="setPreset(1825)">5J</button>
      <button class="preset-btn" onclick="setPreset(0)">Alles</button>
    </div>
    <div class="date-row"><label>Von</label><input type="date" id="dateFrom"></div>
    <div class="date-row"><label>Bis</label><input type="date" id="dateTo"></div>

    <h3>Module</h3>
    <div class="chips" id="moduleChips"></div>
    <div class="lock-panel" id="indoorLockPanel" style="display:none;">
      <div class="lock-row">
        <input type="password" id="indoorPwInput" placeholder="Passwort Innenraum" onkeydown="if(event.key==='Enter')unlockIndoor()">
        <button class="lock-btn" onclick="unlockIndoor()">🔒 Entsperren</button>
      </div>
      <div class="lock-msg" id="indoorLockMsg">Innenraum-Module (Schlafzimmer/Wohnzimmer/Zimmer) sind passwortgeschützt.</div>
    </div>

    <h3>Sensoren</h3>
    <div class="chips" id="sensorChips"></div>

    <h3>Aggregation</h3>
    <div class="agg-row">
      <button class="agg-btn active" data-mode="raw"    onclick="setAgg(this)">Roh</button>
      <button class="agg-btn" data-mode="hourly" onclick="setAgg(this)">Stündl.</button>
      <button class="agg-btn" data-mode="daily"  onclick="setAgg(this)">Tägl.</button>
    </div>
    <div class="agg-row">
      <button class="agg-btn" data-mode3="weekly"  onclick="setAgg(this)">Wöch.</button>
      <button class="agg-btn" data-mode="monthly" onclick="setAgg(this)">Monatl.</button>
      <button class="agg-btn" data-mode="yearly"  onclick="setAgg(this)">Jährl.</button>
    </div>

    <button class="btn-apply" onclick="applyFilters()">Aktualisieren</button>
    <button class="btn-reset" onclick="resetFilters()">Zurücksetzen</button>
  </div>

  <div class="main">
    <div class="tabs">
      <button class="tab active" data-view="charts"   onclick="switchView(this)">📈 Charts</button>
      <button class="tab"        data-view="records"  onclick="switchView(this)">🏆 Rekorde &amp; Tops</button>
      <button class="tab"        data-view="trends"   onclick="switchView(this)">📊 Trends</button>
      <button class="tab"        data-view="longterm" onclick="switchView(this)">📅 Langzeit</button>
      <button class="tab"        data-view="compare"  onclick="switchView(this)">🔄 Vergleich</button>
      <button class="tab"        data-view="climate"  onclick="switchView(this)">🌡️ Klimawandel</button>
      <button class="tab"        data-view="extremes" onclick="switchView(this)">⚡ Extreme</button>
      <button class="tab"        data-view="indoor"   onclick="switchView(this)">🏠 Innenraum</button>
    </div>
    <div id="view-charts"   class="view"><div id="chartsContainer"></div></div>
    <div id="view-records"  class="view hidden"><div id="recordsContainer"></div></div>
    <div id="view-trends"   class="view hidden"><div id="trendsContainer"></div></div>
    <div id="view-longterm" class="view hidden"><div id="longtermContainer"></div></div>
    <div id="view-compare"  class="view hidden"><div id="compareContainer"></div></div>
    <div id="view-climate"  class="view hidden"><div id="climateContainer"></div></div>
    <div id="view-extremes" class="view hidden"><div id="extremesContainer"></div></div>
    <div id="view-indoor"   class="view hidden"><div id="indoorContainer"></div></div>
  </div>
</div>

<script>
const PAYLOAD = __PAYLOAD__;

let aggMode    = 'raw';
let selModules = new Set();
let selSensors = new Set();
let chartInstances = {};
let favTropThreshold = 20;
let favTropCompareThreshold = 18;
let favTropCompareEnabled = true;
let favHotThreshold = 30;
let favHotCompareThreshold = 25;
let favHotCompareEnabled = true;

/* ── Helpers ── */
// Math.max/min(...array) blows the call-stack for very large arrays (raw multi-year data).
function arrMax(arr) { let m = -Infinity; for (let i = 0; i < arr.length; i++) if (arr[i] > m) m = arr[i]; return m; }
function arrMin(arr) { let m = Infinity; for (let i = 0; i < arr.length; i++) if (arr[i] < m) m = arr[i]; return m; }
function fmt(n, unit) {
  if (n === undefined || n === null) return '–';
  return n.toLocaleString('de-DE', {maximumFractionDigits:1}) + (unit ? ' ' + unit : '');
}
function fmtDate(ts) {
  return new Date(ts).toLocaleDateString('de-DE', {day:'2-digit',month:'2-digit',year:'numeric'});
}

/* ── Bucket helpers ── */
function bucketDate(dt, mode) {
  const b = new Date(dt);
  if (mode === 'hourly')  { b.setMinutes(0,0,0); }
  else if (mode === 'daily')   { b.setHours(0,0,0,0); }
  else if (mode === 'weekly')  {
    b.setHours(0,0,0,0);
    // Monday-based week start
    const dow = (b.getDay() + 6) % 7;
    b.setDate(b.getDate() - dow);
  }
  else if (mode === 'monthly') { b.setHours(0,0,0,0); b.setDate(1); }
  else if (mode === 'yearly')  { b.setHours(0,0,0,0); b.setMonth(0,1); }
  return b;
}
function bucketKey(d, mode) {
  const dt = new Date(d.ts);
  switch (mode) {
    case 'hourly':  return `${dt.getFullYear()}-${dt.getMonth()}-${dt.getDate()}-${dt.getHours()}`;
    case 'daily':   return `${dt.getFullYear()}-${dt.getMonth()}-${dt.getDate()}`;
    case 'weekly':  { const b=bucketDate(dt,'weekly'); return `${b.getFullYear()}-${b.getMonth()}-${b.getDate()}`; }
    case 'monthly': return `${dt.getFullYear()}-${dt.getMonth()}`;
    case 'yearly':  return `${dt.getFullYear()}`;
    default: return '';
  }
}

/* ── Aggregate (mean) ── */
function aggregateData(data, mode) {
  if (mode === 'raw') return data;
  const groups = {};
  data.forEach(d => {
    const key = `${d.module}|${d.sensor}|` + bucketKey(d, mode);
    if (!groups[key]) {
      const bucket = bucketDate(new Date(d.ts), mode);
      groups[key] = { ts: bucket.getTime(), module: d.module, sensor: d.sensor, unit: d.unit, vals: [] };
    }
    groups[key].vals.push(d.value);
  });
  return Object.values(groups).map(g => ({
    ts:     g.ts,
    module: g.module,
    sensor: g.sensor,
    unit:   g.unit,
    value:  Math.round(g.vals.reduce((a,b)=>a+b,0)/g.vals.length * 100) / 100,
  })).sort((a,b)=>a.ts-b.ts);
}

/* Aggregate to min/avg/max per bucket */
function aggregateMinAvgMax(data, mode) {
  const groups = {};
  data.forEach(d => {
    const key = `${d.module}|${d.sensor}|` + bucketKey(d, mode);
    if (!groups[key]) {
      const bucket = bucketDate(new Date(d.ts), mode);
      groups[key] = { ts: bucket.getTime(), module: d.module, sensor: d.sensor, unit: d.unit, vals: [] };
    }
    groups[key].vals.push(d.value);
  });
  return Object.values(groups).map(g => {
    const vs = g.vals;
    const sum = vs.reduce((a,b)=>a+b,0);
    return {
      ts: g.ts, module: g.module, sensor: g.sensor, unit: g.unit,
      min: arrMin(vs), max: arrMax(vs),
      avg: Math.round(sum/vs.length*100)/100, n: vs.length
    };
  }).sort((a,b)=>a.ts-b.ts);
}

/* ── Build flat list from compact payload ── */
function buildFlat() {
  const flat = [];
  for (const sensor of PAYLOAD.sensors) {
    const unit = PAYLOAD.units[sensor] || '';
    if (!PAYLOAD.data[sensor]) continue;
    for (const module of PAYLOAD.modules) {
      const pts = PAYLOAD.data[sensor][module];
      if (!pts) continue;
      pts.forEach(([ts, value]) => flat.push({ ts, sensor, module, unit, value }));
    }
  }
  return flat;
}
let FLAT = buildFlat();

/* ── Innenraum-Entschlüsselung (Passwort → PBKDF2 → HMAC-Keystream) ── */
function b64ToBytes(b64) { return Uint8Array.from(atob(b64), c => c.charCodeAt(0)); }

function bytesEqual(a, b) {
  if (a.length !== b.length) return false;
  let diff = 0;
  for (let i = 0; i < a.length; i++) diff |= a[i] ^ b[i];
  return diff === 0;
}

async function deriveLockKey(password, saltB64, iterations) {
  const enc = new TextEncoder();
  const salt = b64ToBytes(saltB64);
  const baseKey = await crypto.subtle.importKey('raw', enc.encode(password), 'PBKDF2', false, ['deriveBits']);
  const bits = await crypto.subtle.deriveBits({ name: 'PBKDF2', salt, iterations, hash: 'SHA-256' }, baseKey, 256);
  return new Uint8Array(bits);
}

async function hmacRaw(keyBytes, msgBytes) {
  const key = await crypto.subtle.importKey('raw', keyBytes, { name: 'HMAC', hash: 'SHA-256' }, false, ['sign']);
  return new Uint8Array(await crypto.subtle.sign('HMAC', key, msgBytes));
}

async function hmacKeystream(keyBytes, nonceBytes, length) {
  const out = new Uint8Array(length);
  let offset = 0, counter = 0;
  while (offset < length) {
    const counterBytes = new Uint8Array(4);
    new DataView(counterBytes.buffer).setUint32(0, counter, false);
    const input = new Uint8Array(nonceBytes.length + 4);
    input.set(nonceBytes, 0);
    input.set(counterBytes, nonceBytes.length);
    const block = await hmacRaw(keyBytes, input);
    out.set(block.subarray(0, Math.min(block.length, length - offset)), offset);
    offset += block.length;
    counter += 1;
  }
  return out;
}

async function unlockIndoor() {
  const enc = PAYLOAD.locked_enc;
  const msgEl = document.getElementById('indoorLockMsg');
  const pwInput = document.getElementById('indoorPwInput');
  const pw = pwInput.value;
  if (!enc || !pw) return;
  msgEl.className = 'lock-msg';
  msgEl.textContent = 'Prüfe Passwort…';
  try {
    const key = await deriveLockKey(pw, enc.salt, enc.iterations);
    const nonce = b64ToBytes(enc.nonce);
    const ciphertext = b64ToBytes(enc.ciphertext);
    const macMsg = new Uint8Array(nonce.length + ciphertext.length);
    macMsg.set(nonce, 0);
    macMsg.set(ciphertext, nonce.length);
    const computedMac = await hmacRaw(key, macMsg);
    if (!bytesEqual(computedMac, b64ToBytes(enc.mac))) {
      msgEl.className = 'lock-msg err';
      msgEl.textContent = '❌ Falsches Passwort.';
      return;
    }
    const ks = await hmacKeystream(key, nonce, ciphertext.length);
    const plain = new Uint8Array(ciphertext.length);
    for (let i = 0; i < ciphertext.length; i++) plain[i] = ciphertext[i] ^ ks[i];
    const obj = JSON.parse(new TextDecoder().decode(plain));
    Object.keys(obj).forEach(sensor => {
      PAYLOAD.data[sensor] = PAYLOAD.data[sensor] || {};
      Object.assign(PAYLOAD.data[sensor], obj[sensor]);
    });
    FLAT = buildFlat();
    pwInput.value = '';
    msgEl.className = 'lock-msg ok';
    msgEl.textContent = '✅ Innenraum-Daten entsperrt.';
    applyFilters();
  } catch (e) {
    msgEl.className = 'lock-msg err';
    msgEl.textContent = '❌ Fehler beim Entschlüsseln: ' + e.message;
  }
}

/* ── Get date bounds from data ── */
function getDateBounds() {
  let mn = Infinity, mx = -Infinity;
  FLAT.forEach(d => { if(d.ts<mn) mn=d.ts; if(d.ts>mx) mx=d.ts; });
  return { min: new Date(mn).toISOString().slice(0,10), max: new Date(mx).toISOString().slice(0,10) };
}

function getDefaultSelection() {
  const outdoorRe = /outdoor|outside|außen|aussen|draußen|draussen|garten|balkon|terrasse|external|extern/i;
  const tempRe = /temperature|temperatur|temp/i;
  const temperatureSensors = PAYLOAD.sensors.filter(s => tempRe.test(s) || /°?\s*C/i.test(PAYLOAD.units[s] || ''));
  const hasPoints = (module, sensor) => !!(PAYLOAD.data[sensor]?.[module]?.length);
  const sensors = temperatureSensors.length ? temperatureSensors : PAYLOAD.sensors;
  const moduleCandidates = PAYLOAD.modules.filter(m => sensors.some(s => hasPoints(m, s)));
  const outdoorModules = PAYLOAD.modules.filter(m => outdoorRe.test(m) && sensors.some(s => hasPoints(m, s)));
  const modules = outdoorModules.length ? outdoorModules : moduleCandidates;
  const selectedSensors = sensors.filter(s => modules.some(m => hasPoints(m, s)));
  const selectedModules = modules.filter(m => selectedSensors.some(s => hasPoints(m, s)));

  return {
    modules: new Set(selectedModules.length ? selectedModules : PAYLOAD.modules.slice(0, 1)),
    sensors: new Set(selectedSensors.length ? selectedSensors : PAYLOAD.sensors.slice(0, 1)),
  };
}

function applyDefaultSelection() {
  const defaults = getDefaultSelection();
  selModules = defaults.modules;
  selSensors = defaults.sensors;
  setAggMode('raw');
}

/* ── Chips ── */
function buildChips(containerId, items, selSet, colorMap) {
  const el = document.getElementById(containerId);
  el.innerHTML = '';
  items.forEach((item, i) => {
    const color = colorMap ? colorMap[item] : PAYLOAD.colors[item] || '#58a6ff';
    const chip  = document.createElement('div');
    chip.className = 'chip' + (selSet.has(item) ? '' : ' off');
    chip.style.borderColor = color;
    chip.style.color       = color;
    chip.textContent       = item;
    chip.onclick = () => {
      if (selSet.has(item)) { selSet.delete(item); chip.classList.add('off'); }
      else                  { selSet.add(item);    chip.classList.remove('off'); }
      applyFilters();
    };
    el.appendChild(chip);
  });
}

function setAggMode(mode) {
  aggMode = mode;
  document.querySelectorAll('.agg-btn').forEach(b => b.classList.toggle('active', b.dataset.mode === mode));
}

function setAgg(btn) {
  if (!btn.dataset.mode) return;
  setAggMode(btn.dataset.mode);
}

function isTemperatureSensor(sensor, unit) {
  return /°?\s*C/i.test(unit || PAYLOAD.units[sensor] || '') || /temp|temperatur|temperature/i.test(sensor);
}

function renderFavoriteThresholdCard(container, days, years, cfg) {
  const yearIndex = Object.fromEntries(years.map((year, index) => [year, index]));
  const countFor = threshold => {
    const counts = years.map(() => 0);
    days.forEach(day => {
      if (cfg.match(day, threshold)) counts[yearIndex[day.year]]++;
    });
    return counts;
  };
  const countBetween = (a, b) => {
    const lo = Math.min(a, b), hi = Math.max(a, b);
    const counts = years.map(() => 0);
    if (lo === hi) return counts;
    days.forEach(day => {
      const value = cfg.value(day);
      if (value > lo && value <= hi) counts[yearIndex[day.year]]++;
    });
    return counts;
  };
  const topOf = counts => counts.reduce(
    (best, value, index) => value > best.value ? { year: years[index], value } : best,
    { year: years[0], value: counts[0] || 0 });
  const fmtT = value => value.toLocaleString('de-DE', {maximumFractionDigits:1});

  const card = document.createElement('div');
  card.className = 'chart-card';
  card.innerHTML = `
    <div class="chart-header"><h2>${cfg.title}</h2><span class="unit-badge">Tage / Jahr</span></div>
    <div class="thr-slider-row">
      <label for="${cfg.sliderId}">${cfg.sliderLabel}</label>
      <input type="range" id="${cfg.sliderId}" min="${cfg.min}" max="${cfg.max}" step="${cfg.step}" value="${cfg.threshold}">
      <span class="thr-val" id="${cfg.valueId}">${cfg.formatLabel(cfg.threshold)}</span>
    </div>
    <div class="thr-slider-row compare-row">
      <label for="${cfg.compareToggleId}"><input type="checkbox" id="${cfg.compareToggleId}" ${cfg.compareEnabled ? 'checked' : ''}> Vergleich</label>
      <input type="range" id="${cfg.compareSliderId}" min="${cfg.min}" max="${cfg.max}" step="${cfg.step}" value="${cfg.compareThreshold}">
      <span class="thr-val compare" id="${cfg.compareValueId}">${cfg.formatLabel(cfg.compareThreshold)}</span>
    </div>
    <div class="insight-pills" id="${cfg.insightId}"></div>
    <div class="chart-wrap tall"><canvas id="${cfg.chartId}"></canvas></div>
    <div class="chart-hint">${cfg.hint}</div>`;
  container.appendChild(card);

  const slider = document.getElementById(cfg.sliderId);
  const compareToggle = document.getElementById(cfg.compareToggleId);
  const compareSlider = document.getElementById(cfg.compareSliderId);
  const valueEl = document.getElementById(cfg.valueId);
  const compareValueEl = document.getElementById(cfg.compareValueId);
  const insightEl = document.getElementById(cfg.insightId);

  const chart = new Chart(document.getElementById(cfg.chartId).getContext('2d'), {
    type: 'bar',
    data: {
      labels: years,
      datasets: [
        { label: cfg.formatLabel(cfg.threshold), data: countFor(cfg.threshold), backgroundColor: cfg.color + 'aa', borderColor: cfg.color, borderWidth: 1 },
        { type: 'line', label: cfg.formatLabel(cfg.compareThreshold), data: countFor(cfg.compareThreshold), borderColor: cfg.compareColor, backgroundColor: cfg.compareColor + '22', borderWidth: 2, borderDash: [6, 4], pointRadius: 3, tension: 0.25, hidden: !cfg.compareEnabled },
      ],
    },
    options: {
      responsive: true, maintainAspectRatio: false, interaction: { mode: 'index', intersect: false },
      scales: {
        x: { grid: { color: '#30363d' }, ticks: { color: '#8b949e' } },
        y: { grid: { color: '#30363d' }, ticks: { color: '#8b949e' }, beginAtZero: true, title: { display: true, text: 'Tage / Jahr', color: '#8b949e' } },
      },
      plugins: {
        legend: { labels: { color: '#e6edf3', boxWidth: 12, padding: 10 } },
        tooltip: { backgroundColor: '#161b22', borderColor: '#30363d', borderWidth: 1, titleColor: '#e6edf3', bodyColor: '#8b949e', callbacks: { label: ctx => ` ${ctx.dataset.label}: ${ctx.parsed.y} Tage` } },
      },
    },
  });
  chartInstances[cfg.chartKey] = chart;

  const updateInsights = () => {
    const threshold = +slider.value;
    const compareThreshold = +compareSlider.value;
    const top = topOf(countFor(threshold));
    const parts = [`<span class="insight-pill ${cfg.kind}"><b>${top.year}</b> stärkstes Jahr: ${top.value} Tage</span>`];
    if (compareToggle.checked) {
      const betweenTop = topOf(countBetween(threshold, compareThreshold));
      const lo = fmtT(Math.min(threshold, compareThreshold));
      const hi = fmtT(Math.max(threshold, compareThreshold));
      parts.push(`<span class="insight-pill"><b>${betweenTop.year}</b> zwischen ${lo} und ${hi} °C: ${betweenTop.value} Tage</span>`);
    }
    insightEl.innerHTML = parts.join('');
  };

  const updateChart = () => {
    const threshold = +slider.value;
    const compareThreshold = +compareSlider.value;
    cfg.setThreshold(threshold);
    cfg.setCompareThreshold(compareThreshold);
    cfg.setCompareEnabled(compareToggle.checked);
    valueEl.textContent = cfg.formatLabel(threshold);
    compareValueEl.textContent = cfg.formatLabel(compareThreshold);
    chart.data.datasets[0].label = cfg.formatLabel(threshold);
    chart.data.datasets[0].data = countFor(threshold);
    chart.data.datasets[1].label = cfg.formatLabel(compareThreshold);
    chart.data.datasets[1].data = countFor(compareThreshold);
    chart.data.datasets[1].hidden = !compareToggle.checked;
    updateInsights();
    chart.update('none');
  };

  slider.addEventListener('input', updateChart);
  compareSlider.addEventListener('input', updateChart);
  compareToggle.addEventListener('change', updateChart);
  updateInsights();
}

/* ── Rekord-Rangliste (Top-N Tage, Balken nach Jahr eingefärbt) ── */
function renderRecordRankingCard(container, daily, cfg) {
  const yearOf = r => new Date(r.ts).getFullYear();
  const ranked = daily
    .map(r => ({ ts: r.ts, module: r.module, v: cfg.pickValue(r) }))
    .sort((a, b) => cfg.sortDesc ? b.v - a.v : a.v - b.v)
    .slice(0, cfg.topN);
  if (!ranked.length) return;

  const currentYear = Math.max(...daily.map(yearOf));
  const top10 = ranked.slice(0, 10);
  const placesCurrent = ranked.filter(r => yearOf(r) === currentYear).length;
  const top10Current = top10.filter(r => yearOf(r) === currentYear).length;
  const yearCounts = {};
  ranked.forEach(r => { const y = yearOf(r); yearCounts[y] = (yearCounts[y] || 0) + 1; });
  const sortedYears = Object.keys(yearCounts).map(Number).sort((a, b) => a - b);
  const otherYears = sortedYears.filter(y => y !== currentYear);
  const fmtNum = n => n.toLocaleString('de-DE', {maximumFractionDigits:1});

  const card = document.createElement('div');
  card.className = 'chart-card';
  card.innerHTML = `
    <div class="chart-header"><h2>${cfg.title}</h2><span class="unit-badge">${cfg.unit}</span></div>
    <div class="rank-stats">
      <div class="rank-stat"><div class="label">Plätze von ${currentYear}</div><div class="value">${placesCurrent}<span class="unit">/${ranked.length}</span></div></div>
      <div class="rank-stat"><div class="label">${cfg.recordLabel}</div><div class="value">${fmtNum(ranked[0].v)}<span class="unit">${cfg.unit}</span></div></div>
      <div class="rank-stat"><div class="label">Top 10</div><div class="value">${top10Current}<span class="unit">/10</span></div></div>
      <div class="rank-stat"><div class="label">Jahre in der Liste</div><div class="value">${sortedYears.length}</div></div>
    </div>
    <div class="rank-legend">
      <span><span class="swatch" style="background:${cfg.color}"></span>${currentYear} — ${placesCurrent} Einträge</span>
      ${otherYears.length ? `<span><span class="swatch" style="background:${cfg.otherColor}"></span>${otherYears.join(', ')} — ${ranked.length - placesCurrent} Einträge</span>` : ''}
    </div>
    <div class="chart-wrap tall"><canvas id="${cfg.chartId}"></canvas></div>
    <div class="rank-year-badges">
      ${sortedYears.map(y => `<div class="rank-year-badge${y === currentYear ? ' current' : ''}"><span class="y">${y}</span>${yearCounts[y]}</div>`).join('')}
    </div>
    <div class="chart-hint">${cfg.hint}</div>`;
  container.appendChild(card);

  const labels = ranked.map(r => fmtDate(r.ts));
  const values = ranked.map(r => r.v);
  const colors = ranked.map(r => yearOf(r) === currentYear ? cfg.color : cfg.otherColor);
  const axisMin = Math.floor(Math.min(...values) / 5) * 5;

  const chart = new Chart(document.getElementById(cfg.chartId).getContext('2d'), {
    type: 'bar',
    data: { labels, datasets: [{ data: values, backgroundColor: colors, borderRadius: 3, barThickness: 14 }] },
    options: {
      indexAxis: 'y',
      responsive: true, maintainAspectRatio: false,
      scales: {
        x: { min: axisMin, grid: { color: '#30363d' }, ticks: { color: '#8b949e', callback: v => v + (cfg.unit ? ' ' + cfg.unit : '') } },
        y: { grid: { display: false }, ticks: { color: '#8b949e' } },
      },
      plugins: {
        legend: { display: false },
        tooltip: {
          backgroundColor: '#161b22', borderColor: '#30363d', borderWidth: 1,
          titleColor: '#e6edf3', bodyColor: '#8b949e',
          callbacks: { label: ctx => ` ${fmt(ctx.parsed.x, cfg.unit)}` },
        },
        datalabels: {
          anchor: 'end', align: 'right', color: '#e6edf3', font: { size: 10, weight: '600' },
          formatter: v => fmtNum(v) + ' ' + cfg.unit,
        },
      },
    },
    plugins: [ChartDataLabels],
  });
  chartInstances[cfg.chartKey] = chart;
}

/* ── Tagesextrema-Chart (Min/Max + Tropennächte/Hitzetage) ── */
function renderDailyExtremes(container, raw, sensor, unit, prefix) {
  if (!isTemperatureSensor(sensor, unit) || !raw.length) return;

  const modules = [...selModules].filter(m => raw.some(d => d.module === m));
  if (!modules.length) return;

  const allTs = raw.map(d => d.ts);
  const dataMaxTs = arrMax(allTs);
  const dataMinTs = arrMin(allTs);

  const chartId  = `dexChart_${prefix}`;
  const rangeId  = `dexRange_${prefix}`;
  let extrChart  = null;
  let curDays    = 30;

  // Build daily {ts, module, min, max} from raw, filtered by tsFrom
  function buildDaily(tsFrom) {
    const map = {};
    raw.forEach(d => {
      if (!selModules.has(d.module) || d.ts < tsFrom) return;
      const dt  = new Date(d.ts);
      const key = `${d.module}|${dt.getFullYear()}-${dt.getMonth()}-${dt.getDate()}`;
      const bkt = new Date(dt.getFullYear(), dt.getMonth(), dt.getDate()).getTime();
      if (!map[key]) map[key] = { ts: bkt, module: d.module, min: d.value, max: d.value };
      else { if (d.value < map[key].min) map[key].min = d.value; if (d.value > map[key].max) map[key].max = d.value; }
    });
    return Object.values(map).sort((a, b) => a.ts - b.ts);
  }

  function buildDatasets(days) {
    const tsFrom = days === 0 ? dataMinTs : dataMaxTs - days * 86400000;
    const daily  = buildDaily(tsFrom);
    const ds     = [];
    modules.forEach(mod => {
      const color = PAYLOAD.colors[mod] || '#58a6ff';
      const md    = daily.filter(d => d.module === mod);
      if (!md.length) return;
      const sfx   = modules.length > 1 ? ` (${mod})` : '';

      // Min/Max-Linien ohne Band-Füllung (Füllung übernehmen die Schwellwert-Flächen unten)
      ds.push({ label: `Tmin${sfx}`, data: md.map(d => ({ x: d.ts, y: Math.round(d.min*10)/10 })),
        borderColor: '#58a6ff', backgroundColor: 'transparent', borderWidth: 1.8,
        pointRadius: 0, pointHoverRadius: 4, tension: 0.3, fill: false, order: 4 });
      ds.push({ label: `Tmax${sfx}`, data: md.map(d => ({ x: d.ts, y: Math.round(d.max*10)/10 })),
        borderColor: '#f85149', backgroundColor: 'transparent', borderWidth: 1.8,
        pointRadius: 0, pointHoverRadius: 4, tension: 0.3, fill: false, order: 4 });

      // Rote Fläche: Hitzetage – Bereich zwischen Tmax-Linie und 30°C, nur wo Tmax > 30°C
      ds.push({ label: `Hitzefläche${sfx}`, data: md.map(d => ({ x: d.ts, y: Math.max(Math.round(d.max*10)/10, 30) })),
        borderWidth: 0, pointRadius: 0, backgroundColor: 'rgba(248,81,73,0.45)',
        fill: { value: 30 }, tension: 0.3, order: 10 });

      // Rote Fläche: Tropennächte – Bereich zwischen Tmin-Linie und 20°C, nur wo Tmin > 20°C
      ds.push({ label: `Tropenfläche${sfx}`, data: md.map(d => ({ x: d.ts, y: Math.max(Math.round(d.min*10)/10, 20) })),
        borderWidth: 0, pointRadius: 0, backgroundColor: 'rgba(248,81,73,0.45)',
        fill: { value: 20 }, tension: 0.3, order: 10 });

      // Tropical night markers on min line
      const tropPts = md.filter(d => d.min > 20).map(d => ({ x: d.ts, y: Math.round(d.min*10)/10 }));
      if (tropPts.length) ds.push({
        label: `🌴 Tropennacht Tmin>20°C${sfx}`, data: tropPts, type: 'scatter',
        backgroundColor: '#bc8cff', borderColor: '#fff', borderWidth: 1.2,
        pointRadius: 6, pointHoverRadius: 8, showLine: false, order: 1 });

      // Heat day markers on max line
      const heatPts = md.filter(d => d.max > 30).map(d => ({ x: d.ts, y: Math.round(d.max*10)/10 }));
      if (heatPts.length) ds.push({
        label: `☀️ Hitzetag Tmax>30°C${sfx}`, data: heatPts, type: 'scatter',
        backgroundColor: '#f59e0b', borderColor: '#fff', borderWidth: 1.2,
        pointRadius: 6, pointHoverRadius: 8, showLine: false, order: 1 });
    });
    return ds;
  }

  function buildOrUpdate(days) {
    const datasets = buildDatasets(days);
    if (extrChart) { extrChart.data.datasets = datasets; extrChart.update('none'); return; }
    const ctx = document.getElementById(chartId).getContext('2d');
    extrChart = new Chart(ctx, {
      type: 'line',
      data: { datasets },
      options: {
        responsive: true, maintainAspectRatio: false, parsing: false, normalized: true,
        interaction: { mode: 'index', intersect: false },
        scales: {
          x: { type: 'time',
               time: { tooltipFormat: 'dd.MM.yyyy', displayFormats: { day:'dd.MM', week:'dd.MM.yy', month:'MMM yy' } },
               grid: { color:'#30363d' }, ticks: { color:'#8b949e', maxTicksLimit:10 } },
          y: { grid: { color:'#30363d' },
               ticks: { color:'#8b949e', callback: v => v.toLocaleString('de-DE',{maximumFractionDigits:1}) + (unit?' '+unit:'') } },
        },
        plugins: {
          legend: { labels: { color:'#e6edf3', boxWidth:12, padding:12,
            // hide raw Tmin/Tmax lines + threshold fill helper datasets from legend, keep marker layers
            filter: it => !/^(Tmin|Tmax|Hitzefläche|Tropenfläche)/.test(it.text) } },
          tooltip: {
            backgroundColor:'#161b22', borderColor:'#30363d', borderWidth:1,
            titleColor:'#e6edf3', bodyColor:'#8b949e',
            callbacks: { label: ctx => {
              const lbl = ctx.dataset.label || '';
              if (/^(Hitzefläche|Tropenfläche)/.test(lbl)) return null;
              const v = ctx.parsed.y;
              if (v === undefined || v === null) return null;
              return ` ${lbl}: ${v.toLocaleString('de-DE',{maximumFractionDigits:1})} ${unit}`;
            }},
          },
          zoom: { zoom:{wheel:{enabled:true},pinch:{enabled:true},mode:'x'}, pan:{enabled:true,mode:'x'} },
        },
      },
    });
    document.getElementById(chartId).addEventListener('dblclick', () => extrChart?.resetZoom());
  }

  // Build info pills (counts)
  function buildPills(days) {
    const tsFrom = days === 0 ? dataMinTs : dataMaxTs - days * 86400000;
    const daily  = buildDaily(tsFrom);
    const trop   = daily.filter(d => d.min > 20).length;
    const heat   = daily.filter(d => d.max > 30).length;
    const frost  = daily.filter(d => d.min < 0).length;
    const parts  = [];
    if (heat)  parts.push(`<span class="insight-pill hot"><b>${heat}</b> Hitzetag${heat!==1?'e':''} (Tmax&gt;30°C)</span>`);
    if (trop)  parts.push(`<span class="insight-pill trop"><b>${trop}</b> Tropennacht${trop!==1?'e':''} (Tmin&gt;20°C)</span>`);
    if (frost) parts.push(`<span class="insight-pill"><b>${frost}</b> Frosttag${frost!==1?'e':''} (Tmin&lt;0°C)</span>`);
    const el = document.getElementById(`dexPills_${prefix}`);
    if (el) el.innerHTML = parts.join('') || '<span class="insight-pill">Keine Extremtage im gewählten Zeitraum</span>';
  }

  const card = document.createElement('div');
  card.className = 'chart-card';
  card.innerHTML = `
    <div class="chart-header">
      <h2>🌡️ Tagesextrema – Min / Max</h2>
      <span class="unit-badge">${unit}</span>
    </div>
    <div style="display:flex;align-items:center;gap:8px;margin-bottom:6px;flex-wrap:wrap;">
      <span style="color:var(--muted);font-size:12px;">Zeitraum:</span>
      <div style="display:flex;gap:4px;" id="${rangeId}">
        <button class="preset-btn" data-days="7">7T</button>
        <button class="preset-btn" data-days="14">14T</button>
        <button class="preset-btn active-range" data-days="30">30T</button>
        <button class="preset-btn" data-days="60">60T</button>
        <button class="preset-btn" data-days="90">90T</button>
        <button class="preset-btn" data-days="365">1J</button>
        <button class="preset-btn" data-days="0">Alles</button>
      </div>
      <span style="color:var(--muted);font-size:11px;margin-left:4px;">
        <span style="display:inline-block;width:9px;height:9px;border-radius:50%;background:#f59e0b;margin-right:3px;"></span>Hitzetag &gt;30°C
        &nbsp;
        <span style="display:inline-block;width:9px;height:9px;border-radius:50%;background:#bc8cff;margin-right:3px;"></span>Tropennacht &gt;20°C
        &nbsp;
        <span style="display:inline-block;width:26px;height:3px;background:#f85149;margin-right:3px;vertical-align:middle;"></span>Tmax
        &nbsp;
        <span style="display:inline-block;width:26px;height:3px;background:#58a6ff;margin-right:3px;vertical-align:middle;"></span>Tmin
      </span>
    </div>
    <div class="insight-pills" id="dexPills_${prefix}"></div>
    <div class="chart-wrap tall"><canvas id="${chartId}"></canvas></div>
    <div class="chart-hint">Scrollen zum Zoomen · Ziehen zum Verschieben · Doppelklick zum Zurücksetzen</div>`;
  container.appendChild(card);

  document.getElementById(rangeId).addEventListener('click', e => {
    const btn = e.target.closest('.preset-btn');
    if (!btn) return;
    curDays = +btn.dataset.days;
    document.querySelectorAll(`#${rangeId} .preset-btn`).forEach(b =>
      b.classList.toggle('active-range', b === btn));
    buildOrUpdate(curDays);
    buildPills(curDays);
  });

  buildOrUpdate(curDays);
  buildPills(curDays);
  // Register destroy handle so main chart cleanup works
  chartInstances[`dex_${prefix}`] = { destroy() { if (extrChart) { extrChart.destroy(); extrChart = null; } }, canvas: null };
}

function renderTemperatureFavorites(container, raw, sensor, unit, prefix) {
  if (!isTemperatureSensor(sensor, unit)) return;
  const dayMap = {};
  raw.forEach(d => {
    const dt = new Date(d.ts);
    const key = `${dt.getFullYear()}-${dt.getMonth()}-${dt.getDate()}`;
    if (!dayMap[key]) dayMap[key] = { year: dt.getFullYear(), min: d.value, max: d.value };
    else {
      if (d.value < dayMap[key].min) dayMap[key].min = d.value;
      if (d.value > dayMap[key].max) dayMap[key].max = d.value;
    }
  });
  const days = Object.values(dayMap);
  const years = [...new Set(days.map(d => d.year))].sort((a,b) => a-b);
  if (!days.length || !years.length) return;

  renderFavoriteThresholdCard(container, days, years, {
    kind: 'trop', chartKey: `fav_${prefix}_trop`, title: '🌴 Tropennächte – Nächte mit Tmin über Schwellwert',
    sliderId: `favTropSlider_${prefix}`, valueId: `favTropVal_${prefix}`, compareToggleId: `favTropCompare_${prefix}`,
    compareSliderId: `favTropCompareSlider_${prefix}`, compareValueId: `favTropCompareVal_${prefix}`, insightId: `favTropInsight_${prefix}`, chartId: `favTropChart_${prefix}`,
    sliderLabel: 'Schwellwert (Tmin >):', color: '#bc8cff', compareColor: '#58a6ff', min: 12, max: 25, step: 0.5,
    threshold: favTropThreshold, compareThreshold: favTropCompareThreshold, compareEnabled: favTropCompareEnabled,
    formatLabel: t => `Tmin > ${t.toLocaleString('de-DE', {maximumFractionDigits:1})} °C`,
    match: (d, t) => d.min > t, value: d => d.min,
    setThreshold: t => { favTropThreshold = t; }, setCompareThreshold: t => { favTropCompareThreshold = t; }, setCompareEnabled: v => { favTropCompareEnabled = v; },
    hint: 'Balken = aktiver Grenzwert. Gestrichelte Linie = Vergleichswert. Der Hinweis zeigt, welches Jahr beim aktuellen Wert und im Bereich zwischen beiden Reglern am stärksten heraussticht.',
  });

  renderFavoriteThresholdCard(container, days, years, {
    kind: 'hot', chartKey: `fav_${prefix}_hot`, title: '☀️ Heiße Tage – Tage mit Tmax über Schwellwert',
    sliderId: `favHotSlider_${prefix}`, valueId: `favHotVal_${prefix}`, compareToggleId: `favHotCompare_${prefix}`,
    compareSliderId: `favHotCompareSlider_${prefix}`, compareValueId: `favHotCompareVal_${prefix}`, insightId: `favHotInsight_${prefix}`, chartId: `favHotChart_${prefix}`,
    sliderLabel: 'Schwellwert (Tmax >):', color: '#f85149', compareColor: '#f0e442', min: 20, max: 38, step: 0.5,
    threshold: favHotThreshold, compareThreshold: favHotCompareThreshold, compareEnabled: favHotCompareEnabled,
    formatLabel: t => `Tmax > ${t.toLocaleString('de-DE', {maximumFractionDigits:1})} °C`,
    match: (d, t) => d.max > t, value: d => d.max,
    setThreshold: t => { favHotThreshold = t; }, setCompareThreshold: t => { favHotCompareThreshold = t; }, setCompareEnabled: v => { favHotCompareEnabled = v; },
    hint: 'Balken = aktiver Grenzwert. Gestrichelte Linie = Vergleichswert. Der Bereich zwischen beiden Reglern zeigt, welches Jahr besonders viele Tage in genau dieser Temperaturzone hatte.',
  });

  const dailyAgg = aggregateMinAvgMax(raw, 'daily');
  renderRecordRankingCard(container, dailyAgg, {
    chartKey: `rank_${prefix}_hot`, chartId: `rankHotChart_${prefix}`,
    title: '🔥 Rekord-Rangliste – Höchstwerte', unit, topN: 50,
    color: '#f85149', otherColor: '#30363d', sortDesc: true,
    recordLabel: 'Höchstwert der Reihe', pickValue: r => r.max,
    hint: 'Top 50 Tageshöchstwerte der gesamten Messreihe (nicht klimatologisch). Rot = aktuelles Jahr, grau = frühere Jahre.',
  });
  renderRecordRankingCard(container, dailyAgg, {
    chartKey: `rank_${prefix}_cold`, chartId: `rankColdChart_${prefix}`,
    title: '🥶 Rekord-Rangliste – Tiefstwerte', unit, topN: 50,
    color: '#58a6ff', otherColor: '#30363d', sortDesc: false,
    recordLabel: 'Tiefstwert der Reihe', pickValue: r => r.min,
    hint: 'Top 50 Tagestiefstwerte der gesamten Messreihe (nicht klimatologisch). Blau = aktuelles Jahr, grau = frühere Jahre.',
  });
}

/* ── Charts ── */
function renderCharts(filtered) {
  const container = document.getElementById('chartsContainer');

  // Destroy old charts
  Object.values(chartInstances).forEach(c => c.destroy());
  chartInstances = {};
  container.innerHTML = '';

  const sensors = [...selSensors].filter(s => filtered.some(d => d.sensor === s));
  if (sensors.length === 0) {
    container.innerHTML = '<div class="no-data">Keine Daten für die gewählten Filter.</div>';
    return;
  }

  sensors.forEach((sensor, sensorIndex) => {
    const sData = filtered.filter(d => d.sensor === sensor);
    const unit  = sData[0]?.unit || '';
    const modules = [...selModules].filter(m => sData.some(d => d.module === m));
    const canvasId = `chart_${sensorIndex}`;

    const card = document.createElement('div');
    card.className = 'chart-card';
    card.innerHTML = `
      <div class="chart-header">
        <h2>${sensor}</h2>
        <span class="unit-badge">${unit}</span>
      </div>
      <div class="chart-wrap"><canvas id="${canvasId}"></canvas></div>
      <div class="chart-hint">Scrollen zum Zoomen · Ziehen zum Verschieben · Doppelklick zum Zurücksetzen</div>`;
    container.appendChild(card);

    const datasets = modules.map(module => {
      const pts = sData
        .filter(d => d.module === module)
        .map(d => ({ x: d.ts, y: d.value }))
        .sort((a,b) => a.x - b.x);
      const color = PAYLOAD.colors[module] || '#58a6ff';
      return {
        label: module,
        data: pts,
        borderColor: color,
        backgroundColor: color + '22',
        borderWidth: aggMode === 'daily' ? 2 : 1.5,
        pointRadius: aggMode === 'daily' || pts.length === 1 ? 3 : 0,
        pointHoverRadius: 5,
        spanGaps: true,
        tension: 0.3,
        fill: modules.length === 1,
      };
    });

    const ctx = document.getElementById(canvasId).getContext('2d');
    chartInstances[sensor] = new Chart(ctx, {
      type: 'line',
      data: { datasets },
      options: {
        responsive: true,
        maintainAspectRatio: false,
        parsing: false,
        normalized: true,
        animation: { duration: 400 },
        interaction: { mode: 'index', intersect: false },
        scales: {
          x: {
            type: 'time',
            time: { tooltipFormat: 'dd.MM.yyyy HH:mm', displayFormats: { hour:'dd.MM HH:mm', day:'dd.MM.yy', week:'dd.MM.yy', month:'MMM yy' } },
            grid: { color: '#30363d' },
            ticks: { color: '#8b949e', maxTicksLimit: 8 },
          },
          y: {
            grid: { color: '#30363d' },
            ticks: { color: '#8b949e',
              callback: v => v.toLocaleString('de-DE', {maximumFractionDigits:1}) + (unit?' '+unit:'') },
          },
        },
        plugins: {
          legend: { labels: { color: '#e6edf3', boxWidth: 12, padding: 14 } },
          tooltip: {
            backgroundColor: '#161b22',
            borderColor: '#30363d',
            borderWidth: 1,
            titleColor: '#e6edf3',
            bodyColor: '#8b949e',
            callbacks: {
              label: ctx => ` ${ctx.dataset.label}: ${ctx.parsed.y?.toLocaleString('de-DE',{maximumFractionDigits:2})} ${unit}`,
            },
          },
          zoom: {
            zoom: { wheel: { enabled: true }, pinch: { enabled: true }, mode: 'x' },
            pan:  { enabled: true, mode: 'x' },
          },
          decimation: {
            enabled: aggMode === 'raw',
            algorithm: 'min-max',
          },
        },
      },
    });

    // Doppelklick → Reset Zoom
    document.getElementById(canvasId).addEventListener('dblclick', () => {
      chartInstances[sensor]?.resetZoom();
    });

    if (isTemperatureSensor(sensor, unit)) {
      const rawForSensor = lastFilteredRaw.filter(d => d.sensor === sensor && selModules.has(d.module));
      renderDailyExtremes(container, rawForSensor, sensor, unit, sensorIndex);
      renderTemperatureFavorites(container, rawForSensor, sensor, unit, sensorIndex);
    }
  });

  requestAnimationFrame(() => {
    requestAnimationFrame(() => {
      Object.values(chartInstances).forEach(chart => {
        if (chart?.canvas?.isConnected) chart.resize();
      });
    });
  });
}

/* ── View switching & presets ── */
let currentView = 'charts';
let lastFiltered = [];
let lastFilteredRaw = [];

function switchView(btn) {
  currentView = btn.dataset.view;
  document.querySelectorAll('.tab').forEach(t => t.classList.toggle('active', t === btn));
  ['charts','records','trends','longterm','compare','climate','extremes','indoor'].forEach(v =>
    document.getElementById('view-' + v).classList.toggle('hidden', v !== currentView));
  renderCurrentView();
}

function setPreset(days) {
  const bounds = getDateBounds();
  if (days === 0) {
    document.getElementById('dateFrom').value = bounds.min;
  } else {
    const max = new Date(bounds.max);
    const from = new Date(max.getTime() - days * 86400000);
    const fromIso = from.toISOString().slice(0,10);
    document.getElementById('dateFrom').value = fromIso < bounds.min ? bounds.min : fromIso;
  }
  document.getElementById('dateTo').value = bounds.max;
  applyFilters();
}

/* ── Linear regression (least squares) on [{ts,value}] ── */
function linearRegression(pts) {
  const n = pts.length;
  if (n < 2) return null;
  let sx=0, sy=0, sxx=0, sxy=0;
  const t0 = pts[0].ts;
  pts.forEach(p => {
    const x = (p.ts - t0) / 86400000; // days
    const y = p.value;
    sx += x; sy += y; sxx += x*x; sxy += x*y;
  });
  const denom = n*sxx - sx*sx;
  if (denom === 0) return null;
  const slope = (n*sxy - sx*sy) / denom;        // per day
  const intercept = (sy - slope*sx) / n;
  return { slope, intercept, t0, slopePerYear: slope * 365.25 };
}

/* Moving average over already-sorted [{ts,value}] */
function movingAverage(pts, windowDays) {
  if (!pts.length) return [];
  const w = windowDays * 86400000;
  const out = [];
  let lo = 0, sum = 0;
  for (let i=0; i<pts.length; i++) {
    sum += pts[i].value;
    while (pts[i].ts - pts[lo].ts > w) { sum -= pts[lo].value; lo++; }
    out.push({ x: pts[i].ts, y: sum / (i - lo + 1) });
  }
  return out;
}

/* ── Records & Tops ── */
function renderRecords(raw) {
  const container = document.getElementById('recordsContainer');
  container.innerHTML = '';
  const sensors = [...selSensors].filter(s => raw.some(d => d.sensor === s));
  if (!sensors.length) { container.innerHTML = '<div class="no-data">Keine Daten.</div>'; return; }

  const grid = document.createElement('div');
  grid.className = 'records-grid';
  container.appendChild(grid);

  sensors.forEach(sensor => {
    const sData = raw.filter(d => d.sensor === sensor);
    const unit = sData[0]?.unit || '';

    // Top 20 highest / lowest – ein Rekord pro Tag (Tagesmax / Tagesmin)
    const daily = aggregateMinAvgMax(sData, 'daily');
    const top = [...daily].sort((a,b) => b.max - a.max).slice(0, 20);
    const bot = [...daily].sort((a,b) => a.min - b.min).slice(0, 20);

    // Daily extremes (kurzliste)
    const hottestDay  = [...daily].sort((a,b) => b.max - a.max).slice(0, 5);
    const coldestDay  = [...daily].sort((a,b) => a.min - b.min).slice(0, 5);

    // Monthly extremes (avg)
    const monthly = aggregateMinAvgMax(sData, 'monthly');
    const bestMonths  = [...monthly].sort((a,b) => b.avg - a.avg).slice(0, 5);
    const worstMonths = [...monthly].sort((a,b) => a.avg - b.avg).slice(0, 5);

    // Yearly stats
    const yearly = aggregateMinAvgMax(sData, 'yearly');

    const card = document.createElement('div');
    card.className = 'rec-card';
    card.innerHTML = `
      <h3>${sensor} <span class="badge">${unit}</span></h3>
      <table class="rec-table">
        <tr><th colspan="3">🔥 Top 20 Höchstwerte (Tagesmax, 1 Eintrag/Tag)</th></tr>
        <tr><th>Tag</th><th>Modul</th><th style="text-align:right">Tagesmax</th></tr>
        ${top.map(r => `<tr><td>${fmtDay(r.ts)}</td><td>${r.module}</td><td class="val max">${fmt(r.max, unit)}</td></tr>`).join('')}
      </table>
      <table class="rec-table" style="margin-top:12px">
        <tr><th colspan="3">❄️ Top 20 Tiefstwerte (Tagesmin, 1 Eintrag/Tag)</th></tr>
        <tr><th>Tag</th><th>Modul</th><th style="text-align:right">Tagesmin</th></tr>
        ${bot.map(r => `<tr><td>${fmtDay(r.ts)}</td><td>${r.module}</td><td class="val min">${fmt(r.min, unit)}</td></tr>`).join('')}
      </table>
      <table class="rec-table" style="margin-top:12px">
        <tr><th colspan="4">☀️ Heisseste Tage (Tagesmax)</th></tr>
        <tr><th>Tag</th><th>Modul</th><th style="text-align:right">Max</th><th style="text-align:right">Mittel</th></tr>
        ${hottestDay.map(r => `<tr><td>${fmtDay(r.ts)}</td><td>${r.module}</td><td class="val max">${fmt(r.max, unit)}</td><td class="val">${fmt(r.avg, unit)}</td></tr>`).join('')}
      </table>
      <table class="rec-table" style="margin-top:12px">
        <tr><th colspan="4">❄️ Kälteste Tage (Tagesmin)</th></tr>
        <tr><th>Tag</th><th>Modul</th><th style="text-align:right">Min</th><th style="text-align:right">Mittel</th></tr>
        ${coldestDay.map(r => `<tr><td>${fmtDay(r.ts)}</td><td>${r.module}</td><td class="val min">${fmt(r.min, unit)}</td><td class="val">${fmt(r.avg, unit)}</td></tr>`).join('')}
      </table>
      <table class="rec-table" style="margin-top:12px">
        <tr><th colspan="3">📅 Wärmste Monate (Ø)</th></tr>
        <tr><th>Monat</th><th>Modul</th><th style="text-align:right">Ø</th></tr>
        ${bestMonths.map(r => `<tr><td>${fmtMonth(r.ts)}</td><td>${r.module}</td><td class="val max">${fmt(r.avg, unit)}</td></tr>`).join('')}
      </table>
      <table class="rec-table" style="margin-top:12px">
        <tr><th colspan="3">❄️ Kälteste Monate (Ø)</th></tr>
        <tr><th>Monat</th><th>Modul</th><th style="text-align:right">Ø</th></tr>
        ${worstMonths.map(r => `<tr><td>${fmtMonth(r.ts)}</td><td>${r.module}</td><td class="val min">${fmt(r.avg, unit)}</td></tr>`).join('')}
      </table>
      <table class="rec-table" style="margin-top:12px">
        <tr><th colspan="5">📊 Jahresübersicht</th></tr>
        <tr><th>Jahr</th><th>Modul</th><th style="text-align:right">Min</th><th style="text-align:right">Ø</th><th style="text-align:right">Max</th></tr>
        ${yearly.map(r => `<tr><td>${new Date(r.ts).getFullYear()}</td><td>${r.module}</td><td class="val min">${fmt(r.min,unit)}</td><td class="val">${fmt(r.avg,unit)}</td><td class="val max">${fmt(r.max,unit)}</td></tr>`).join('')}
      </table>`;
    grid.appendChild(card);
  });
}

/* ── Trends ── */
let trendCharts = {};
let trendWindow = 30;
let trendSensor = null;
let yoyYears = null;  // Set<number> of selected years for YoY (null = all)

function renderTrends(raw) {
  const container = document.getElementById('trendsContainer');
  Object.values(trendCharts).forEach(c => c.destroy());
  trendCharts = {};
  container.innerHTML = '';

  const sensors = [...selSensors].filter(s => raw.some(d => d.sensor === s));
  if (!sensors.length) { container.innerHTML = '<div class="no-data">Keine Daten.</div>'; return; }
  if (!trendSensor || !sensors.includes(trendSensor)) trendSensor = sensors[0];

  // Controls
  const ctrl = document.createElement('div');
  ctrl.className = 'trend-controls';
  ctrl.innerHTML = `
    <label>Sensor:</label>
    <select id="trendSensorSel">${sensors.map(s => `<option ${s===trendSensor?'selected':''}>${s}</option>`).join('')}</select>
    <label>Gleitender Mittelwert:</label>
    <select id="trendWinSel">
      <option value="7"${trendWindow===7?' selected':''}>7 Tage</option>
      <option value="30"${trendWindow===30?' selected':''}>30 Tage</option>
      <option value="90"${trendWindow===90?' selected':''}>90 Tage</option>
      <option value="365"${trendWindow===365?' selected':''}>365 Tage</option>
    </select>`;
  container.appendChild(ctrl);
  ctrl.querySelector('#trendSensorSel').onchange = e => { trendSensor = e.target.value; renderTrends(raw); };
  ctrl.querySelector('#trendWinSel').onchange    = e => { trendWindow = +e.target.value; renderTrends(raw); };

  const sData = raw.filter(d => d.sensor === trendSensor);
  const unit  = sData[0]?.unit || '';
  const modules = [...selModules].filter(m => sData.some(d => d.module === m));

  // ----- Card 1: daily mean + moving avg + regression -----
  const card1 = document.createElement('div');
  card1.className = 'chart-card';
  card1.innerHTML = `
    <div class="chart-header"><h2>${trendSensor} – Tagesmittel, ${trendWindow}-Tage-Mittel, lineare Regression</h2>
      <span class="unit-badge">${unit}</span></div>
    <div class="chart-wrap tall"><canvas id="trendChart"></canvas></div>
    <div class="trend-info" id="trendInfo"></div>`;
  container.appendChild(card1);

  const datasets = [];
  const infoEl = card1.querySelector('#trendInfo');
  modules.forEach((module, mi) => {
    const color = PAYLOAD.colors[module] || '#58a6ff';
    const daily = aggregateData(sData.filter(d => d.module === module), 'daily')
      .map(d => ({ ts: d.ts, value: d.value }));
    if (!daily.length) return;

    datasets.push({
      label: `${module} (Tag Ø)`, data: daily.map(d => ({x: d.ts, y: d.value})),
      borderColor: color + '66', backgroundColor: 'transparent',
      borderWidth: 1, pointRadius: 0, tension: 0.2,
    });
    const ma = movingAverage(daily, trendWindow);
    datasets.push({
      label: `${module} (Ø ${trendWindow}T)`, data: ma,
      borderColor: color, backgroundColor: 'transparent',
      borderWidth: 2.5, pointRadius: 0, tension: 0.3,
    });
    const reg = linearRegression(daily);
    if (reg) {
      const x1 = daily[0].ts, x2 = daily[daily.length-1].ts;
      const y1 = reg.intercept;
      const y2 = reg.intercept + reg.slope * ((x2 - reg.t0) / 86400000);
      datasets.push({
        label: `${module} Trend`, data: [{x:x1,y:y1},{x:x2,y:y2}],
        borderColor: color, backgroundColor: 'transparent',
        borderWidth: 1.5, borderDash:[6,4], pointRadius:0,
      });
      const sign = reg.slopePerYear >= 0 ? '↗️' : '↘️';
      const cls  = reg.slopePerYear >= 0 ? 'up' : 'down';
      const pill = document.createElement('span');
      pill.className = 'pill ' + cls;
      pill.innerHTML = `${sign} <b>${module}</b>: ${reg.slopePerYear >= 0 ? '+' : ''}${reg.slopePerYear.toLocaleString('de-DE',{maximumFractionDigits:2})} ${unit}/Jahr`;
      infoEl.appendChild(pill);
    }
  });

  trendCharts.trend = new Chart(document.getElementById('trendChart').getContext('2d'), {
    type: 'line', data: { datasets },
    options: commonChartOpts(unit, true),
  });

  // ----- Card 2: Year-over-Year overlay -----
  // Build day-of-year buckets per (module, year)
  const card2 = document.createElement('div');
  card2.className = 'chart-card';
  card2.innerHTML = `
    <div class="chart-header"><h2>${trendSensor} – Jahresvergleich (Year-over-Year)</h2>
      <span class="unit-badge">${unit}</span></div>
    <div id="yoyYearChips" class="year-chips"></div>
    <div class="chart-wrap tall"><canvas id="yoyChart"></canvas></div>
    <div class="chart-hint">Tagesmittel über alle gewählten Module, je Jahr eine Linie. X-Achse: Tag im Jahr.</div>`;
  container.appendChild(card2);

  const dailyAll = aggregateData(sData, 'daily');
  const byYear = {};
  dailyAll.forEach(d => {
    const dt = new Date(d.ts);
    const y = dt.getFullYear();
    // synthetic date in year 2000 (leap) so DOY aligns
    const synth = new Date(2000, dt.getMonth(), dt.getDate()).getTime();
    if (!byYear[y]) byYear[y] = {};
    if (!byYear[y][synth]) byYear[y][synth] = [];
    byYear[y][synth].push(d.value);
  });
  const years = Object.keys(byYear).map(Number).sort();
  const palette = ['#58a6ff','#3fb950','#d29922','#f85149','#bc8cff','#ff7b72','#39c5cf','#ffa657','#7ee787','#f0883e'];
  const yearColor = {};
  years.forEach((y, i) => { yearColor[y] = palette[i % palette.length]; });

  // Initialise / sync year selection: default = all years
  if (!yoyYears) yoyYears = new Set(years);
  else {
    // drop years that no longer exist in current data
    [...yoyYears].forEach(y => { if (!years.includes(y)) yoyYears.delete(y); });
    // if filter wiped out everything, fall back to all
    if (!yoyYears.size) yoyYears = new Set(years);
  }

  // Render year-chip selector
  const chipsEl = document.getElementById('yoyYearChips');
  const renderChips = () => {
    chipsEl.innerHTML =
      `<span class="label">Jahre:</span>` +
      years.map(y => {
        const active = yoyYears.has(y);
        return `<button class="year-chip${active?' active':''}" data-year="${y}">`
             + `<span class="swatch" style="background:${yearColor[y]}"></span>${y}</button>`;
      }).join('') +
      `<button class="link-btn" data-act="all">Alle</button>` +
      `<button class="link-btn" data-act="none">Keine</button>` +
      `<button class="link-btn" data-act="last2">Letzte 2</button>` +
      `<button class="link-btn" data-act="last5">Letzte 5</button>`;
  };
  renderChips();
  chipsEl.addEventListener('click', e => {
    const chip = e.target.closest('.year-chip');
    const link = e.target.closest('.link-btn');
    if (chip) {
      const y = +chip.dataset.year;
      if (yoyYears.has(y)) yoyYears.delete(y); else yoyYears.add(y);
      if (!yoyYears.size) yoyYears.add(y); // never empty
    } else if (link) {
      const act = link.dataset.act;
      if (act === 'all')   yoyYears = new Set(years);
      if (act === 'none')  yoyYears = new Set([years[years.length-1]]);
      if (act === 'last2') yoyYears = new Set(years.slice(-2));
      if (act === 'last5') yoyYears = new Set(years.slice(-5));
    } else return;
    renderChips();
    rebuildYoyChart();
  });

  const buildYoyDatasets = () => years.filter(y => yoyYears.has(y)).map(y => {
    const pts = Object.entries(byYear[y])
      .map(([x, vs]) => ({ x: +x, y: vs.reduce((a,b)=>a+b,0)/vs.length }))
      .sort((a,b)=>a.x-b.x);
    return {
      label: String(y), data: pts,
      borderColor: yearColor[y],
      backgroundColor: 'transparent', borderWidth: 1.8,
      pointRadius: 0, tension: 0.25,
    };
  });

  const rebuildYoyChart = () => {
    if (trendCharts.yoy) { trendCharts.yoy.data.datasets = buildYoyDatasets(); trendCharts.yoy.update(); }
  };

  trendCharts.yoy = new Chart(document.getElementById('yoyChart').getContext('2d'), {
    type: 'line', data: { datasets: buildYoyDatasets() },
    options: {
      responsive: true, maintainAspectRatio: false,
      interaction: { mode: 'index', intersect: false },
      scales: {
        x: { type: 'time',
             time: { tooltipFormat: 'dd.MM', displayFormats: { day:'dd.MM', month:'MMM' }, unit: 'month' },
             grid: { color: '#30363d' }, ticks: { color: '#8b949e' } },
        y: { grid: { color: '#30363d' }, ticks: { color: '#8b949e',
             callback: v => v.toLocaleString('de-DE',{maximumFractionDigits:1}) + (unit?' '+unit:'') } },
      },
      plugins: {
        legend: { labels: { color: '#e6edf3', boxWidth: 12, padding: 10 } },
        tooltip: { backgroundColor: '#161b22', borderColor: '#30363d', borderWidth: 1,
                   titleColor: '#e6edf3', bodyColor: '#8b949e',
                   callbacks: { label: ctx => ` ${ctx.dataset.label}: ${ctx.parsed.y?.toLocaleString('de-DE',{maximumFractionDigits:2})} ${unit}` } },
      },
    },
  });

  // ----- Card 3: Tagesprofil (24h) -----
  const card3 = document.createElement('div');
  card3.className = 'chart-card';
  card3.innerHTML = `
    <div class="chart-header"><h2>${trendSensor} – Tagesprofil (Ø nach Stunde)</h2>
      <span class="unit-badge">${unit}</span></div>
    <div class="chart-wrap"><canvas id="hourChart"></canvas></div>
    <div class="chart-hint">Mittelwert ± Standardabweichung über alle Tage des gewählten Zeitraums.</div>`;
  container.appendChild(card3);

  const hourDatasets = [];
  modules.forEach(module => {
    const color = PAYLOAD.colors[module] || '#58a6ff';
    const byHour = Array.from({length:24}, () => []);
    sData.filter(d => d.module === module).forEach(d => {
      byHour[new Date(d.ts).getHours()].push(d.value);
    });
    const means = byHour.map(vs => vs.length ? vs.reduce((a,b)=>a+b,0)/vs.length : null);
    const stds  = byHour.map(vs => {
      if (!vs.length) return 0;
      const m = vs.reduce((a,b)=>a+b,0)/vs.length;
      return Math.sqrt(vs.reduce((s,v)=>s+(v-m)*(v-m),0)/vs.length);
    });
    hourDatasets.push({
      label: module + ' Ø+σ', data: means.map((m,i) => m===null?null:m+stds[i]),
      borderColor: 'transparent', backgroundColor: color + '22',
      fill: '+1', pointRadius: 0, tension: 0.35, borderWidth: 0,
    });
    hourDatasets.push({
      label: module + ' Ø-σ', data: means.map((m,i) => m===null?null:m-stds[i]),
      borderColor: 'transparent', backgroundColor: 'transparent',
      fill: false, pointRadius: 0, tension: 0.35, borderWidth: 0,
    });
    hourDatasets.push({
      label: module + ' Ø', data: means,
      borderColor: color, backgroundColor: color,
      borderWidth: 2.5, pointRadius: 3, tension: 0.35, fill: false,
    });
  });
  trendCharts.hour = new Chart(document.getElementById('hourChart').getContext('2d'), {
    type: 'line',
    data: { labels: Array.from({length:24}, (_,i) => i.toString().padStart(2,'0')+':00'), datasets: hourDatasets },
    options: {
      responsive: true, maintainAspectRatio: false,
      interaction: { mode: 'index', intersect: false },
      scales: {
        x: { grid: { color: '#30363d' }, ticks: { color: '#8b949e' } },
        y: { grid: { color: '#30363d' }, ticks: { color: '#8b949e',
             callback: v => v.toLocaleString('de-DE',{maximumFractionDigits:1}) + (unit?' '+unit:'') } },
      },
      plugins: {
        legend: { labels: { color: '#e6edf3', boxWidth: 12, padding: 10,
          filter: it => !it.text.endsWith('Ø+σ') && !it.text.endsWith('Ø-σ') } },
        tooltip: { backgroundColor: '#161b22', borderColor: '#30363d', borderWidth: 1,
                   titleColor: '#e6edf3', bodyColor: '#8b949e' },
      },
    },
  });

  // ----- Card 4: Wochenprofil (7 Tage) -----
  const card4 = document.createElement('div');
  card4.className = 'chart-card';
  card4.innerHTML = `
    <div class="chart-header"><h2>${trendSensor} – Wochenprofil (Ø nach Wochentag)</h2>
      <span class="unit-badge">${unit}</span></div>
    <div class="chart-wrap"><canvas id="weekChart"></canvas></div>`;
  container.appendChild(card4);

  const weekLabels = ['Mo','Di','Mi','Do','Fr','Sa','So'];
  const weekDatasets = modules.map(module => {
    const color = PAYLOAD.colors[module] || '#58a6ff';
    const byDow = Array.from({length:7}, () => []);
    sData.filter(d => d.module === module).forEach(d => {
      // Convert JS getDay (0=Sun..6=Sat) to Mo=0..So=6
      const dow = (new Date(d.ts).getDay() + 6) % 7;
      byDow[dow].push(d.value);
    });
    return {
      label: module,
      data: byDow.map(vs => vs.length ? Math.round(vs.reduce((a,b)=>a+b,0)/vs.length*100)/100 : null),
      backgroundColor: color + 'cc', borderColor: color, borderWidth: 1,
    };
  });
  trendCharts.week = new Chart(document.getElementById('weekChart').getContext('2d'), {
    type: 'bar',
    data: { labels: weekLabels, datasets: weekDatasets },
    options: {
      responsive: true, maintainAspectRatio: false,
      scales: {
        x: { grid: { color: '#30363d' }, ticks: { color: '#8b949e' } },
        y: { grid: { color: '#30363d' }, ticks: { color: '#8b949e',
             callback: v => v.toLocaleString('de-DE',{maximumFractionDigits:1}) + (unit?' '+unit:'') } },
      },
      plugins: {
        legend: { labels: { color: '#e6edf3', boxWidth: 12, padding: 10 } },
        tooltip: { backgroundColor: '#161b22', borderColor: '#30363d', borderWidth: 1,
                   titleColor: '#e6edf3', bodyColor: '#8b949e' },
      },
    },
  });
}
let longCharts = {};
function renderLongterm(raw) {
  const container = document.getElementById('longtermContainer');
  Object.values(longCharts).forEach(c => c.destroy());
  longCharts = {};
  container.innerHTML = '';

  const sensors = [...selSensors].filter(s => raw.some(d => d.sensor === s));
  if (!sensors.length) { container.innerHTML = '<div class="no-data">Keine Daten.</div>'; return; }

  const monthNames = ['Jan','Feb','Mär','Apr','Mai','Jun','Jul','Aug','Sep','Okt','Nov','Dez'];

  sensors.forEach(sensor => {
    const sData = raw.filter(d => d.sensor === sensor);
    const unit = sData[0]?.unit || '';

    // ---- Heatmap (year x month, mean of all selected modules) ----
    const monthly = aggregateMinAvgMax(
      sData.map(d => ({...d})), 'monthly');
    // Average across modules within the same year-month
    const ym = {};
    monthly.forEach(r => {
      const dt = new Date(r.ts);
      const k = `${dt.getFullYear()}-${dt.getMonth()}`;
      if (!ym[k]) ym[k] = { y: dt.getFullYear(), m: dt.getMonth(), vals: [] };
      ym[k].vals.push(r.avg);
    });
    const cells = Object.values(ym).map(c => ({...c, avg: c.vals.reduce((a,b)=>a+b,0)/c.vals.length }));
    if (!cells.length) return;
    const years = [...new Set(cells.map(c => c.y))].sort();
    const mn = Math.min(...cells.map(c => c.avg));
    const mx = Math.max(...cells.map(c => c.avg));

    let rowsHtml = '';
    years.forEach(yr => {
      let row = `<tr><td class="year-label">${yr}</td>`;
      for (let m = 0; m < 12; m++) {
        const c = cells.find(x => x.y === yr && x.m === m);
        if (!c) { row += `<td class="cell empty">–</td>`; }
        else {
          const t = (c.avg - mn) / (mx - mn || 1);
          row += `<td class="cell" style="background:${heatColor(t)}" title="${yr}-${m+1}: ${fmt(c.avg, unit)}">${c.avg.toLocaleString('de-DE',{maximumFractionDigits:1})}</td>`;
        }
      }
      row += '</tr>';
      rowsHtml += row;
    });

    const card = document.createElement('div');
    card.className = 'heatmap-card';
    card.innerHTML = `
      <div class="chart-header"><h2>${sensor} – Monatsmittel-Heatmap</h2>
        <span class="unit-badge">${unit}</span></div>
      <table class="heatmap">
        <tr><th></th>${monthNames.map(m => `<th>${m}</th>`).join('')}</tr>
        ${rowsHtml}
      </table>
      <div class="heatmap-legend">
        <span>${fmt(mn, unit)}</span>
        <span class="legend-bar"></span>
        <span>${fmt(mx, unit)}</span>
      </div>`;
    container.appendChild(card);

    // ---- Yearly bar chart (min/avg/max per year) ----
    const yearly = aggregateMinAvgMax(sData, 'yearly');
    // group modules together
    const ybar = {};
    yearly.forEach(r => {
      const yr = new Date(r.ts).getFullYear();
      if (!ybar[yr]) ybar[yr] = { min:[], avg:[], max:[] };
      ybar[yr].min.push(r.min); ybar[yr].avg.push(r.avg); ybar[yr].max.push(r.max);
    });
    const yrs = Object.keys(ybar).map(Number).sort();
    const mins = yrs.map(y => Math.min(...ybar[y].min));
    const avgs = yrs.map(y => ybar[y].avg.reduce((a,b)=>a+b,0)/ybar[y].avg.length);
    const maxs = yrs.map(y => Math.max(...ybar[y].max));

    const barCard = document.createElement('div');
    barCard.className = 'chart-card';
    const barId = `barChart_${sensor.replace(/[^a-z0-9]/gi,'')}`;
    barCard.innerHTML = `
      <div class="chart-header"><h2>${sensor} – Jahreswerte (Min / Ø / Max)</h2>
        <span class="unit-badge">${unit}</span></div>
      <div class="chart-wrap tall"><canvas id="${barId}"></canvas></div>`;
    container.appendChild(barCard);

    longCharts[sensor] = new Chart(document.getElementById(barId).getContext('2d'), {
      type: 'bar',
      data: {
        labels: yrs,
        datasets: [
          { label: 'Min', data: mins, backgroundColor: '#58a6ff99', borderColor:'#58a6ff', borderWidth:1 },
          { label: 'Ø',   data: avgs, backgroundColor: '#3fb95099', borderColor:'#3fb950', borderWidth:1 },
          { label: 'Max', data: maxs, backgroundColor: '#f8514999', borderColor:'#f85149', borderWidth:1 },
        ],
      },
      options: {
        responsive: true, maintainAspectRatio: false,
        scales: {
          x: { grid: { color:'#30363d' }, ticks: { color:'#8b949e' } },
          y: { grid: { color:'#30363d' }, ticks: { color:'#8b949e',
               callback: v => v.toLocaleString('de-DE',{maximumFractionDigits:1}) + (unit?' '+unit:'') } },
        },
        plugins: {
          legend: { labels: { color:'#e6edf3', boxWidth: 12, padding: 10 } },
          tooltip: { backgroundColor: '#161b22', borderColor:'#30363d', borderWidth: 1,
                     titleColor: '#e6edf3', bodyColor: '#8b949e',
                     callbacks: { label: ctx => ` ${ctx.dataset.label}: ${ctx.parsed.y?.toLocaleString('de-DE',{maximumFractionDigits:2})} ${unit}` } },
        },
      },
    });

    // ---- Histogramm (Werteverteilung) ----
    const vals = sData.map(d => d.value);
    const vmin = Math.min(...vals), vmax = Math.max(...vals);
    const binCount = 30;
    const binWidth = (vmax - vmin) / binCount || 1;
    const bins = Array.from({length: binCount}, () => 0);
    vals.forEach(v => {
      const i = Math.min(binCount - 1, Math.floor((v - vmin) / binWidth));
      bins[i]++;
    });
    const labels = Array.from({length: binCount}, (_, i) =>
      (vmin + i * binWidth).toLocaleString('de-DE', {maximumFractionDigits:1}));
    const histCard = document.createElement('div');
    histCard.className = 'chart-card';
    const histId = `hist_${sensor.replace(/[^a-z0-9]/gi,'')}`;
    histCard.innerHTML = `
      <div class="chart-header"><h2>${sensor} – Werteverteilung (Histogramm)</h2>
        <span class="unit-badge">${unit}</span></div>
      <div class="chart-wrap"><canvas id="${histId}"></canvas></div>
      <div class="chart-hint">n = ${vals.length.toLocaleString('de-DE')} · Bin-Breite ≈ ${binWidth.toLocaleString('de-DE',{maximumFractionDigits:2})} ${unit}</div>`;
    container.appendChild(histCard);

    // Colored bars by bin position
    const histColors = bins.map((_, i) => heatColor(i / (binCount - 1)));
    longCharts[sensor + '_hist'] = new Chart(document.getElementById(histId).getContext('2d'), {
      type: 'bar',
      data: { labels, datasets: [{ label: 'Häufigkeit', data: bins,
              backgroundColor: histColors, borderWidth: 0 }] },
      options: {
        responsive: true, maintainAspectRatio: false,
        scales: {
          x: { grid: { color:'#30363d' }, ticks: { color:'#8b949e', maxTicksLimit: 12,
               callback: function(v) { return this.getLabelForValue(v) + ' ' + unit; } } },
          y: { grid: { color:'#30363d' }, ticks: { color:'#8b949e' }, title: { display: true, text: 'Anzahl', color: '#8b949e' } },
        },
        plugins: {
          legend: { display: false },
          tooltip: { backgroundColor: '#161b22', borderColor:'#30363d', borderWidth: 1,
                     titleColor: '#e6edf3', bodyColor: '#8b949e' },
        },
      },
    });
  });
}

/* ── Vergleich: Module gegeneinander ── */
let cmpCharts = {};
let cmpSensor = null, cmpModA = null, cmpModB = null;

function renderCompare(raw) {
  const container = document.getElementById('compareContainer');
  Object.values(cmpCharts).forEach(c => c.destroy());
  cmpCharts = {};
  container.innerHTML = '';

  const sensors = [...selSensors].filter(s => raw.some(d => d.sensor === s));
  if (!sensors.length) { container.innerHTML = '<div class="no-data">Keine Daten.</div>'; return; }
  if (!cmpSensor || !sensors.includes(cmpSensor)) cmpSensor = sensors[0];

  const sData = raw.filter(d => d.sensor === cmpSensor);
  const modules = [...new Set(sData.map(d => d.module))].sort();
  const unit = sData[0]?.unit || '';

  if (modules.length < 2) {
    container.innerHTML = `<div class="no-data">Für Sensor <b>${cmpSensor}</b> ist nur 1 Modul vorhanden. Vergleich benötigt mindestens 2 Module mit demselben Sensor.</div>`;
    return;
  }
  if (!cmpModA || !modules.includes(cmpModA)) cmpModA = modules[0];
  if (!cmpModB || !modules.includes(cmpModB) || cmpModB === cmpModA) cmpModB = modules.find(m => m !== cmpModA);

  // Controls
  const ctrl = document.createElement('div');
  ctrl.className = 'trend-controls';
  ctrl.innerHTML = `
    <label>Sensor:</label>
    <select id="cmpSensorSel">${sensors.map(s => `<option ${s===cmpSensor?'selected':''}>${s}</option>`).join('')}</select>
    <label>Modul A:</label>
    <select id="cmpASel">${modules.map(m => `<option ${m===cmpModA?'selected':''}>${m}</option>`).join('')}</select>
    <label>Modul B:</label>
    <select id="cmpBSel">${modules.map(m => `<option ${m===cmpModB?'selected':''}>${m}</option>`).join('')}</select>`;
  container.appendChild(ctrl);
  ctrl.querySelector('#cmpSensorSel').onchange = e => { cmpSensor = e.target.value; cmpModA = cmpModB = null; renderCompare(raw); };
  ctrl.querySelector('#cmpASel').onchange      = e => { cmpModA = e.target.value; renderCompare(raw); };
  ctrl.querySelector('#cmpBSel').onchange      = e => { cmpModB = e.target.value; renderCompare(raw); };

  const colA = PAYLOAD.colors[cmpModA] || '#58a6ff';
  const colB = PAYLOAD.colors[cmpModB] || '#f85149';

  // Align hourly buckets (so timestamps line up)
  const aHour = aggregateData(sData.filter(d => d.module === cmpModA), 'hourly');
  const bHour = aggregateData(sData.filter(d => d.module === cmpModB), 'hourly');
  const bMap = new Map(bHour.map(d => [d.ts, d.value]));
  const paired = aHour.filter(d => bMap.has(d.ts))
                      .map(d => ({ ts: d.ts, a: d.value, b: bMap.get(d.ts) }));

  if (!paired.length) {
    const warn = document.createElement('div');
    warn.className = 'no-data';
    warn.textContent = `Keine gemeinsamen Zeitstempel zwischen ${cmpModA} und ${cmpModB} im Zeitraum.`;
    container.appendChild(warn);
    return;
  }

  // ----- Stats: Pearson r + lineare Regression A→B -----
  const n = paired.length;
  let sa=0, sb=0, saa=0, sbb=0, sab=0;
  paired.forEach(p => { sa+=p.a; sb+=p.b; saa+=p.a*p.a; sbb+=p.b*p.b; sab+=p.a*p.b; });
  const ma = sa/n, mb = sb/n;
  const cov = sab/n - ma*mb;
  const va  = saa/n - ma*ma;
  const vb  = sbb/n - mb*mb;
  const r = cov / (Math.sqrt(va*vb) || 1);
  const slope = cov / (va || 1);
  const intercept = mb - slope*ma;

  // Deltas
  const deltas = paired.map(p => p.a - p.b);
  const meanDelta = deltas.reduce((a,b)=>a+b,0)/n;
  const maxDelta  = Math.max(...deltas);
  const minDelta  = Math.min(...deltas);

  // Info pills
  const info = document.createElement('div');
  info.className = 'trend-info';
  info.innerHTML = `
    <span class="pill">📊 n = <b>${n.toLocaleString('de-DE')}</b> Stunden</span>
    <span class="pill">🔗 Pearson r = <b>${r.toLocaleString('de-DE',{maximumFractionDigits:3})}</b></span>
    <span class="pill">📈 ${cmpModB} ≈ <b>${slope.toLocaleString('de-DE',{maximumFractionDigits:3})}</b> · ${cmpModA} + <b>${intercept.toLocaleString('de-DE',{maximumFractionDigits:2})}</b></span>
    <span class="pill">Δ Ø: <b>${fmt(meanDelta, unit)}</b></span>
    <span class="pill">Δ Min: <b>${fmt(minDelta, unit)}</b></span>
    <span class="pill">Δ Max: <b>${fmt(maxDelta, unit)}</b></span>`;
  container.appendChild(info);

  // ----- Card 1: Scatter A vs B with regression line -----
  const card1 = document.createElement('div');
  card1.className = 'chart-card';
  card1.innerHTML = `
    <div class="chart-header"><h2>Korrelation: ${cmpModA} vs ${cmpModB}</h2>
      <span class="unit-badge">${unit}</span></div>
    <div class="chart-wrap tall"><canvas id="cmpScatter"></canvas></div>`;
  container.appendChild(card1);

  const xMin = Math.min(...paired.map(p => p.a));
  const xMax = Math.max(...paired.map(p => p.a));
  cmpCharts.scatter = new Chart(document.getElementById('cmpScatter').getContext('2d'), {
    type: 'scatter',
    data: {
      datasets: [
        { label: 'Stundenmittel', data: paired.map(p => ({x: p.a, y: p.b})),
          backgroundColor: colA + '55', borderColor: 'transparent', pointRadius: 2 },
        { label: 'Regression', type: 'line',
          data: [{x: xMin, y: slope*xMin + intercept}, {x: xMax, y: slope*xMax + intercept}],
          borderColor: colB, backgroundColor: 'transparent', borderWidth: 2,
          pointRadius: 0, showLine: true },
        { label: 'Identität (1:1)', type: 'line',
          data: [{x: xMin, y: xMin}, {x: xMax, y: xMax}],
          borderColor: '#8b949e', backgroundColor: 'transparent', borderWidth: 1,
          borderDash:[4,4], pointRadius: 0, showLine: true },
      ],
    },
    options: {
      responsive: true, maintainAspectRatio: false,
      scales: {
        x: { type: 'linear', title: { display: true, text: cmpModA + ' [' + unit + ']', color: '#8b949e' },
             grid: { color: '#30363d' }, ticks: { color: '#8b949e' } },
        y: { type: 'linear', title: { display: true, text: cmpModB + ' [' + unit + ']', color: '#8b949e' },
             grid: { color: '#30363d' }, ticks: { color: '#8b949e' } },
      },
      plugins: {
        legend: { labels: { color: '#e6edf3', boxWidth: 12, padding: 10 } },
        tooltip: { backgroundColor: '#161b22', borderColor: '#30363d', borderWidth: 1,
                   titleColor: '#e6edf3', bodyColor: '#8b949e',
                   callbacks: {
                     label: ctx => ` ${cmpModA}: ${ctx.parsed.x?.toLocaleString('de-DE',{maximumFractionDigits:2})} ${unit}, ${cmpModB}: ${ctx.parsed.y?.toLocaleString('de-DE',{maximumFractionDigits:2})} ${unit}`
                   } },
      },
    },
  });

  // ----- Card 2: Delta-Verlauf (A - B) über Zeit (tägl.) -----
  const card2 = document.createElement('div');
  card2.className = 'chart-card';
  card2.innerHTML = `
    <div class="chart-header"><h2>Differenz im Zeitverlauf: ${cmpModA} − ${cmpModB} (Tagesmittel)</h2>
      <span class="unit-badge">${unit}</span></div>
    <div class="chart-wrap tall"><canvas id="cmpDelta"></canvas></div>`;
  container.appendChild(card2);

  // Build daily-aligned deltas
  const aDay = aggregateData(sData.filter(d => d.module === cmpModA), 'daily');
  const bDay = aggregateData(sData.filter(d => d.module === cmpModB), 'daily');
  const bDayMap = new Map(bDay.map(d => [d.ts, d.value]));
  const dailyDeltas = aDay.filter(d => bDayMap.has(d.ts))
                          .map(d => ({ x: d.ts, y: Math.round((d.value - bDayMap.get(d.ts))*100)/100 }));

  cmpCharts.delta = new Chart(document.getElementById('cmpDelta').getContext('2d'), {
    type: 'line',
    data: {
      datasets: [
        { label: `${cmpModA} − ${cmpModB}`, data: dailyDeltas,
          borderColor: colA, backgroundColor: colA + '33',
          borderWidth: 1.5, pointRadius: 0, tension: 0.3,
          fill: { target: 'origin', above: colA + '22', below: colB + '22' } },
      ],
    },
    options: commonChartOpts(unit, true),
  });

  // ----- Card 3: Monthly side-by-side bars per module -----
  const card3 = document.createElement('div');
  card3.className = 'chart-card';
  card3.innerHTML = `
    <div class="chart-header"><h2>Monatsmittel je Modul</h2>
      <span class="unit-badge">${unit}</span></div>
    <div class="chart-wrap tall"><canvas id="cmpMonthly"></canvas></div>`;
  container.appendChild(card3);

  const aMon = aggregateData(sData.filter(d => d.module === cmpModA), 'monthly');
  const bMon = aggregateData(sData.filter(d => d.module === cmpModB), 'monthly');
  const allMonths = [...new Set([...aMon.map(d => d.ts), ...bMon.map(d => d.ts)])].sort();
  const aByT = new Map(aMon.map(d => [d.ts, d.value]));
  const bByT = new Map(bMon.map(d => [d.ts, d.value]));

  cmpCharts.monthly = new Chart(document.getElementById('cmpMonthly').getContext('2d'), {
    type: 'bar',
    data: {
      labels: allMonths.map(t => fmtMonth(t)),
      datasets: [
        { label: cmpModA, data: allMonths.map(t => aByT.get(t) ?? null),
          backgroundColor: colA + 'cc', borderColor: colA, borderWidth: 1 },
        { label: cmpModB, data: allMonths.map(t => bByT.get(t) ?? null),
          backgroundColor: colB + 'cc', borderColor: colB, borderWidth: 1 },
      ],
    },
    options: {
      responsive: true, maintainAspectRatio: false,
      scales: {
        x: { grid: { color:'#30363d' }, ticks: { color:'#8b949e', maxTicksLimit: 12 } },
        y: { grid: { color:'#30363d' }, ticks: { color:'#8b949e',
             callback: v => v.toLocaleString('de-DE',{maximumFractionDigits:1}) + (unit?' '+unit:'') } },
      },
      plugins: {
        legend: { labels: { color:'#e6edf3', boxWidth: 12, padding: 10 } },
        tooltip: { backgroundColor: '#161b22', borderColor:'#30363d', borderWidth: 1,
                   titleColor: '#e6edf3', bodyColor: '#8b949e',
                   callbacks: { label: ctx => ` ${ctx.dataset.label}: ${ctx.parsed.y?.toLocaleString('de-DE',{maximumFractionDigits:2})} ${unit}` } },
      },
    },
  });
}
function heatColor(t) {
  const stops = [
    [0.00, [30, 58,138]], [0.25, [88,166,255]], [0.50, [240,228, 66]],
    [0.75, [245,158, 11]], [1.00, [220, 38, 38]],
  ];
  for (let i=0; i<stops.length-1; i++) {
    if (t <= stops[i+1][0]) {
      const a = stops[i], b = stops[i+1];
      const f = (t - a[0]) / (b[0] - a[0]);
      const c = a[1].map((v,j) => Math.round(v + (b[1][j]-v) * f));
      return `rgb(${c[0]},${c[1]},${c[2]})`;
    }
  }
  return 'rgb(220,38,38)';
}

function fmtDT(ts)    { return new Date(ts).toLocaleString('de-DE', {day:'2-digit',month:'2-digit',year:'numeric',hour:'2-digit',minute:'2-digit'}); }
function fmtDay(ts)   { return new Date(ts).toLocaleDateString('de-DE', {weekday:'short',day:'2-digit',month:'2-digit',year:'numeric'}); }
function fmtMonth(ts) { return new Date(ts).toLocaleDateString('de-DE', {month:'long', year:'numeric'}); }

function commonChartOpts(unit, withZoom) {
  const opts = {
    responsive: true, maintainAspectRatio: false,
    animation: { duration: 300 },
    interaction: { mode: 'index', intersect: false },
    scales: {
      x: { type: 'time',
           time: { tooltipFormat: 'dd.MM.yyyy HH:mm', displayFormats: { hour:'dd.MM HH:mm', day:'dd.MM.yy', week:'dd.MM.yy', month:'MMM yy', year:'yyyy' } },
           grid: { color: '#30363d' }, ticks: { color: '#8b949e', maxTicksLimit: 8 } },
      y: { grid: { color: '#30363d' }, ticks: { color: '#8b949e',
           callback: v => v.toLocaleString('de-DE',{maximumFractionDigits:1}) + (unit?' '+unit:'') } },
    },
    plugins: {
      legend: { labels: { color: '#e6edf3', boxWidth: 12, padding: 12 } },
      tooltip: { backgroundColor: '#161b22', borderColor: '#30363d', borderWidth: 1,
                 titleColor: '#e6edf3', bodyColor: '#8b949e',
                 callbacks: { label: ctx => ` ${ctx.dataset.label}: ${ctx.parsed.y?.toLocaleString('de-DE',{maximumFractionDigits:2})} ${unit}` } },
    },
  };
  if (withZoom) {
    opts.plugins.zoom = {
      zoom: { wheel: { enabled: true }, pinch: { enabled: true }, mode: 'x' },
      pan:  { enabled: true, mode: 'x' },
    };
  }
  return opts;
}

/* ── Climate change: warming trend, anomalies, stripes, decades, threshold days ── */
let climateCharts = {};
let climateSensor = null, climateModule = null;
let climateRefStart = null, climateRefEnd = null;
let climateMinMaxYears = null;  // Set<number> of years for daily min/max overlay (null = last 3)
let climateTropThreshold = 20;  // °C, slider value for tropical-nights chart (Tmin > threshold)
let climateHotThreshold  = 30;  // °C, slider value for hot-days chart (Tmax > threshold)

/* ── Extreme state ── */
let extremeCharts = {};
let extremeSensor = null, extremeModule = null;
let extremeHeatThr = 25;
let extremeColdThr = 0;

function renderClimate(raw) {
  const container = document.getElementById('climateContainer');
  Object.values(climateCharts).forEach(c => c.destroy());
  climateCharts = {};
  container.innerHTML = '';

  // Prefer temperature sensors (°C unit)
  const allSensors = [...new Set(raw.map(d => d.sensor))];
  const tempSensors = allSensors.filter(s => {
    const u = PAYLOAD.units[s] || '';
    return /°?\s*C/i.test(u) || /temp/i.test(s);
  });
  const sensors = tempSensors.length ? tempSensors : allSensors;
  if (!sensors.length) { container.innerHTML = '<div class="no-data">Keine Daten.</div>'; return; }
  if (!climateSensor || !sensors.includes(climateSensor)) climateSensor = sensors[0];

  const sData = raw.filter(d => d.sensor === climateSensor);
  const modules = ['(alle gewählten Module)', ...new Set(sData.map(d => d.module))].sort();
  if (!climateModule || !modules.includes(climateModule)) climateModule = modules[0];

  const mData = climateModule === '(alle gewählten Module)' ? sData : sData.filter(d => d.module === climateModule);
  const unit = PAYLOAD.units[climateSensor] || sData[0]?.unit || '';

  // Yearly means / mins / maxes
  const yearly = aggregateMinAvgMax(mData, 'yearly');
  // Collapse over multiple modules when "(alle ...)": average per year
  const yMap = {};
  yearly.forEach(r => {
    const y = new Date(r.ts).getFullYear();
    if (!yMap[y]) yMap[y] = { y, avgs:[], mins:[], maxs:[], ns:0 };
    yMap[y].avgs.push(r.avg); yMap[y].mins.push(r.min); yMap[y].maxs.push(r.max); yMap[y].ns += r.n;
  });
  const yearsArr = Object.values(yMap).map(y => ({
    year: y.y,
    avg: y.avgs.reduce((a,b)=>a+b,0)/y.avgs.length,
    min: Math.min(...y.mins),
    max: Math.max(...y.maxs),
    n:   y.ns,
  })).sort((a,b) => a.year - b.year);

  const coverageByYear = {};
  mData.forEach(d => {
    const dt = new Date(d.ts);
    const yr = dt.getFullYear();
    if (!coverageByYear[yr]) coverageByYear[yr] = { months: new Set(), days: new Set() };
    coverageByYear[yr].months.add(dt.getMonth());
    coverageByYear[yr].days.add(`${yr}-${dt.getMonth()}-${dt.getDate()}`);
  });
  const daysInYear = (yr) => new Date(yr, 1, 29).getMonth() === 1 ? 366 : 365;
  const maxCoveredDays = Math.max(0, ...Object.values(coverageByYear).map(c => c.days.size));
  const currentCalendarYear = new Date().getFullYear();
  const recordEligibleYears = new Set(yearsArr
    .filter(y => {
      const cov = coverageByYear[y.year];
      if (!cov || y.year >= currentCalendarYear || cov.months.size !== 12) return false;
      const minCoveredDays = Math.min(
        Math.floor(daysInYear(y.year) * 0.9),
        Math.floor(maxCoveredDays * 0.9)
      );
      return cov.days.size >= minCoveredDays;
    })
    .map(y => y.year));

  if (yearsArr.length < 2) {
    container.innerHTML = `<div class="no-data">Für Klima-Analysen werden mindestens 2 Jahre Daten benötigt. Aktuell: ${yearsArr.length} Jahr(e). Tipp: links den Zeitraum auf <b>„Alles"</b> stellen.</div>`;
    // still render controls so user can switch sensor
  }

  const minYr = yearsArr[0]?.year, maxYr = yearsArr[yearsArr.length-1]?.year;
  // Default reference: first 10 years (or first half if dataset shorter)
  if (climateRefStart === null || climateRefStart < minYr || climateRefStart > maxYr) {
    climateRefStart = minYr;
  }
  if (climateRefEnd === null || climateRefEnd < minYr || climateRefEnd > maxYr) {
    climateRefEnd = Math.min(maxYr, minYr + Math.max(0, Math.min(9, Math.floor(yearsArr.length/2) - 1)));
    if (climateRefEnd < climateRefStart) climateRefEnd = climateRefStart;
  }

  // ----- Controls -----
  const ctrl = document.createElement('div');
  ctrl.className = 'trend-controls';
  const yearOpts = (sel) => yearsArr.map(y => `<option value="${y.year}" ${y.year===sel?'selected':''}>${y.year}</option>`).join('');
  ctrl.innerHTML = `
    <label>Sensor:</label>
    <select id="climSensorSel">${sensors.map(s => `<option ${s===climateSensor?'selected':''}>${s}</option>`).join('')}</select>
    <label>Modul:</label>
    <select id="climModSel">${modules.map(m => `<option ${m===climateModule?'selected':''}>${m}</option>`).join('')}</select>
    <label>Referenzperiode:</label>
    <select id="climRefStart">${yearOpts(climateRefStart)}</select>
    <span style="color:var(--muted);font-size:12px;">bis</span>
    <select id="climRefEnd">${yearOpts(climateRefEnd)}</select>`;
  container.appendChild(ctrl);
  ctrl.querySelector('#climSensorSel').onchange = e => { climateSensor = e.target.value; climateModule = null; climateRefStart = climateRefEnd = null; renderClimate(raw); };
  ctrl.querySelector('#climModSel').onchange    = e => { climateModule = e.target.value; renderClimate(raw); };
  ctrl.querySelector('#climRefStart').onchange  = e => { climateRefStart = +e.target.value; if (climateRefEnd < climateRefStart) climateRefEnd = climateRefStart; renderClimate(raw); };
  ctrl.querySelector('#climRefEnd').onchange    = e => { climateRefEnd   = +e.target.value; if (climateRefStart > climateRefEnd) climateRefStart = climateRefEnd; renderClimate(raw); };

  if (yearsArr.length < 2) return;

  // Reference mean
  const refYears = yearsArr.filter(y => y.year >= climateRefStart && y.year <= climateRefEnd);
  const refMean = refYears.reduce((a,b) => a + b.avg, 0) / (refYears.length || 1);

  // Linear regression on yearly means (x = year as days since first year)
  const regPts = yearsArr.map(y => ({ ts: new Date(y.year, 6, 1).getTime(), value: y.avg }));
  const reg = linearRegression(regPts);
  const slopePerDecade = reg ? reg.slopePerYear * 10 : 0;
  const totalChange    = reg ? reg.slopePerYear * (maxYr - minYr) : 0;

  // Anomalies
  const anomalies = yearsArr.map(y => ({ year: y.year, anom: y.avg - refMean, avg: y.avg }));
  const anomMin = Math.min(...anomalies.map(a => a.anom));
  const anomMax = Math.max(...anomalies.map(a => a.anom));
  const anomAbs = Math.max(Math.abs(anomMin), Math.abs(anomMax)) || 1;

  // Color helpers
  const stripeColor = (anom) => {
    // -1..0 blue → white, 0..+1 white → red (Ed Hawkins style)
    const t = Math.max(-1, Math.min(1, anom / anomAbs));
    if (t >= 0) {
      // white (255,245,235) → dark red (165,15,21)
      const r = Math.round(255 + (165-255)*t);
      const g = Math.round(245 + ( 15-245)*t);
      const b = Math.round(235 + ( 21-235)*t);
      return `rgb(${r},${g},${b})`;
    } else {
      const k = -t;
      const r = Math.round(255 + ( 33-255)*k);
      const g = Math.round(245 + (102-245)*k);
      const b = Math.round(235 + (172-235)*k);
      return `rgb(${r},${g},${b})`;
    }
  };

  // ----- Stats pills -----
  const info = document.createElement('div');
  info.className = 'trend-info';
  const upDown = slopePerDecade >= 0 ? 'up' : 'down';
  const recordYearsArr = yearsArr.filter(y => recordEligibleYears.has(y.year));
  const warmestRecordYear = recordYearsArr.length ? recordYearsArr.reduce((a,b)=>b.avg>a.avg?b:a) : null;
  const coldestRecordYear = recordYearsArr.length ? recordYearsArr.reduce((a,b)=>b.avg<a.avg?b:a) : null;
  info.innerHTML = `
    <span class="pill">📅 Zeitraum: <b>${minYr} – ${maxYr}</b> (${yearsArr.length} Jahre)</span>
    <span class="pill">📐 Referenz Ø: <b>${fmt(refMean, unit)}</b> <span style="opacity:.6">(${climateRefStart}–${climateRefEnd})</span></span>
    <span class="pill ${upDown}">📈 Trend: <b>${(slopePerDecade>=0?'+':'')}${slopePerDecade.toLocaleString('de-DE',{maximumFractionDigits:2})} ${unit}/Dekade</b></span>
    <span class="pill ${upDown}">Σ ${minYr}→${maxYr}: <b>${(totalChange>=0?'+':'')}${totalChange.toLocaleString('de-DE',{maximumFractionDigits:2})} ${unit}</b></span>
    ${warmestRecordYear ? `<span class="pill">🔥 wärmstes: <b>${warmestRecordYear.year}</b> (${fmt(warmestRecordYear.avg, unit)})</span>` : ''}
    ${coldestRecordYear ? `<span class="pill">❄️ kältestes: <b>${coldestRecordYear.year}</b> (${fmt(coldestRecordYear.avg, unit)})</span>` : ''}`;
  container.appendChild(info);

  // ----- Card 1: Warming Stripes (Ed Hawkins style) -----
  const card1 = document.createElement('div');
  card1.className = 'chart-card';
  card1.innerHTML = `
    <div class="chart-header"><h2>🌡️ Erwärmungsstreifen (Warming Stripes)</h2>
      <span class="unit-badge">${unit}</span></div>
    <div class="chart-wrap"><canvas id="climStripes"></canvas></div>
    <div class="chart-hint">Jeder Streifen = 1 Jahr. Blau = kühler, Rot = wärmer als Ø ${climateRefStart}–${climateRefEnd}. Inspiriert von Ed Hawkins (#ShowYourStripes).</div>`;
  container.appendChild(card1);

  climateCharts.stripes = new Chart(document.getElementById('climStripes').getContext('2d'), {
    type: 'bar',
    data: {
      labels: yearsArr.map(y => y.year),
      datasets: [{
        label: 'Anomalie',
        data: yearsArr.map(() => 1),
        backgroundColor: anomalies.map(a => stripeColor(a.anom)),
        borderWidth: 0, barPercentage: 1.0, categoryPercentage: 1.0,
      }],
    },
    options: {
      responsive: true, maintainAspectRatio: false,
      scales: {
        x: { grid: { display: false }, ticks: { color: '#8b949e', maxRotation: 0, autoSkipPadding: 20 } },
        y: { display: false, min: 0, max: 1 },
      },
      plugins: {
        legend: { display: false },
        tooltip: {
          backgroundColor: '#161b22', borderColor: '#30363d', borderWidth: 1,
          titleColor: '#e6edf3', bodyColor: '#8b949e',
          callbacks: {
            title: items => `Jahr ${items[0].label}`,
            label: ctx => {
              const a = anomalies[ctx.dataIndex];
              const sign = a.anom >= 0 ? '+' : '';
              return [` Ø ${fmt(a.avg, unit)}`,
                      ` Anomalie: ${sign}${a.anom.toLocaleString('de-DE',{maximumFractionDigits:2})} ${unit}`];
            },
          },
        },
      },
    },
  });

  // ----- Card 2: Anomalies bar chart with trend line -----
  const card2 = document.createElement('div');
  card2.className = 'chart-card';
  card2.innerHTML = `
    <div class="chart-header"><h2>📊 Jahres-Anomalien (relativ zu ${climateRefStart}–${climateRefEnd})</h2>
      <span class="unit-badge">${unit}</span></div>
    <div class="chart-wrap tall"><canvas id="climAnom"></canvas></div>
    <div class="chart-hint">Balken = Abweichung des Jahresmittels von der Referenzperiode. Linie = lineare Regression (${slopePerDecade>=0?'+':''}${slopePerDecade.toLocaleString('de-DE',{maximumFractionDigits:2})} ${unit}/Dekade).</div>`;
  container.appendChild(card2);

  // Build regression line over years
  const regLine = yearsArr.map(y => {
    const t = new Date(y.year, 6, 1).getTime();
    return reg ? (reg.intercept + reg.slope * (t - reg.t0) / 86400000) - refMean : 0;
  });

  climateCharts.anom = new Chart(document.getElementById('climAnom').getContext('2d'), {
    type: 'bar',
    data: {
      labels: yearsArr.map(y => y.year),
      datasets: [
        {
          label: 'Anomalie',
          data: anomalies.map(a => Math.round(a.anom*100)/100),
          backgroundColor: anomalies.map(a => a.anom >= 0 ? '#f8514999' : '#58a6ff99'),
          borderColor:     anomalies.map(a => a.anom >= 0 ? '#f85149'   : '#58a6ff'),
          borderWidth: 1, order: 2,
        },
        {
          label: 'Linearer Trend',
          type: 'line', data: regLine,
          borderColor: '#d29922', borderWidth: 2.5, pointRadius: 0,
          tension: 0, fill: false, order: 1,
        },
      ],
    },
    options: {
      responsive: true, maintainAspectRatio: false,
      scales: {
        x: { grid: { color: '#30363d' }, ticks: { color: '#8b949e' } },
        y: { grid: { color: '#30363d' }, ticks: { color: '#8b949e',
             callback: v => (v>=0?'+':'') + v.toLocaleString('de-DE',{maximumFractionDigits:1}) + (unit?' '+unit:'') },
             title: { display: true, text: `Δ vs. Ø ${climateRefStart}–${climateRefEnd}`, color: '#8b949e' } },
      },
      plugins: {
        legend: { labels: { color: '#e6edf3', boxWidth: 12, padding: 10 } },
        tooltip: { backgroundColor: '#161b22', borderColor: '#30363d', borderWidth: 1,
                   titleColor: '#e6edf3', bodyColor: '#8b949e',
                   callbacks: { label: ctx => ` ${ctx.dataset.label}: ${(ctx.parsed.y>=0?'+':'')}${ctx.parsed.y?.toLocaleString('de-DE',{maximumFractionDigits:2})} ${unit}` } },
      },
    },
  });

  // ----- Card 3: Yearly mean with moving averages -----
  const card3 = document.createElement('div');
  card3.className = 'chart-card';
  card3.innerHTML = `
    <div class="chart-header"><h2>📈 Jahresmittel mit gleitenden Mittelwerten</h2>
      <span class="unit-badge">${unit}</span></div>
    <div class="chart-wrap tall"><canvas id="climLine"></canvas></div>
    <div class="chart-hint">Rohwerte (dünn) glätten sich durch 5- und 10-Jahres-Mittel (dick), die langfristige Trends sichtbar machen.</div>`;
  container.appendChild(card3);

  const movAvg = (arr, w) => arr.map((_, i) => {
    const s = Math.max(0, i - Math.floor(w/2));
    const e = Math.min(arr.length, i + Math.ceil(w/2));
    const slice = arr.slice(s, e);
    return slice.reduce((a,b)=>a+b,0)/slice.length;
  });
  const yAvgs = yearsArr.map(y => y.avg);
  const ma5  = yearsArr.length >= 5  ? movAvg(yAvgs, 5)  : null;
  const ma10 = yearsArr.length >= 10 ? movAvg(yAvgs, 10) : null;

  // Regression line in absolute units
  const regAbsLine = reg ? yearsArr.map(y => {
    const t = new Date(y.year, 6, 1).getTime();
    return reg.intercept + reg.slope * (t - reg.t0) / 86400000;
  }) : null;

  const lineDatasets = [
    { label: 'Jahresmittel', data: yAvgs, borderColor: '#8b949e',
      backgroundColor: 'transparent', borderWidth: 1.2, pointRadius: 2, tension: 0.2 },
  ];
  if (ma5)  lineDatasets.push({ label: '5-Jahres-Mittel',  data: ma5,
            borderColor: '#3fb950', backgroundColor: 'transparent', borderWidth: 2.5, pointRadius: 0, tension: 0.3 });
  if (ma10) lineDatasets.push({ label: '10-Jahres-Mittel', data: ma10,
            borderColor: '#bc8cff', backgroundColor: 'transparent', borderWidth: 2.5, pointRadius: 0, tension: 0.3 });
  if (regAbsLine) lineDatasets.push({ label: 'Linearer Trend', data: regAbsLine,
            borderColor: '#d29922', backgroundColor: 'transparent', borderWidth: 2, borderDash: [6,4], pointRadius: 0, tension: 0 });

  climateCharts.line = new Chart(document.getElementById('climLine').getContext('2d'), {
    type: 'line',
    data: { labels: yearsArr.map(y => y.year), datasets: lineDatasets },
    options: {
      responsive: true, maintainAspectRatio: false,
      interaction: { mode: 'index', intersect: false },
      scales: {
        x: { grid: { color: '#30363d' }, ticks: { color: '#8b949e' } },
        y: { grid: { color: '#30363d' }, ticks: { color: '#8b949e',
             callback: v => v.toLocaleString('de-DE',{maximumFractionDigits:1}) + (unit?' '+unit:'') } },
      },
      plugins: {
        legend: { labels: { color: '#e6edf3', boxWidth: 12, padding: 10 } },
        tooltip: { backgroundColor: '#161b22', borderColor: '#30363d', borderWidth: 1,
                   titleColor: '#e6edf3', bodyColor: '#8b949e',
                   callbacks: { label: ctx => ` ${ctx.dataset.label}: ${ctx.parsed.y?.toLocaleString('de-DE',{maximumFractionDigits:2})} ${unit}` } },
      },
    },
  });

  // ----- Card 4: Decade comparison -----
  const card4 = document.createElement('div');
  card4.className = 'chart-card';
  card4.innerHTML = `
    <div class="chart-header"><h2>🗓️ Dekaden-Vergleich</h2>
      <span class="unit-badge">${unit}</span></div>
    <div class="chart-wrap"><canvas id="climDecade"></canvas></div>
    <div class="chart-hint">Mittelwert pro Dekade (10 Jahre). Glättet kurzfristige Schwankungen und zeigt strukturelle Verschiebungen.</div>`;
  container.appendChild(card4);

  const decades = {};
  yearsArr.forEach(y => {
    const d = Math.floor(y.year / 10) * 10;
    if (!decades[d]) decades[d] = { sum:0, n:0, min:Infinity, max:-Infinity };
    decades[d].sum += y.avg; decades[d].n += 1;
    if (y.min < decades[d].min) decades[d].min = y.min;
    if (y.max > decades[d].max) decades[d].max = y.max;
  });
  const decKeys = Object.keys(decades).map(Number).sort();
  const decMeans = decKeys.map(k => decades[k].sum / decades[k].n);
  const decMins  = decKeys.map(k => decades[k].min);
  const decMaxs  = decKeys.map(k => decades[k].max);
  // color decades by warming
  const decMinV = Math.min(...decMeans), decMaxV = Math.max(...decMeans);
  const decColors = decMeans.map(v => {
    const t = (v - decMinV) / (decMaxV - decMinV || 1);
    return stripeColor((v - refMean) / (anomAbs || 1) * anomAbs);
  });

  climateCharts.decade = new Chart(document.getElementById('climDecade').getContext('2d'), {
    type: 'bar',
    data: {
      labels: decKeys.map(k => k + 'er'),
      datasets: [
        { label: 'Min',  data: decMins,  backgroundColor: '#58a6ff66', borderColor:'#58a6ff', borderWidth: 1 },
        { label: 'Ø',    data: decMeans, backgroundColor: decColors,   borderColor:'#e6edf3', borderWidth: 1 },
        { label: 'Max',  data: decMaxs,  backgroundColor: '#f8514966', borderColor:'#f85149', borderWidth: 1 },
      ],
    },
    options: {
      responsive: true, maintainAspectRatio: false,
      scales: {
        x: { grid: { color: '#30363d' }, ticks: { color: '#8b949e' } },
        y: { grid: { color: '#30363d' }, ticks: { color: '#8b949e',
             callback: v => v.toLocaleString('de-DE',{maximumFractionDigits:1}) + (unit?' '+unit:'') } },
      },
      plugins: {
        legend: { labels: { color: '#e6edf3', boxWidth: 12, padding: 10 } },
        tooltip: { backgroundColor: '#161b22', borderColor: '#30363d', borderWidth: 1,
                   titleColor: '#e6edf3', bodyColor: '#8b949e',
                   callbacks: { label: ctx => ` ${ctx.dataset.label}: ${ctx.parsed.y?.toLocaleString('de-DE',{maximumFractionDigits:2})} ${unit}` } },
      },
    },
  });

  // ----- Card 5: Threshold days per year (only meaningful for temperature in °C) -----
  if (/°?\s*C/i.test(unit) || /temp/i.test(climateSensor)) {
    const card5 = document.createElement('div');
    card5.className = 'chart-card';
    card5.innerHTML = `
      <div class="chart-header"><h2>🌞 Schwellwert-Tage pro Jahr</h2>
        <span class="unit-badge">Tage</span></div>
      <div class="chart-wrap tall"><canvas id="climThresh"></canvas></div>
      <div class="chart-hint">Heiße Tage (Tmax ≥ 30 °C), Sommertage (≥ 25 °C), Tropennächte (Tmin ≥ 20 °C), Frosttage (Tmin < 0 °C), Eistage (Tmax < 0 °C). Berechnet aus Tagesmin/-max je Jahr.</div>`;
    container.appendChild(card5);

    // Daily aggregation across modules: per day take min/max across all included modules
    const daily = aggregateMinAvgMax(mData, 'daily');
    const dayAgg = {};
    daily.forEach(r => {
      const d = new Date(r.ts);
      const dk = `${d.getFullYear()}-${d.getMonth()}-${d.getDate()}`;
      if (!dayAgg[dk]) dayAgg[dk] = { year: d.getFullYear(), min: r.min, max: r.max };
      else {
        if (r.min < dayAgg[dk].min) dayAgg[dk].min = r.min;
        if (r.max > dayAgg[dk].max) dayAgg[dk].max = r.max;
      }
    });
    const byYr = {};
    Object.values(dayAgg).forEach(d => {
      if (!byYr[d.year]) byYr[d.year] = { hot:0, summer:0, tropical:0, frost:0, ice:0 };
      if (d.max >= 30) byYr[d.year].hot++;
      if (d.max >= 25) byYr[d.year].summer++;
      if (d.min >= 20) byYr[d.year].tropical++;
      if (d.min <  0)  byYr[d.year].frost++;
      if (d.max <  0)  byYr[d.year].ice++;
    });
    const thrYears = Object.keys(byYr).map(Number).sort();
    const mkDs = (key, color, label) => ({
      label, data: thrYears.map(y => byYr[y][key]),
      borderColor: color, backgroundColor: color + '22',
      borderWidth: 2, pointRadius: 2, tension: 0.3, fill: false,
    });

    climateCharts.thresh = new Chart(document.getElementById('climThresh').getContext('2d'), {
      type: 'line',
      data: {
        labels: thrYears,
        datasets: [
          mkDs('hot',      '#f85149', 'Heiße Tage (≥30 °C)'),
          mkDs('summer',   '#d29922', 'Sommertage (≥25 °C)'),
          mkDs('tropical', '#bc8cff', 'Tropennächte (Tmin≥20 °C)'),
          mkDs('frost',    '#58a6ff', 'Frosttage (Tmin<0 °C)'),
          mkDs('ice',      '#39c5cf', 'Eistage (Tmax<0 °C)'),
        ],
      },
      options: {
        responsive: true, maintainAspectRatio: false,
        interaction: { mode: 'index', intersect: false },
        scales: {
          x: { grid: { color: '#30363d' }, ticks: { color: '#8b949e' } },
          y: { grid: { color: '#30363d' }, ticks: { color: '#8b949e' }, beginAtZero: true,
               title: { display: true, text: 'Tage / Jahr', color: '#8b949e' } },
        },
        plugins: {
          legend: { labels: { color: '#e6edf3', boxWidth: 12, padding: 10 } },
          tooltip: { backgroundColor: '#161b22', borderColor: '#30363d', borderWidth: 1,
                     titleColor: '#e6edf3', bodyColor: '#8b949e' },
        },
      },
    });
  }

  // ----- Card 6: Monthly trend grid (warming per calendar month) -----
  if (yearsArr.length >= 3) {
    const card6 = document.createElement('div');
    card6.className = 'chart-card';
    card6.innerHTML = `
      <div class="chart-header"><h2>📆 Erwärmung nach Kalendermonat</h2>
        <span class="unit-badge">${unit}/Dekade</span></div>
      <div class="chart-wrap"><canvas id="climMonthTrend"></canvas></div>
      <div class="chart-hint">Linearer Trend (°C pro Dekade) separat für jeden Kalendermonat – zeigt, ob z.B. Winter oder Sommer stärker aufheizen.</div>`;
    container.appendChild(card6);

    // Monthly mean per (year, month)
    const monthly = aggregateData(mData, 'monthly');
    const byYM = {};
    monthly.forEach(r => {
      const dt = new Date(r.ts);
      const y = dt.getFullYear(), m = dt.getMonth();
      const k = `${y}-${m}`;
      if (!byYM[k]) byYM[k] = { y, m, vals: [] };
      byYM[k].vals.push(r.value);
    });
    const monthSlopes = [];
    for (let m = 0; m < 12; m++) {
      const pts = Object.values(byYM)
        .filter(o => o.m === m)
        .map(o => ({ ts: new Date(o.y, m, 1).getTime(), value: o.vals.reduce((a,b)=>a+b,0)/o.vals.length }))
        .sort((a,b) => a.ts - b.ts);
      const r = linearRegression(pts);
      monthSlopes.push(r ? r.slopePerYear * 10 : 0);
    }
    const monthNames = ['Jan','Feb','Mär','Apr','Mai','Jun','Jul','Aug','Sep','Okt','Nov','Dez'];
    const msAbs = Math.max(...monthSlopes.map(Math.abs)) || 1;
    const monthColors = monthSlopes.map(s => {
      const t = s / msAbs;
      return t >= 0 ? `rgba(248,81,73,${0.4 + 0.6*t})` : `rgba(88,166,255,${0.4 + 0.6*(-t)})`;
    });

    climateCharts.monthTrend = new Chart(document.getElementById('climMonthTrend').getContext('2d'), {
      type: 'bar',
      data: {
        labels: monthNames,
        datasets: [{
          label: 'Trend',
          data: monthSlopes.map(s => Math.round(s*100)/100),
          backgroundColor: monthColors,
          borderColor: monthColors.map(c => c.replace(/[\d.]+\)$/, '1)')),
          borderWidth: 1,
        }],
      },
      options: {
        responsive: true, maintainAspectRatio: false,
        scales: {
          x: { grid: { color: '#30363d' }, ticks: { color: '#8b949e' } },
          y: { grid: { color: '#30363d' }, ticks: { color: '#8b949e',
               callback: v => (v>=0?'+':'') + v.toLocaleString('de-DE',{maximumFractionDigits:2}) + ' ' + unit + '/Dek.' } },
        },
        plugins: {
          legend: { display: false },
          tooltip: { backgroundColor: '#161b22', borderColor: '#30363d', borderWidth: 1,
                     titleColor: '#e6edf3', bodyColor: '#8b949e',
                     callbacks: { label: ctx => ` ${(ctx.parsed.y>=0?'+':'')}${ctx.parsed.y.toLocaleString('de-DE',{maximumFractionDigits:2})} ${unit}/Dekade` } },
        },
      },
    });
  }

  // ----- Card 7: Daily Min/Max envelopes per selected year (DOY x-axis) -----
  {
    // Daily min/max per (year, day-of-year using synthetic 2000 base)
    const daily = aggregateMinAvgMax(mData, 'daily');
    const dayBy = {};  // {year: {synthTs: {min, max}}}
    daily.forEach(r => {
      const d = new Date(r.ts);
      const y = d.getFullYear();
      const synth = new Date(2000, d.getMonth(), d.getDate()).getTime();
      if (!dayBy[y]) dayBy[y] = {};
      const e = dayBy[y][synth];
      if (!e) dayBy[y][synth] = { min: r.min, max: r.max };
      else {
        if (r.min < e.min) e.min = r.min;
        if (r.max > e.max) e.max = r.max;
      }
    });
    const mmYears = Object.keys(dayBy).map(Number).sort();

    // Initialise / sync year selection
    if (!climateMinMaxYears) {
      climateMinMaxYears = new Set(mmYears.slice(-3));  // default: last 3 years
    } else {
      [...climateMinMaxYears].forEach(y => { if (!mmYears.includes(y)) climateMinMaxYears.delete(y); });
      if (!climateMinMaxYears.size && mmYears.length) climateMinMaxYears.add(mmYears[mmYears.length-1]);
    }

    const palette = ['#58a6ff','#3fb950','#d29922','#f85149','#bc8cff','#ff7b72','#39c5cf','#ffa657','#7ee787','#f0883e'];
    const mmYearColor = {};
    mmYears.forEach((y, i) => { mmYearColor[y] = palette[i % palette.length]; });

    const card7 = document.createElement('div');
    card7.className = 'chart-card';
    card7.innerHTML = `
      <div class="chart-header"><h2>🌡️ Tagesextrema im Jahresverlauf (Min/Max pro Tag)</h2>
        <span class="unit-badge">${unit}</span></div>
      <div id="climMinMaxYearChips" class="year-chips"></div>
      <div class="chart-wrap tall"><canvas id="climMinMax"></canvas></div>
      <div class="chart-hint">Pro Jahr ein farbiges Band zwischen Tagesminimum und Tagesmaximum. X-Achse: Tag im Jahr. Jahre per Klick überlagern und vergleichen.</div>`;
    container.appendChild(card7);

    const chipsEl = document.getElementById('climMinMaxYearChips');
    const renderChips = () => {
      chipsEl.innerHTML =
        `<span class="label">Jahre:</span>` +
        mmYears.map(y => {
          const active = climateMinMaxYears.has(y);
          return `<button class="year-chip${active?' active':''}" data-year="${y}">`
               + `<span class="swatch" style="background:${mmYearColor[y]}"></span>${y}</button>`;
        }).join('') +
        `<button class="link-btn" data-act="all">Alle</button>` +
        `<button class="link-btn" data-act="last2">Letzte 2</button>` +
        `<button class="link-btn" data-act="last3">Letzte 3</button>` +
        `<button class="link-btn" data-act="last5">Letzte 5</button>` +
        `<button class="link-btn" data-act="oldnew">Ältestes + Neuestes</button>`;
    };
    renderChips();
    chipsEl.addEventListener('click', e => {
      const chip = e.target.closest('.year-chip');
      const link = e.target.closest('.link-btn');
      if (chip) {
        const y = +chip.dataset.year;
        if (climateMinMaxYears.has(y)) climateMinMaxYears.delete(y); else climateMinMaxYears.add(y);
        if (!climateMinMaxYears.size) climateMinMaxYears.add(y);
      } else if (link) {
        const act = link.dataset.act;
        if (act === 'all')    climateMinMaxYears = new Set(mmYears);
        if (act === 'last2')  climateMinMaxYears = new Set(mmYears.slice(-2));
        if (act === 'last3')  climateMinMaxYears = new Set(mmYears.slice(-3));
        if (act === 'last5')  climateMinMaxYears = new Set(mmYears.slice(-5));
        if (act === 'oldnew') climateMinMaxYears = new Set([mmYears[0], mmYears[mmYears.length-1]]);
      } else return;
      renderChips();
      rebuildMinMaxChart();
    });

    const hexToRgba = (hex, a) => {
      const h = hex.replace('#','');
      const r = parseInt(h.substring(0,2),16), g = parseInt(h.substring(2,4),16), b = parseInt(h.substring(4,6),16);
      return `rgba(${r},${g},${b},${a})`;
    };

    const buildMinMaxDatasets = () => {
      const ds = [];
      mmYears.filter(y => climateMinMaxYears.has(y)).forEach(y => {
        const color = mmYearColor[y];
        const days = Object.entries(dayBy[y])
          .map(([x, v]) => ({ x: +x, min: v.min, max: v.max }))
          .sort((a,b) => a.x - b.x);
        // Min line first (no fill), then Max with fill back to previous dataset (Min)
        ds.push({
          label: `${y} Min`,
          data: days.map(d => ({ x: d.x, y: d.min })),
          borderColor: color, backgroundColor: 'transparent',
          borderWidth: 1.2, pointRadius: 0, tension: 0.25,
          fill: false, _year: y, _kind: 'min',
        });
        ds.push({
          label: `${y} Max`,
          data: days.map(d => ({ x: d.x, y: d.max })),
          borderColor: color, backgroundColor: hexToRgba(color, 0.18),
          borderWidth: 1.2, pointRadius: 0, tension: 0.25,
          fill: '-1', _year: y, _kind: 'max',
        });
      });
      return ds;
    };

    const rebuildMinMaxChart = () => {
      if (climateCharts.minmax) {
        climateCharts.minmax.data.datasets = buildMinMaxDatasets();
        climateCharts.minmax.update();
      }
    };

    climateCharts.minmax = new Chart(document.getElementById('climMinMax').getContext('2d'), {
      type: 'line',
      data: { datasets: buildMinMaxDatasets() },
      options: {
        responsive: true, maintainAspectRatio: false,
        interaction: { mode: 'index', intersect: false },
        scales: {
          x: { type: 'time',
               time: { tooltipFormat: 'dd.MM', displayFormats: { day:'dd.MM', month:'MMM' }, unit: 'month' },
               grid: { color: '#30363d' }, ticks: { color: '#8b949e' } },
          y: { grid: { color: '#30363d' }, ticks: { color: '#8b949e',
               callback: v => v.toLocaleString('de-DE',{maximumFractionDigits:1}) + (unit?' '+unit:'') } },
        },
        plugins: {
          legend: { labels: { color: '#e6edf3', boxWidth: 12, padding: 10,
            filter: it => it.text.endsWith(' Max'),
            generateLabels: (ch) => {
              const seen = new Set();
              return ch.data.datasets.flatMap((d, i) => {
                if (d._kind !== 'max' || seen.has(d._year)) return [];
                seen.add(d._year);
                return [{ text: String(d._year), fillStyle: d.borderColor, strokeStyle: d.borderColor,
                          lineWidth: 1, hidden: false, datasetIndex: i }];
              });
            },
          } },
          tooltip: { backgroundColor: '#161b22', borderColor: '#30363d', borderWidth: 1,
                     titleColor: '#e6edf3', bodyColor: '#8b949e',
                     callbacks: { label: ctx => ` ${ctx.dataset.label}: ${ctx.parsed.y?.toLocaleString('de-DE',{maximumFractionDigits:2})} ${unit}` } },
          zoom: {
            zoom: { wheel: { enabled: true }, pinch: { enabled: true }, mode: 'x' },
            pan:  { enabled: true, mode: 'x' },
          },
        },
      },
    });
  }

  // ----- Card 8: Jahres-Spanne (Min/Max/Ø) mit Gefrierpunkt-Linie -----
  {
    const card8 = document.createElement('div');
    card8.className = 'chart-card';
    const isTempUnit = /°?\s*C/i.test(unit) || /temp/i.test(climateSensor);
    card8.innerHTML = `
      <div class="chart-header"><h2>📏 Jahres-Spanne: Min / Max / Ø${isTempUnit ? ' (mit Gefrierpunkt)' : ''}</h2>
        <span class="unit-badge">${unit}</span></div>
      <div class="chart-wrap tall"><canvas id="climSpan"></canvas></div>
      <div class="chart-hint">Pro Jahr: Min und Max als Linien (Band dazwischen = Spanne), Jahresmittel als kräftige Linie.${isTempUnit ? ' Gestrichelte Linie bei 0 °C markiert den Gefrierpunkt.' : ''}</div>`;
    container.appendChild(card8);

    const labels = yearsArr.map(y => y.year);
    const datasetsSpan = [
      { label: 'Min',  data: yearsArr.map(y => Math.round(y.min*100)/100),
        borderColor: '#58a6ff', backgroundColor: 'transparent',
        borderWidth: 1.5, pointRadius: 2, tension: 0.25, fill: false, order: 2 },
      { label: 'Max',  data: yearsArr.map(y => Math.round(y.max*100)/100),
        borderColor: '#f85149', backgroundColor: 'rgba(248,81,73,0.12)',
        borderWidth: 1.5, pointRadius: 2, tension: 0.25, fill: '-1', order: 2 },
      { label: 'Ø',    data: yearsArr.map(y => Math.round(y.avg*100)/100),
        borderColor: '#d29922', backgroundColor: 'transparent',
        borderWidth: 2.5, pointRadius: 3, tension: 0.25, fill: false, order: 1 },
    ];
    if (isTempUnit) {
      datasetsSpan.push({
        label: 'Gefrierpunkt (0 °C)',
        data: yearsArr.map(() => 0),
        borderColor: '#39c5cf', backgroundColor: 'transparent',
        borderWidth: 1.5, pointRadius: 0, borderDash: [6, 4], tension: 0, fill: false, order: 0,
      });
    }

    // Spanne-Pills (Min/Max/Ø/Range)
    const spanPills = document.createElement('div');
    spanPills.className = 'trend-info';
    const yMinAll = Math.min(...yearsArr.map(y => y.min));
    const yMaxAll = Math.max(...yearsArr.map(y => y.max));
    const rangeAll = yMaxAll - yMinAll;
    const minRangeYr = yearsArr.reduce((a,b) => (b.max-b.min) < (a.max-a.min) ? b : a);
    const maxRangeYr = yearsArr.reduce((a,b) => (b.max-b.min) > (a.max-a.min) ? b : a);
    spanPills.innerHTML = `
      <span class="pill">❄️ Absolut-Min: <b>${fmt(yMinAll, unit)}</b></span>
      <span class="pill">🔥 Absolut-Max: <b>${fmt(yMaxAll, unit)}</b></span>
      <span class="pill">📏 Gesamt-Spanne: <b>${fmt(rangeAll, unit)}</b></span>
      <span class="pill">🪶 kleinste Jahres-Spanne: <b>${minRangeYr.year}</b> (${fmt(minRangeYr.max - minRangeYr.min, unit)})</span>
      <span class="pill">⚡ größte Jahres-Spanne: <b>${maxRangeYr.year}</b> (${fmt(maxRangeYr.max - maxRangeYr.min, unit)})</span>`;
    card8.insertBefore(spanPills, card8.querySelector('.chart-wrap'));

    climateCharts.span = new Chart(document.getElementById('climSpan').getContext('2d'), {
      type: 'line',
      data: { labels, datasets: datasetsSpan },
      options: {
        responsive: true, maintainAspectRatio: false,
        interaction: { mode: 'index', intersect: false },
        scales: {
          x: { grid: { color: '#30363d' }, ticks: { color: '#8b949e' } },
          y: { grid: { color: '#30363d' }, ticks: { color: '#8b949e',
               callback: v => v.toLocaleString('de-DE',{maximumFractionDigits:1}) + (unit?' '+unit:'') } },
        },
        plugins: {
          legend: { labels: { color: '#e6edf3', boxWidth: 12, padding: 10 } },
          tooltip: { backgroundColor: '#161b22', borderColor: '#30363d', borderWidth: 1,
                     titleColor: '#e6edf3', bodyColor: '#8b949e',
                     callbacks: { label: ctx => ` ${ctx.dataset.label}: ${ctx.parsed.y?.toLocaleString('de-DE',{maximumFractionDigits:2})} ${unit}` } },
        },
      },
    });
  }

  // ----- Card 9: Anzahl Tage je Temperatur-Schwelle pro Jahr -----
  if (/°?\s*C/i.test(unit) || /temp/i.test(climateSensor)) {
    const card9 = document.createElement('div');
    card9.className = 'chart-card';
    card9.innerHTML = `
      <div class="chart-header"><h2>🌡️ Tage pro Jahr unter / über Schwellwerten</h2>
        <span class="unit-badge">Tage</span></div>
      <div class="chart-wrap tall"><canvas id="climThr2"></canvas></div>
      <div class="chart-hint">Kalte Schwellen werden auf <b>Tagesminimum</b> angewendet, warme auf <b>Tagesmaximum</b>. Klick in die Legende blendet einzelne Schwellen aus.</div>`;
    container.appendChild(card9);

    // Daily min/max collapsed across modules in mData
    const daily9 = aggregateMinAvgMax(mData, 'daily');
    const dayAgg9 = {};
    daily9.forEach(r => {
      const d = new Date(r.ts);
      const dk = `${d.getFullYear()}-${d.getMonth()}-${d.getDate()}`;
      if (!dayAgg9[dk]) dayAgg9[dk] = { year: d.getFullYear(), min: r.min, max: r.max };
      else {
        if (r.min < dayAgg9[dk].min) dayAgg9[dk].min = r.min;
        if (r.max > dayAgg9[dk].max) dayAgg9[dk].max = r.max;
      }
    });

    const coldThresh = [-5, 0, 5, 10];  // Tmin < t
    const warmThresh = [20, 25, 30, 35]; // Tmax >= t
    const counts = {};  // {year: {[label]: count}}
    Object.values(dayAgg9).forEach(d => {
      if (!counts[d.year]) counts[d.year] = {};
      coldThresh.forEach(t => {
        const k = `< ${t} °C`;
        counts[d.year][k] = (counts[d.year][k] || 0) + (d.min < t ? 1 : 0);
      });
      warmThresh.forEach(t => {
        const k = `≥ ${t} °C`;
        counts[d.year][k] = (counts[d.year][k] || 0) + (d.max >= t ? 1 : 0);
      });
    });
    const thrYears9 = Object.keys(counts).map(Number).sort();

    // Color gradients: cold = blue→cyan, warm = orange→deep red
    const coldColors = ['#1e3a8a', '#1f6feb', '#58a6ff', '#79c0ff'];
    const warmColors = ['#fbbf24', '#f59e0b', '#f85149', '#7f1d1d'];

    const mkDs9 = (label, color, key) => ({
      label, data: thrYears9.map(y => counts[y][label] || 0),
      borderColor: color, backgroundColor: color + '22',
      borderWidth: 2, pointRadius: 2, tension: 0.3, fill: false,
    });

    const datasets9 = [
      ...coldThresh.map((t, i) => mkDs9(`< ${t} °C`, coldColors[i])),
      ...warmThresh.map((t, i) => mkDs9(`≥ ${t} °C`, warmColors[i])),
    ];

    climateCharts.thr2 = new Chart(document.getElementById('climThr2').getContext('2d'), {
      type: 'line',
      data: { labels: thrYears9, datasets: datasets9 },
      options: {
        responsive: true, maintainAspectRatio: false,
        interaction: { mode: 'index', intersect: false },
        scales: {
          x: { grid: { color: '#30363d' }, ticks: { color: '#8b949e' } },
          y: { grid: { color: '#30363d' }, ticks: { color: '#8b949e' }, beginAtZero: true,
               title: { display: true, text: 'Tage / Jahr', color: '#8b949e' } },
        },
        plugins: {
          legend: { labels: { color: '#e6edf3', boxWidth: 12, padding: 8, font: { size: 11 } } },
          tooltip: { backgroundColor: '#161b22', borderColor: '#30363d', borderWidth: 1,
                     titleColor: '#e6edf3', bodyColor: '#8b949e' },
        },
      },
    });
  }

  // ----- Card 10: Saison-Anomalien Heatmap (sofort-erkenne wo Extreme lagen) -----
  {
    // Build seasonal means per year. Meteorological seasons:
    //   Winter (Y) = Dec(Y-1) + Jan(Y) + Feb(Y)
    //   Spring (Y) = Mar, Apr, May  (Y)
    //   Summer (Y) = Jun, Jul, Aug  (Y)
    //   Autumn (Y) = Sep, Oct, Nov  (Y)
    const monthly = aggregateData(mData, 'monthly');
    const monthMean = {};  // { 'Y-M': value }
    monthly.forEach(r => {
      const d = new Date(r.ts);
      const k = `${d.getFullYear()}-${d.getMonth()}`;
      // Average across modules within the same year-month
      if (!monthMean[k]) monthMean[k] = { sum: 0, n: 0 };
      monthMean[k].sum += r.value; monthMean[k].n += 1;
    });
    Object.keys(monthMean).forEach(k => { monthMean[k] = monthMean[k].sum / monthMean[k].n; });

    const seasonMonths = {
      Winter:    [[-1, 11], [0, 0], [0, 1]],   // [yearOffset, monthIndex]
      Frühling:  [[0, 2], [0, 3], [0, 4]],
      Sommer:    [[0, 5], [0, 6], [0, 7]],
      Herbst:    [[0, 8], [0, 9], [0, 10]],
    };
    const seasonOrder = ['Winter', 'Frühling', 'Sommer', 'Herbst', 'Jahr'];

    // Compute season means per year (null if any month missing)
    const seasonByYear = {};  // { year: { Winter: x|null, ... } }
    const allYears = yearsArr.map(y => y.year);
    allYears.forEach(yr => {
      seasonByYear[yr] = {};
      Object.entries(seasonMonths).forEach(([s, defs]) => {
        const vals = defs.map(([yo, m]) => monthMean[`${yr + yo}-${m}`]).filter(v => v !== undefined);
        seasonByYear[yr][s] = vals.length === 3 ? vals.reduce((a,b)=>a+b,0)/3 : null;
      });
      // Annual mean from yearsArr
      const yObj = yearsArr.find(o => o.year === yr);
      seasonByYear[yr]['Jahr'] = yObj ? yObj.avg : null;
    });

    // Reference season means over reference period
    const refSeasonMean = {};
    seasonOrder.forEach(s => {
      const vs = [];
      for (let yr = climateRefStart; yr <= climateRefEnd; yr++) {
        const v = seasonByYear[yr] && seasonByYear[yr][s];
        if (v !== null && v !== undefined) vs.push(v);
      }
      refSeasonMean[s] = vs.length ? vs.reduce((a,b)=>a+b,0)/vs.length : null;
    });

    // Anomalies per (year, season)
    const anomGrid = {};  // { year: { season: anom|null } }
    let anomAbsAll = 0;
    allYears.forEach(yr => {
      anomGrid[yr] = {};
      seasonOrder.forEach(s => {
        const v = seasonByYear[yr][s];
        const ref = refSeasonMean[s];
        if (v === null || v === undefined || ref === null) { anomGrid[yr][s] = null; return; }
        const a = v - ref;
        anomGrid[yr][s] = a;
        if (Math.abs(a) > anomAbsAll) anomAbsAll = Math.abs(a);
      });
    });
    if (anomAbsAll === 0) anomAbsAll = 1;

    // Find warmest / coldest year by annual anomaly
    const annualAnoms = allYears
      .filter(yr => recordEligibleYears.has(yr) && anomGrid[yr]['Jahr'] !== null)
      .map(yr => ({ yr, a: anomGrid[yr]['Jahr'] }));
    let warmestYear = null, coldestYear = null;
    if (annualAnoms.length) {
      warmestYear = annualAnoms.reduce((a,b) => b.a > a.a ? b : a).yr;
      coldestYear = annualAnoms.reduce((a,b) => b.a < a.a ? b : a).yr;
    }

    // Find hottest / coldest cell per season across all years
    const seasonExtremes = {};  // { season: {hotYr, hotA, coldYr, coldA} }
    seasonOrder.forEach(s => {
      let hotYr = null, hotA = -Infinity, coldYr = null, coldA = Infinity;
      allYears.forEach(yr => {
        if (!recordEligibleYears.has(yr)) return;
        const a = anomGrid[yr][s];
        if (a === null) return;
        if (a > hotA) { hotA = a; hotYr = yr; }
        if (a < coldA) { coldA = a; coldYr = yr; }
      });
      seasonExtremes[s] = { hotYr, hotA, coldYr, coldA };
    });

    // Diverging color (blue ↔ white ↔ red)
    const divColor = (anom) => {
      const t = Math.max(-1, Math.min(1, anom / anomAbsAll));
      if (t >= 0) {
        const r = Math.round(248 + (185-248)*t);
        const g = Math.round(246 + ( 28-246)*t);
        const b = Math.round(252 + ( 28-252)*t);
        return `rgb(${r},${g},${b})`;
      } else {
        const k = -t;
        const r = Math.round(248 + ( 30-248)*k);
        const g = Math.round(246 + ( 58-246)*k);
        const b = Math.round(252 + (138-252)*k);
        return `rgb(${r},${g},${b})`;
      }
    };

    // Build top banner with key findings
    const findExtremes = () => {
      // Find the single most extreme (hot) seasonal cell across all years
      let topHot = null, topCold = null;
      seasonOrder.filter(s => s !== 'Jahr').forEach(s => {
        const ex = seasonExtremes[s];
        if (ex.hotYr !== null && (topHot === null || ex.hotA > topHot.a)) topHot = { s, yr: ex.hotYr, a: ex.hotA };
        if (ex.coldYr !== null && (topCold === null || ex.coldA < topCold.a)) topCold = { s, yr: ex.coldYr, a: ex.coldA };
      });
      return { topHot, topCold };
    };
    const { topHot, topCold } = findExtremes();

    const banner = document.createElement('div');
    banner.className = 'trend-info';
    const bannerParts = [];
    if (warmestYear !== null) {
      bannerParts.push(`<span class="pill">🔥 Wärmstes Jahr: <b>${warmestYear}</b> (+${anomGrid[warmestYear]['Jahr'].toLocaleString('de-DE',{maximumFractionDigits:2})} ${unit})</span>`);
    }
    if (coldestYear !== null) {
      bannerParts.push(`<span class="pill">🥶 Kältestes Jahr: <b>${coldestYear}</b> (${anomGrid[coldestYear]['Jahr'].toLocaleString('de-DE',{maximumFractionDigits:2})} ${unit})</span>`);
    }
    if (topHot) {
      bannerParts.push(`<span class="pill">☀️ Heißeste Saison: <b>${topHot.s} ${topHot.yr}</b> (+${topHot.a.toLocaleString('de-DE',{maximumFractionDigits:2})} ${unit})</span>`);
    }
    if (topCold) {
      bannerParts.push(`<span class="pill">❄️ Kälteste Saison: <b>${topCold.s} ${topCold.yr}</b> (${topCold.a.toLocaleString('de-DE',{maximumFractionDigits:2})} ${unit})</span>`);
    }
    banner.innerHTML = bannerParts.join('');

    // Build table HTML
    let rowsHtml = '';
    allYears.forEach(yr => {
      const rowCls = yr === warmestYear ? 'year-warmest' : (yr === coldestYear ? 'year-coldest' : '');
      let row = `<tr${rowCls ? ` class="${rowCls}"` : ''}><td class="year-label">${yr}</td>`;
      seasonOrder.forEach(s => {
        const a = anomGrid[yr][s];
        if (a === null) { row += `<td class="cell empty">–</td>`; return; }
        let extra = '';
        if (s !== 'Jahr') {
          const ex = seasonExtremes[s];
          if (yr === ex.hotYr)  extra = ' hot-record';
          if (yr === ex.coldYr) extra = (extra ? extra + ' ' : ' ') + 'cold-record';
        }
        const sign = a >= 0 ? '+' : '';
        const abs = seasonByYear[yr][s];
        const recordHint = recordEligibleYears.has(yr) ? '' : ' - unvollständiges/laufendes Jahr, nicht für Rekorde verwendet';
        const tip = `${s} ${yr}: ${fmt(abs, unit)} (Anomalie ${sign}${a.toLocaleString('de-DE',{maximumFractionDigits:2})} ${unit})${recordHint}`;
        row += `<td class="cell${extra}" style="background:${divColor(a)}" title="${tip}">${sign}${a.toLocaleString('de-DE',{maximumFractionDigits:1})}</td>`;
      });
      row += '</tr>';
      rowsHtml += row;
    });

    const card10 = document.createElement('div');
    card10.className = 'heatmap-card';
    card10.innerHTML = `
      <div class="chart-header"><h2>🍊 Saison-Anomalien: wärmstes/kältestes Jahr auf einen Blick</h2>
        <span class="unit-badge">Δ ${unit}</span></div>
      <div class="chart-hint" style="text-align:left;margin:6px 0 0;">
        Zeilen = Jahre, Spalten = Jahreszeiten (Met. Winter = Dez-Feb). Werte = Abweichung vom Saison-Mittel der Referenzperiode <b>${climateRefStart}-${climateRefEnd}</b>. Jahres-Waermerekord und Jahres-Kaelterekord pro Saison werden markiert; unvollständige oder laufende Jahre werden dafür ignoriert.
      </div>
      ${bannerParts.length ? '<div id="climSeasonBanner"></div>' : ''}
      <table class="heatmap" style="margin-top:12px;">
        <tr><th></th>${seasonOrder.map(s => `<th>${s}</th>`).join('')}</tr>
        ${rowsHtml}
      </table>
      <div class="heatmap-legend">
        <span>kälter</span>
        <span class="legend-bar diverging"></span>
        <span>wärmer</span>
        <span style="margin-left:14px;">□ Goldrand = Hitzerekord pro Saison · Cyanrand = Kälterekord</span>
      </div>`;
    container.appendChild(card10);
    if (bannerParts.length) {
      document.getElementById('climSeasonBanner').replaceWith(banner);
    }
  }

  // ----- Card 11 & 12: Tropennächte und Heiße Tage mit einstellbarem Schwellwert -----
  if (/°?\s*C/i.test(unit) || /temp/i.test(climateSensor)) {
    // Daily min/max collapsed across modules in mData
    const dailyTh = aggregateMinAvgMax(mData, 'daily');
    const dayAggTh = {};
    dailyTh.forEach(r => {
      const d = new Date(r.ts);
      const dk = `${d.getFullYear()}-${d.getMonth()}-${d.getDate()}`;
      if (!dayAggTh[dk]) dayAggTh[dk] = { year: d.getFullYear(), min: r.min, max: r.max };
      else {
        if (r.min < dayAggTh[dk].min) dayAggTh[dk].min = r.min;
        if (r.max > dayAggTh[dk].max) dayAggTh[dk].max = r.max;
      }
    });
    const thrAllYears = [...new Set(Object.values(dayAggTh).map(d => d.year))].sort();

    // Generic builder for one threshold-slider card
    const buildThresholdCard = (cfg) => {
      // cfg: { id, title, sliderId, valueId, chartId, color, min, max, step, init,
      //        compute(threshold) -> {years, counts}, formatLabel(threshold) }
      const card = document.createElement('div');
      card.className = 'chart-card';
      card.innerHTML = `
        <div class="chart-header"><h2>${cfg.title}</h2>
          <span class="unit-badge">Tage / Jahr</span></div>
        <div class="thr-slider-row">
          <label for="${cfg.sliderId}">${cfg.sliderLabel}</label>
          <input type="range" id="${cfg.sliderId}" min="${cfg.min}" max="${cfg.max}" step="${cfg.step}" value="${cfg.init}">
          <span class="thr-val" id="${cfg.valueId}">${cfg.formatLabel(cfg.init)}</span>
        </div>
        <div class="chart-wrap tall"><canvas id="${cfg.chartId}"></canvas></div>
        <div class="thr-slider-row"><span class="thr-info">${cfg.hint}</span></div>`;
      container.appendChild(card);

      const sliderEl = document.getElementById(cfg.sliderId);
      const valueEl  = document.getElementById(cfg.valueId);

      const makeData = (threshold) => {
        const counts = thrAllYears.map(y => 0);
        Object.values(dayAggTh).forEach(d => {
          if (cfg.match(d, threshold)) {
            const idx = thrAllYears.indexOf(d.year);
            if (idx >= 0) counts[idx]++;
          }
        });
        return counts;
      };

      const initData = makeData(cfg.init);
      const chart = new Chart(document.getElementById(cfg.chartId).getContext('2d'), {
        type: 'bar',
        data: {
          labels: thrAllYears,
          datasets: [{
            label: cfg.formatLabel(cfg.init),
            data: initData,
            backgroundColor: cfg.color + 'aa',
            borderColor: cfg.color,
            borderWidth: 1,
          }],
        },
        options: {
          responsive: true, maintainAspectRatio: false,
          scales: {
            x: { grid: { color: '#30363d' }, ticks: { color: '#8b949e' } },
            y: { grid: { color: '#30363d' }, ticks: { color: '#8b949e' }, beginAtZero: true,
                 title: { display: true, text: 'Tage / Jahr', color: '#8b949e' } },
          },
          plugins: {
            legend: { labels: { color: '#e6edf3', boxWidth: 12, padding: 10 } },
            tooltip: { backgroundColor: '#161b22', borderColor: '#30363d', borderWidth: 1,
                       titleColor: '#e6edf3', bodyColor: '#8b949e',
                       callbacks: { label: ctx => ` ${ctx.dataset.label}: ${ctx.parsed.y} Tage` } },
          },
        },
      });
      climateCharts[cfg.id] = chart;

      sliderEl.addEventListener('input', () => {
        const t = +sliderEl.value;
        cfg.setGlobal(t);
        valueEl.textContent = cfg.formatLabel(t);
        chart.data.datasets[0].data = makeData(t);
        chart.data.datasets[0].label = cfg.formatLabel(t);
        chart.update('none');
      });
    };

    // Card 11: Tropennächte (Tmin > Schwellwert)
    buildThresholdCard({
      id: 'tropNights',
      title: '🌴 Tropennächte – Nächte mit Tmin über Schwellwert',
      sliderId: 'climTropSlider',
      sliderLabel: 'Schwellwert (Tmin >):',
      valueId: 'climTropVal',
      chartId: 'climTropChart',
      color: '#bc8cff',
      min: 12, max: 25, step: 0.5, init: climateTropThreshold,
      hint: 'Eine Tropennacht ist klassisch eine Nacht, in der das Tagesminimum nicht unter 20 °C fällt. Schieberegler einstellen, um eigene Grenzwerte zu prüfen (z. B. 18, 19, 20 °C).',
      formatLabel: t => `Tmin > ${t.toLocaleString('de-DE', {maximumFractionDigits:1})} °C`,
      match: (d, t) => d.min > t,
      setGlobal: t => { climateTropThreshold = t; },
    });

    // Card 12: Heiße Tage (Tmax > Schwellwert)
    buildThresholdCard({
      id: 'hotDays',
      title: '☀️ Heiße Tage – Tage mit Tmax über Schwellwert',
      sliderId: 'climHotSlider',
      sliderLabel: 'Schwellwert (Tmax >):',
      valueId: 'climHotVal',
      chartId: 'climHotChart',
      color: '#f85149',
      min: 20, max: 38, step: 0.5, init: climateHotThreshold,
      hint: 'Ein heißer Tag (Hitzetag) ist klassisch ein Tag mit Tagesmaximum ≥ 30 °C. Schieberegler einstellen, um eigene Grenzwerte zu prüfen (z. B. 25, 28, 30, 32 °C).',
      formatLabel: t => `Tmax > ${t.toLocaleString('de-DE', {maximumFractionDigits:1})} °C`,
      match: (d, t) => d.max > t,
      setGlobal: t => { climateHotThreshold = t; },
    });
  }
}

/* ── Streaks helper ── */
function computeStreaks(sortedDays, condFn) {
  const streaks = [];
  let cur = null;
  sortedDays.forEach(d => {
    if (condFn(d)) {
      if (!cur) cur = { start: d.ts, end: d.ts, days: 1, maxTemp: d.max, minTemp: d.min, dates: [d] };
      else { cur.end = d.ts; cur.days++; if (d.max > cur.maxTemp) cur.maxTemp = d.max; if (d.min < cur.minTemp) cur.minTemp = d.min; cur.dates.push(d); }
    } else {
      if (cur && cur.days >= 2) streaks.push(cur);
      cur = null;
    }
  });
  if (cur && cur.days >= 2) streaks.push(cur);
  return streaks.sort((a, b) => b.days - a.days);
}

/* ── Extremes: Hitzewellen, Kältephasen, Ausreißer ── */
function renderExtremes(raw) {
  const container = document.getElementById('extremesContainer');
  Object.values(extremeCharts).forEach(c => c.destroy());
  extremeCharts = {};
  container.innerHTML = '';

  const allSensors = [...new Set(raw.map(d => d.sensor))];
  const tempSensors = allSensors.filter(s => /°?\s*C/i.test(PAYLOAD.units[s] || '') || /temp/i.test(s));
  const sensors = tempSensors.length ? tempSensors : allSensors;
  if (!sensors.length) { container.innerHTML = '<div class="no-data">Keine Daten.</div>'; return; }
  if (!extremeSensor || !sensors.includes(extremeSensor)) extremeSensor = sensors[0];

  const sData = raw.filter(d => d.sensor === extremeSensor);
  const modules = ['(alle gewählten Module)', ...new Set(sData.map(d => d.module))].sort();
  if (!extremeModule || !modules.includes(extremeModule)) extremeModule = modules[0];
  const mData = extremeModule === '(alle gewählten Module)' ? sData : sData.filter(d => d.module === extremeModule);
  const unit = PAYLOAD.units[extremeSensor] || sData[0]?.unit || '';

  // Controls
  const ctrl = document.createElement('div');
  ctrl.className = 'trend-controls';
  ctrl.innerHTML = `
    <label>Sensor:</label>
    <select id="extSensorSel">${sensors.map(s => `<option ${s===extremeSensor?'selected':''}>${s}</option>`).join('')}</select>
    <label>Modul:</label>
    <select id="extModSel">${modules.map(m => `<option ${m===extremeModule?'selected':''}>${m}</option>`).join('')}</select>`;
  container.appendChild(ctrl);
  ctrl.querySelector('#extSensorSel').onchange = e => { extremeSensor = e.target.value; extremeModule = null; renderExtremes(raw); };
  ctrl.querySelector('#extModSel').onchange    = e => { extremeModule = e.target.value; renderExtremes(raw); };

  // Build daily min/max list
  const daily = aggregateMinAvgMax(mData, 'daily');
  const dayMap = {};
  daily.forEach(r => {
    const d = new Date(r.ts);
    const dk = `${d.getFullYear()}-${d.getMonth()}-${d.getDate()}`;
    if (!dayMap[dk]) dayMap[dk] = { ts: r.ts, year: d.getFullYear(), month: d.getMonth(), min: r.min, max: r.max, avg: r.avg };
    else { if (r.min < dayMap[dk].min) dayMap[dk].min = r.min; if (r.max > dayMap[dk].max) dayMap[dk].max = r.max; }
  });
  const dayList = Object.values(dayMap).sort((a, b) => a.ts - b.ts);

  if (!dayList.length) { container.innerHTML = '<div class="no-data">Keine Tagesdaten verfügbar.</div>'; return; }

  // ── Card 1: Längste Hitzewellen ──
  {
    const card = document.createElement('div');
    card.className = 'chart-card';
    card.innerHTML = `
      <div class="chart-header"><h2>🔥 Längste Hitzewellen (Top 15)</h2>
        <span class="unit-badge">Tage</span></div>
      <div class="thr-slider-row">
        <label for="extHeatThr">Schwellwert (Tmax ≥):</label>
        <input type="range" id="extHeatThr" min="20" max="38" step="0.5" value="${extremeHeatThr}">
        <span class="thr-val" id="extHeatThrVal">${extremeHeatThr.toLocaleString('de-DE',{maximumFractionDigits:1})} ${unit}</span>
      </div>
      <div class="chart-wrap tall"><canvas id="extHeatChart"></canvas></div>
      <div id="extHeatTable" style="margin-top:12px;"></div>`;
    container.appendChild(card);

    const renderHeat = (thr) => {
      const streaks = computeStreaks(dayList, d => d.max >= thr);
      const top = streaks.slice(0, 15);
      if (extremeCharts.heat) extremeCharts.heat.destroy();
      extremeCharts.heat = new Chart(document.getElementById('extHeatChart').getContext('2d'), {
        type: 'bar',
        data: {
          labels: top.map(s => fmtDay(s.start)),
          datasets: [{ label: `Dauer (Tmax ≥ ${thr.toLocaleString('de-DE',{maximumFractionDigits:1})} ${unit})`,
            data: top.map(s => s.days),
            backgroundColor: top.map((_, i) => `rgba(248,81,73,${Math.max(0.35, 1 - i*0.045)})`),
            borderColor: '#f85149', borderWidth: 1 }],
        },
        options: {
          indexAxis: 'y', responsive: true, maintainAspectRatio: false,
          scales: {
            x: { grid:{color:'#30363d'}, ticks:{color:'#8b949e'}, title:{display:true,text:'Dauer (Tage)',color:'#8b949e'}, beginAtZero: true },
            y: { grid:{color:'#30363d'}, ticks:{color:'#8b949e', font:{size:11}} },
          },
          plugins: {
            legend: { labels:{color:'#e6edf3',boxWidth:12} },
            tooltip: { backgroundColor:'#161b22', borderColor:'#30363d', borderWidth:1, titleColor:'#e6edf3', bodyColor:'#8b949e',
              callbacks: { label: ctx => ` ${ctx.parsed.x} Tage`,
                afterLabel: ctx => { const s=top[ctx.dataIndex]; return [`  bis: ${fmtDay(s.end)}`,`  Tmax: ${fmt(s.maxTemp,unit)}`]; } } },
          },
        },
      });
      const el = document.getElementById('extHeatTable');
      if (!top.length) { el.innerHTML = '<div class="no-data">Keine Hitzewellen gefunden.</div>'; return; }
      el.innerHTML = `<table class="rec-table">
        <tr><th>#</th><th>Von</th><th>Bis</th><th style="text-align:right">Tage</th><th style="text-align:right">Tmax</th></tr>
        ${top.map((s,i)=>`<tr><td>${i+1}</td><td>${fmtDay(s.start)}</td><td>${fmtDay(s.end)}</td><td class="val max">${s.days}</td><td class="val max">${fmt(s.maxTemp,unit)}</td></tr>`).join('')}
      </table>`;
    };
    renderHeat(extremeHeatThr);
    document.getElementById('extHeatThr').addEventListener('input', e => {
      extremeHeatThr = +e.target.value;
      document.getElementById('extHeatThrVal').textContent = `${extremeHeatThr.toLocaleString('de-DE',{maximumFractionDigits:1})} ${unit}`;
      renderHeat(extremeHeatThr);
    });
  }

  // ── Card 2: Längste Kältephasen ──
  {
    const card = document.createElement('div');
    card.className = 'chart-card';
    card.innerHTML = `
      <div class="chart-header"><h2>❄️ Längste Kältephasen (Top 15)</h2>
        <span class="unit-badge">Tage</span></div>
      <div class="thr-slider-row">
        <label for="extColdThr">Schwellwert (Tmax &lt;):</label>
        <input type="range" id="extColdThr" min="-15" max="15" step="0.5" value="${extremeColdThr}">
        <span class="thr-val" id="extColdThrVal">${extremeColdThr.toLocaleString('de-DE',{maximumFractionDigits:1})} ${unit}</span>
      </div>
      <div class="chart-wrap tall"><canvas id="extColdChart"></canvas></div>
      <div id="extColdTable" style="margin-top:12px;"></div>`;
    container.appendChild(card);

    const renderCold = (thr) => {
      const streaks = computeStreaks(dayList, d => d.max < thr);
      const top = streaks.slice(0, 15);
      if (extremeCharts.cold) extremeCharts.cold.destroy();
      extremeCharts.cold = new Chart(document.getElementById('extColdChart').getContext('2d'), {
        type: 'bar',
        data: {
          labels: top.map(s => fmtDay(s.start)),
          datasets: [{ label: `Dauer (Tmax < ${thr.toLocaleString('de-DE',{maximumFractionDigits:1})} ${unit})`,
            data: top.map(s => s.days),
            backgroundColor: top.map((_, i) => `rgba(88,166,255,${Math.max(0.35, 1 - i*0.045)})`),
            borderColor: '#58a6ff', borderWidth: 1 }],
        },
        options: {
          indexAxis: 'y', responsive: true, maintainAspectRatio: false,
          scales: {
            x: { grid:{color:'#30363d'}, ticks:{color:'#8b949e'}, title:{display:true,text:'Dauer (Tage)',color:'#8b949e'}, beginAtZero: true },
            y: { grid:{color:'#30363d'}, ticks:{color:'#8b949e', font:{size:11}} },
          },
          plugins: {
            legend: { labels:{color:'#e6edf3',boxWidth:12} },
            tooltip: { backgroundColor:'#161b22', borderColor:'#30363d', borderWidth:1, titleColor:'#e6edf3', bodyColor:'#8b949e',
              callbacks: { label: ctx => ` ${ctx.parsed.x} Tage`,
                afterLabel: ctx => { const s=top[ctx.dataIndex]; return [`  bis: ${fmtDay(s.end)}`,`  Tmin: ${fmt(s.minTemp,unit)}`]; } } },
          },
        },
      });
      const el = document.getElementById('extColdTable');
      if (!top.length) { el.innerHTML = '<div class="no-data">Keine Kältephasen gefunden.</div>'; return; }
      el.innerHTML = `<table class="rec-table">
        <tr><th>#</th><th>Von</th><th>Bis</th><th style="text-align:right">Tage</th><th style="text-align:right">Tmin</th></tr>
        ${top.map((s,i)=>`<tr><td>${i+1}</td><td>${fmtDay(s.start)}</td><td>${fmtDay(s.end)}</td><td class="val min">${s.days}</td><td class="val min">${fmt(s.minTemp,unit)}</td></tr>`).join('')}
      </table>`;
    };
    renderCold(extremeColdThr);
    document.getElementById('extColdThr').addEventListener('input', e => {
      extremeColdThr = +e.target.value;
      document.getElementById('extColdThrVal').textContent = `${extremeColdThr.toLocaleString('de-DE',{maximumFractionDigits:1})} ${unit}`;
      renderCold(extremeColdThr);
    });
  }

  // ── Cards 3+4: Monatliche Ausreißer (z-Score) ──
  {
    const dailyAvg = aggregateData(mData, 'daily');
    const byMonth = {};
    dailyAvg.forEach(d => {
      const m = new Date(d.ts).getMonth();
      if (!byMonth[m]) byMonth[m] = [];
      byMonth[m].push(d.value);
    });
    const monthStats = {};
    for (let m = 0; m < 12; m++) {
      const vs = byMonth[m] || [];
      if (!vs.length) { monthStats[m] = { mean: null, std: 1 }; continue; }
      const mean = vs.reduce((a,b)=>a+b,0)/vs.length;
      const std  = Math.sqrt(vs.reduce((s,v)=>s+(v-mean)*(v-mean),0)/vs.length) || 1;
      monthStats[m] = { mean, std };
    }
    const zScored = dailyAvg.map(d => {
      const m = new Date(d.ts).getMonth();
      const { mean, std } = monthStats[m];
      if (mean === null) return null;
      return { ts: d.ts, value: d.value, month: m, z: (d.value - mean) / std, mean, std };
    }).filter(Boolean);

    const hotOutliers  = [...zScored].sort((a,b) => b.z - a.z).slice(0, 15);
    const coldOutliers = [...zScored].sort((a,b) => a.z - b.z).slice(0, 15);
    const monthNames   = ['Jan','Feb','Mär','Apr','Mai','Jun','Jul','Aug','Sep','Okt','Nov','Dez'];
    const maxHotZ      = Math.max(...hotOutliers.map(d => d.z)) || 1;
    const maxColdZAbs  = Math.max(...coldOutliers.map(d => Math.abs(d.z))) || 1;

    // Card 3: Warm outliers
    const card3 = document.createElement('div');
    card3.className = 'chart-card';
    card3.innerHTML = `
      <div class="chart-header"><h2>🌡️ Monatlich ungewöhnlichste Hitzetage – Top 15 Ausreißer nach oben</h2>
        <span class="unit-badge">z-Score</span></div>
      <div class="chart-wrap tall"><canvas id="extHotOutliers"></canvas></div>
      <div class="chart-hint">z-Score = Abweichung in Standardabweichungen vom jeweiligen Monats-Mittel. Hoher Wert = für diesen Monat besonders ungewöhnlich warm.</div>
      <div id="extHotTable" style="margin-top:12px;"></div>`;
    container.appendChild(card3);

    extremeCharts.hotOut = new Chart(document.getElementById('extHotOutliers').getContext('2d'), {
      type: 'bar',
      data: {
        labels: hotOutliers.map(d => fmtDay(d.ts)),
        datasets: [{ label: 'z-Score (Wärme)', data: hotOutliers.map(d => Math.round(d.z*100)/100),
          backgroundColor: hotOutliers.map(d => `rgba(248,81,73,${Math.min(1, 0.35 + d.z/maxHotZ*0.65)})`),
          borderColor: '#f85149', borderWidth: 1 }],
      },
      options: {
        indexAxis: 'y', responsive: true, maintainAspectRatio: false,
        scales: {
          x: { grid:{color:'#30363d'}, ticks:{color:'#8b949e'}, title:{display:true,text:'z-Score (σ)',color:'#8b949e'}, beginAtZero: true },
          y: { grid:{color:'#30363d'}, ticks:{color:'#8b949e', font:{size:11}} },
        },
        plugins: {
          legend: { display: false },
          tooltip: { backgroundColor:'#161b22', borderColor:'#30363d', borderWidth:1, titleColor:'#e6edf3', bodyColor:'#8b949e',
            callbacks: { label: ctx => ` z = +${ctx.parsed.x.toLocaleString('de-DE',{maximumFractionDigits:2})}σ`,
              afterLabel: ctx => { const d=hotOutliers[ctx.dataIndex]; return [`  Tagesmittel: ${fmt(d.value,unit)}`,`  Monats-Ø: ${fmt(d.mean,unit)} ± ${d.std.toLocaleString('de-DE',{maximumFractionDigits:2})} ${unit}`,`  Monat: ${monthNames[d.month]}`]; } } },
        },
      },
    });
    document.getElementById('extHotTable').innerHTML = `<table class="rec-table">
      <tr><th>#</th><th>Datum</th><th>Monat</th><th style="text-align:right">Tagesmittel</th><th style="text-align:right">Monats-Ø</th><th style="text-align:right">z-Score</th></tr>
      ${hotOutliers.map((d,i)=>`<tr><td>${i+1}</td><td>${fmtDay(d.ts)}</td><td>${monthNames[d.month]}</td><td class="val max">${fmt(d.value,unit)}</td><td class="val">${fmt(d.mean,unit)}</td><td class="val max">+${d.z.toLocaleString('de-DE',{maximumFractionDigits:2})}σ</td></tr>`).join('')}
    </table>`;

    // Card 4: Cold outliers
    const card4 = document.createElement('div');
    card4.className = 'chart-card';
    card4.innerHTML = `
      <div class="chart-header"><h2>❄️ Monatlich ungewöhnlichste Kältetage – Top 15 Ausreißer nach unten</h2>
        <span class="unit-badge">z-Score</span></div>
      <div class="chart-wrap tall"><canvas id="extColdOutliers"></canvas></div>
      <div class="chart-hint">Negativer z-Score = für den Monat besonders ungewöhnlich kalt.</div>
      <div id="extColdTable2" style="margin-top:12px;"></div>`;
    container.appendChild(card4);

    extremeCharts.coldOut = new Chart(document.getElementById('extColdOutliers').getContext('2d'), {
      type: 'bar',
      data: {
        labels: coldOutliers.map(d => fmtDay(d.ts)),
        datasets: [{ label: 'z-Score (Kälte)', data: coldOutliers.map(d => Math.round(d.z*100)/100),
          backgroundColor: coldOutliers.map(d => `rgba(88,166,255,${Math.min(1, 0.35 + Math.abs(d.z)/maxColdZAbs*0.65)})`),
          borderColor: '#58a6ff', borderWidth: 1 }],
      },
      options: {
        indexAxis: 'y', responsive: true, maintainAspectRatio: false,
        scales: {
          x: { grid:{color:'#30363d'}, ticks:{color:'#8b949e'}, title:{display:true,text:'z-Score (σ)',color:'#8b949e'} },
          y: { grid:{color:'#30363d'}, ticks:{color:'#8b949e', font:{size:11}} },
        },
        plugins: {
          legend: { display: false },
          tooltip: { backgroundColor:'#161b22', borderColor:'#30363d', borderWidth:1, titleColor:'#e6edf3', bodyColor:'#8b949e',
            callbacks: { label: ctx => ` z = ${ctx.parsed.x.toLocaleString('de-DE',{maximumFractionDigits:2})}σ`,
              afterLabel: ctx => { const d=coldOutliers[ctx.dataIndex]; return [`  Tagesmittel: ${fmt(d.value,unit)}`,`  Monats-Ø: ${fmt(d.mean,unit)} ± ${d.std.toLocaleString('de-DE',{maximumFractionDigits:2})} ${unit}`,`  Monat: ${monthNames[d.month]}`]; } } },
        },
      },
    });
    document.getElementById('extColdTable2').innerHTML = `<table class="rec-table">
      <tr><th>#</th><th>Datum</th><th>Monat</th><th style="text-align:right">Tagesmittel</th><th style="text-align:right">Monats-Ø</th><th style="text-align:right">z-Score</th></tr>
      ${coldOutliers.map((d,i)=>`<tr><td>${i+1}</td><td>${fmtDay(d.ts)}</td><td>${monthNames[d.month]}</td><td class="val min">${fmt(d.value,unit)}</td><td class="val">${fmt(d.mean,unit)}</td><td class="val min">${d.z.toLocaleString('de-DE',{maximumFractionDigits:2})}σ</td></tr>`).join('')}
    </table>`;

    // Card 5: Scatter aller Ausreißer (|z|>2) pro Monat
    const card5 = document.createElement('div');
    card5.className = 'chart-card';
    card5.innerHTML = `
      <div class="chart-header"><h2>📊 Ausreißer-Übersicht pro Monat (|z| &gt; 2σ)</h2>
        <span class="unit-badge">z-Score</span></div>
      <div class="chart-wrap tall"><canvas id="extMonthScatter"></canvas></div>
      <div class="chart-hint">Jeder Punkt = 1 Tag mit |z| > 2 (mehr als 2 Standardabweichungen vom Monats-Mittel). Rot = wärmer, Blau = kälter als erwartet. Hover für Datum und Wert.</div>`;
    container.appendChild(card5);

    const outliersFlat = zScored.filter(d => Math.abs(d.z) > 2);
    extremeCharts.scatter = new Chart(document.getElementById('extMonthScatter').getContext('2d'), {
      type: 'scatter',
      data: {
        datasets: [
          { label: 'Wärmer als normal (z > 2σ)',
            data: outliersFlat.filter(d => d.z > 0).map(d => ({ x: monthNames[d.month], y: Math.round(d.z*100)/100, ts: d.ts, value: d.value, month: d.month })),
            backgroundColor: 'rgba(248,81,73,0.7)', borderColor: '#f85149', pointRadius: 5, pointHoverRadius: 7 },
          { label: 'Kälter als normal (z < −2σ)',
            data: outliersFlat.filter(d => d.z < 0).map(d => ({ x: monthNames[d.month], y: Math.round(d.z*100)/100, ts: d.ts, value: d.value, month: d.month })),
            backgroundColor: 'rgba(88,166,255,0.7)', borderColor: '#58a6ff', pointRadius: 5, pointHoverRadius: 7 },
        ],
      },
      options: {
        responsive: true, maintainAspectRatio: false,
        scales: {
          x: { type: 'category', labels: monthNames, grid:{color:'#30363d'}, ticks:{color:'#8b949e'} },
          y: { grid:{color:'#30363d'}, ticks:{color:'#8b949e'}, title:{display:true,text:'z-Score (σ)',color:'#8b949e'} },
        },
        plugins: {
          legend: { labels:{color:'#e6edf3',boxWidth:12} },
          tooltip: { backgroundColor:'#161b22', borderColor:'#30363d', borderWidth:1, titleColor:'#e6edf3', bodyColor:'#8b949e',
            callbacks: { title: ctx => fmtDay(ctx[0].raw.ts),
              label: ctx => ` z = ${ctx.parsed.y.toLocaleString('de-DE',{maximumFractionDigits:2})}σ  ·  ${fmt(ctx.raw.value, unit)}` } },
        },
      },
    });
  }

  // ── Card 6: Timeline mit markierten Hitze- und Kältephasen ──
  {
    const card6 = document.createElement('div');
    card6.className = 'chart-card';
    card6.innerHTML = `
      <div class="chart-header"><h2>📅 Tages-Extrema im Zeitverlauf – Hitze- und Kältephasen markiert</h2>
        <span class="unit-badge">${unit}</span></div>
      <div class="chart-wrap tall"><canvas id="extTimeline"></canvas></div>
      <div class="chart-hint">Tagesmax (rot) und Tagesmin (blau). Orangene Punkte = Hitzewellen-Tage (Tmax ≥ Schwellwert), Cyanfarbene Punkte = Kältephasen-Tage. Scrollen/Ziehen zum Zoomen.</div>`;
    container.appendChild(card6);

    const heatSet = new Set();
    const coldSet = new Set();
    computeStreaks(dayList, d => d.max >= extremeHeatThr).forEach(s => s.dates.forEach(d => heatSet.add(d.ts)));
    computeStreaks(dayList, d => d.max < extremeColdThr).forEach(s => s.dates.forEach(d => coldSet.add(d.ts)));

    extremeCharts.timeline = new Chart(document.getElementById('extTimeline').getContext('2d'), {
      type: 'line',
      data: {
        datasets: [
          { label: 'Tagesmax', data: dayList.map(d=>({x:d.ts,y:d.max})),
            borderColor:'#f8514966', backgroundColor:'transparent', borderWidth:1.2, pointRadius:0, tension:0.2 },
          { label: 'Tagesmin', data: dayList.map(d=>({x:d.ts,y:d.min})),
            borderColor:'#58a6ff66', backgroundColor:'transparent', borderWidth:1.2, pointRadius:0, tension:0.2 },
          { label: `Hitzewelle (≥${extremeHeatThr}°C)`, type:'scatter',
            data: dayList.filter(d=>heatSet.has(d.ts)).map(d=>({x:d.ts,y:d.max})),
            backgroundColor:'#f59e0b', borderColor:'#f59e0b', pointRadius:3, pointHoverRadius:5 },
          { label: `Kältephase (<${extremeColdThr}°C)`, type:'scatter',
            data: dayList.filter(d=>coldSet.has(d.ts)).map(d=>({x:d.ts,y:d.min})),
            backgroundColor:'#39c5cf', borderColor:'#39c5cf', pointRadius:3, pointHoverRadius:5 },
        ],
      },
      options: commonChartOpts(unit, true),
    });
  }
}

/* ── Indoor / Raumklima-Analysen ── */
let indoorCharts = {};
let indoorModule = null, outdoorModuleInd = null;

function dewPoint(t, rh) {
  if (!rh || rh <= 0 || rh > 100) return null;
  const a = 17.625, b = 243.04;
  const gamma = Math.log(rh / 100) + a * t / (b + t);
  return b * gamma / (a - gamma);
}

function absoluteHumidity(t, rh) {
  return 6.112 * Math.exp(17.67 * t / (t + 243.5)) * rh * 2.1674 / (273.15 + t);
}

function heatIndexC(t, rh) {
  if (t < 26 || rh < 40) return t;
  return -8.78469475556 + 1.61139411*t + 2.33854883889*rh
       - 0.14611605*t*rh - 0.012308094*t*t - 0.016424828*rh*rh
       + 0.002211732*t*t*rh + 0.00072546*t*rh*rh - 0.000003582*t*t*rh*rh;
}

function comfortScoreInd(t, rh, co2) {
  let wSum = 0, sSum = 0;
  if (t !== null && !isNaN(t)) {
    const w = 0.3; let s = 0;
    if (t >= 20 && t <= 23) s = 100;
    else if (t >= 18 && t < 20) s = 60 + (t-18)*20;
    else if (t > 23 && t <= 26) s = 60 + (26-t)*13.3;
    else if (t < 18) s = Math.max(0, 60-(18-t)*15);
    else s = Math.max(0, 60-(t-26)*15);
    wSum += w; sSum += s * w;
  }
  if (rh !== null && !isNaN(rh)) {
    const w = 0.3; let s = 0;
    if (rh >= 40 && rh <= 60) s = 100;
    else if (rh >= 30 && rh < 40) s = (rh-30)*10;
    else if (rh > 60 && rh <= 70) s = (70-rh)*10;
    wSum += w; sSum += s * w;
  }
  if (co2 !== null && !isNaN(co2)) {
    const w = 0.4; let s = 0;
    if (co2 < 800) s = 100;
    else if (co2 < 1000) s = 100-(co2-800)/2;
    else if (co2 < 1400) s = Math.max(0, 50-(co2-1000)/8);
    wSum += w; sSum += s * w;
  }
  return wSum > 0 ? sSum / wSum : null;
}

function renderIndoor(raw) {
  const container = document.getElementById('indoorContainer');
  Object.values(indoorCharts).forEach(c => c.destroy());
  indoorCharts = {};
  container.innerHTML = '';

  const modules = [...new Set(raw.map(d => d.module))].sort();
  if (!modules.length) { container.innerHTML = '<div class="no-data">Keine Daten.</div>'; return; }

  const isTempD = d => /°?\s*C/i.test(d.unit) || /temp/i.test(d.sensor);
  const isHumD  = d => d.unit === '%' || /hum|feuch/i.test(d.sensor);
  const isCO2D  = d => /ppm/i.test(d.unit) || /co2/i.test(d.sensor);

  const tempSensors = [...new Set(raw.filter(isTempD).map(d => d.sensor))];
  const humSensors  = [...new Set(raw.filter(isHumD).map(d => d.sensor))];
  const co2Sensors  = [...new Set(raw.filter(isCO2D).map(d => d.sensor))];

  const co2Mods = co2Sensors.length ? [...new Set(raw.filter(isCO2D).map(d => d.module))] : [];
  if (!indoorModule || !modules.includes(indoorModule)) indoorModule = co2Mods[0] || modules[0];
  if (!outdoorModuleInd || !modules.includes(outdoorModuleInd))
    outdoorModuleInd = modules.find(m => m !== indoorModule) || modules[0];

  const ctrl = document.createElement('div');
  ctrl.className = 'trend-controls';
  ctrl.innerHTML = `
    <label>🏠 Innenraum-Modul:</label>
    <select id="indModSel">${modules.map(m=>`<option ${m===indoorModule?'selected':''}>${m}</option>`).join('')}</select>
    <label>🌤️ Außen-Modul:</label>
    <select id="outModSel">${modules.map(m=>`<option ${m===outdoorModuleInd?'selected':''}>${m}</option>`).join('')}</select>`;
  container.appendChild(ctrl);
  ctrl.querySelector('#indModSel').onchange = e => { indoorModule = e.target.value; renderIndoor(raw); };
  ctrl.querySelector('#outModSel').onchange = e => { outdoorModuleInd = e.target.value; renderIndoor(raw); };

  const inData  = raw.filter(d => d.module === indoorModule);
  const outData = raw.filter(d => d.module === outdoorModuleInd);

  const inTempSen  = tempSensors.find(s => inData.some(d => d.sensor === s));
  const inHumSen   = humSensors.find(s => inData.some(d => d.sensor === s));
  const inCO2Sen   = co2Sensors.find(s => inData.some(d => d.sensor === s));
  const outTempSen = tempSensors.find(s => outData.some(d => d.sensor === s));
  const outHumSen  = humSensors.find(s => outData.some(d => d.sensor === s));

  const inTempH  = inTempSen  ? aggregateData(inData.filter(d=>d.sensor===inTempSen),  'hourly') : [];
  const inHumH   = inHumSen   ? aggregateData(inData.filter(d=>d.sensor===inHumSen),   'hourly') : [];
  const inCO2H   = inCO2Sen   ? aggregateData(inData.filter(d=>d.sensor===inCO2Sen),   'hourly') : [];
  const outTempH = outTempSen ? aggregateData(outData.filter(d=>d.sensor===outTempSen),'hourly') : [];
  const outHumH  = outHumSen  ? aggregateData(outData.filter(d=>d.sensor===outHumSen), 'hourly') : [];

  const mkMap = arr => new Map(arr.map(d=>[d.ts,d.value]));
  const inTempMap=mkMap(inTempH), inHumMap=mkMap(inHumH), inCO2Map=mkMap(inCO2H);
  const outTempMap=mkMap(outTempH), outHumMap=mkMap(outHumH);

  const allTs = [...new Set([...inTempH,...inHumH,...inCO2H].map(d=>d.ts))].sort((a,b)=>a-b);

  const derived = allTs.map(ts => {
    const t    = inTempMap.get(ts)  ?? null;
    const rh   = inHumMap.get(ts)   ?? null;
    const co2  = inCO2Map.get(ts)   ?? null;
    const tOut = outTempMap.get(ts) ?? null;
    const rhOut= outHumMap.get(ts)  ?? null;
    const dp    = (t!==null&&rh!==null) ? Math.round((dewPoint(t,rh)??0)*10)/10 : null;
    const ah    = (t!==null&&rh!==null) ? Math.round(absoluteHumidity(t,rh)*10)/10 : null;
    const hi    = (t!==null&&rh!==null) ? Math.round(heatIndexC(t,rh)*10)/10 : null;
    const dpOut = (tOut!==null&&rhOut!==null) ? Math.round((dewPoint(tOut,rhOut)??0)*10)/10 : null;
    const ahOut = (tOut!==null&&rhOut!==null) ? Math.round(absoluteHumidity(tOut,rhOut)*10)/10 : null;
    const cs    = comfortScoreInd(t, rh, co2);
    return { ts, t, rh, co2, tOut, rhOut, dp, ah, hi, dpOut, ahOut, cs };
  });

  const latest = [...derived].reverse().find(d=>d.t!==null||d.rh!==null||d.co2!==null) || {};

  /* ── Cockpit ── */
  const INF = 1e9;
  const mkCard = (icon, label, val, unit, lo, hi2, zones) => {
    if (val===null||val===undefined||isNaN(val)) return `
      <div class="cockpit-card ck-na">
        <div class="ck-icon">${icon}</div><div class="ck-label">${label}</div>
        <div class="ck-val">–</div>
        <div class="ck-bar-bg"><div class="ck-bar-fill" style="width:0%"></div></div>
        <div class="ck-status">Kein Sensor</div></div>`;
    let cls='ck-ok', st='OK';
    for(const z of zones){if(val>=z.lo&&val<z.hi){cls='ck-'+z.c;st=z.t;break;}}
    const pct=Math.min(100,Math.max(0,(val-lo)/(hi2-lo)*100)).toFixed(1);
    return `
      <div class="cockpit-card ${cls}">
        <div class="ck-icon">${icon}</div><div class="ck-label">${label}</div>
        <div class="ck-val">${Number(val).toLocaleString('de-DE',{maximumFractionDigits:1})} ${unit}</div>
        <div class="ck-bar-bg"><div class="ck-bar-fill" style="width:${pct}%"></div></div>
        <div class="ck-status">${st}</div></div>`;
  };

  const cockpit = document.createElement('div');
  cockpit.className = 'cockpit-grid';
  cockpit.innerHTML = [
    mkCard('🌡️','Temperatur',   latest.t,  '°C',  5,  35, [{lo:-INF,hi:15,c:'bad',t:'❄️ Zu kalt'},{lo:15,hi:18,c:'warn',t:'Kühl'},{lo:18,hi:24,c:'ok',t:'✅ Behaglich'},{lo:24,hi:26,c:'warn',t:'Warm'},{lo:26,hi:INF,c:'bad',t:'🔥 Zu warm'}]),
    mkCard('💧','Feuchte',       latest.rh, '%',   0, 100, [{lo:-INF,hi:30,c:'bad',t:'🏜️ Sehr trocken'},{lo:30,hi:40,c:'warn',t:'Trocken'},{lo:40,hi:60,c:'ok',t:'✅ Ideal'},{lo:60,hi:70,c:'warn',t:'Feucht'},{lo:70,hi:INF,c:'bad',t:'💦 Zu feucht'}]),
    mkCard('🌿','CO₂',          latest.co2,'ppm',300,2500, [{lo:-INF,hi:800,c:'ok',t:'✅ Frische Luft'},{lo:800,hi:1000,c:'ok',t:'Gut'},{lo:1000,hi:1400,c:'warn',t:'⚠️ Lüften!'},{lo:1400,hi:2000,c:'bad',t:'😮 Schlecht'},{lo:2000,hi:INF,c:'bad',t:'🚨 Kritisch'}]),
    mkCard('🌫️','Taupunkt',    latest.dp, '°C',-15,  30, [{lo:-INF,hi:0,c:'ok',t:'Sehr trocken'},{lo:0,hi:10,c:'ok',t:'✅ Angenehm'},{lo:10,hi:16,c:'ok',t:'Komfortabel'},{lo:16,hi:18,c:'warn',t:'Leicht feucht'},{lo:18,hi:21,c:'warn',t:'⚠️ Schwül'},{lo:21,hi:INF,c:'bad',t:'🥵 Sehr schwül'}]),
    mkCard('🌊','Abs. Feuchte', latest.ah, 'g/m³',0,  20, [{lo:-INF,hi:4,c:'warn',t:'Trocken'},{lo:4,hi:7,c:'ok',t:'OK'},{lo:7,hi:12,c:'ok',t:'✅ Ideal (7–12)'},{lo:12,hi:17,c:'warn',t:'Feucht'},{lo:17,hi:INF,c:'bad',t:'💦 Sehr feucht'}]),
    latest.cs!==null&&latest.cs!==undefined ? `
      <div class="cockpit-card ${latest.cs>=70?'ck-ok':latest.cs>=40?'ck-warn':'ck-bad'}">
        <div class="ck-icon">🏠</div><div class="ck-label">Wohlfühl-Score</div>
        <div class="ck-val">${Math.round(latest.cs)} / 100</div>
        <div class="ck-bar-bg"><div class="ck-bar-fill" style="width:${latest.cs.toFixed(1)}%"></div></div>
        <div class="ck-status">${latest.cs>=70?'✅ Behaglich':latest.cs>=40?'⚠️ Mäßig':'❌ Unbehaglich'}</div></div>` : '',
  ].join('');
  container.appendChild(cockpit);

  /* Stats pills */
  {
    const tTot=derived.filter(d=>d.t!==null).length, tOK=derived.filter(d=>d.t!==null&&d.t>=18&&d.t<=24).length;
    const rTot=derived.filter(d=>d.rh!==null).length, rOK=derived.filter(d=>d.rh!==null&&d.rh>=40&&d.rh<=60).length;
    const cTot=derived.filter(d=>d.co2!==null).length, cOK=derived.filter(d=>d.co2!==null&&d.co2<1000).length;
    const scVals=derived.filter(d=>d.cs!==null).map(d=>d.cs);
    const meanCS=scVals.length?scVals.reduce((a,b)=>a+b,0)/scVals.length:null;
    const mkPill=(txt,pct,t)=>{ const cls=pct>=t[0]?'good':pct>=t[1]?'warn':'bad'; return `<span class="comfort-pill ${cls}">${txt}: <b>${Math.round(pct)} %</b></span>`; };
    const pills=document.createElement('div'); pills.className='comfort-info';
    pills.innerHTML=[
      tTot?mkPill('🌡️ Behagliche Temp (18–24°C)',tOK/tTot*100,[70,40]):'',
      rTot?mkPill('💧 Ideale Feuchte (40–60%)',rOK/rTot*100,[70,40]):'',
      cTot?mkPill('🌿 CO₂ &lt; 1000 ppm',cOK/cTot*100,[80,50]):'',
      meanCS!==null?`<span class="comfort-pill ${meanCS>=70?'good':meanCS>=40?'warn':'bad'}">🏠 Ø Wohlfühl-Score: <b>${Math.round(meanCS)}/100</b></span>`:'',
    ].join('');
    container.appendChild(pills);
  }

  /* Chart 1: Taupunkt */
  if (inTempSen && inHumSen) {
    const card=document.createElement('div'); card.className='chart-card';
    card.innerHTML=`<div class="chart-header"><h2>🌫️ Taupunkt (innen) vs. Außentemperatur – Kondensationsrisiko</h2><span class="unit-badge">°C</span></div>
      <div class="chart-wrap tall"><canvas id="indDewChart"></canvas></div>
      <div class="chart-hint">Wenn Außen-/Wandtemperatur unter den Taupunkt fällt → Kondenswasser / Schimmelrisiko. Berechnung nach Magnus-Formel (WMO). Tooltip zeigt Warnung bei Unterschreitung.</div>`;
    container.appendChild(card);
    const dpPts=derived.filter(d=>d.dp!==null).map(d=>({x:d.ts,y:d.dp}));
    const tIPts=derived.filter(d=>d.t!==null).map(d=>({x:d.ts,y:d.t}));
    const tOPts=outTempH.map(d=>({x:d.ts,y:d.value}));
    const dsDew=[
      {label:'Taupunkt Innen',data:dpPts,borderColor:'#bc8cff',backgroundColor:'#bc8cff11',borderWidth:2,pointRadius:0,tension:0.3,fill:false},
      {label:`Temp. Innen (${indoorModule})`,data:tIPts,borderColor:'#4dc9f6',backgroundColor:'transparent',borderWidth:1.5,pointRadius:0,tension:0.3,fill:false},
    ];
    if(tOPts.length) dsDew.push({label:`Temp. Außen (${outdoorModuleInd})`,data:tOPts,borderColor:'#f59e0b',backgroundColor:'transparent',borderWidth:1.5,pointRadius:0,tension:0.3,fill:false,borderDash:[5,3]});
    indoorCharts.dew=new Chart(document.getElementById('indDewChart').getContext('2d'),{
      type:'line', data:{datasets:dsDew},
      options:{...commonChartOpts('°C',true),
        plugins:{...commonChartOpts('°C',true).plugins,
          tooltip:{backgroundColor:'#161b22',borderColor:'#30363d',borderWidth:1,titleColor:'#e6edf3',bodyColor:'#8b949e',
            callbacks:{
              label:ctx=>` ${ctx.dataset.label}: ${ctx.parsed.y?.toLocaleString('de-DE',{maximumFractionDigits:1})} °C`,
              afterBody:items=>{
                const dp=items.find(i=>i.dataset.label.includes('Taupunkt'))?.parsed.y;
                const tO=items.find(i=>i.dataset.label.includes('Außen'))?.parsed.y;
                return(dp!==undefined&&tO!==undefined&&tO<dp)?['⚠️ Kondensationsrisiko bei diesen Bedingungen!']:[];
              },
            },
          },
        },
      },
    });
  }

  /* Chart 2: CO2 mit Ampelzonen */
  if (inCO2Sen) {
    const card=document.createElement('div'); card.className='chart-card';
    card.innerHTML=`<div class="chart-header"><h2>🌿 CO₂-Verlauf mit Luftqualitäts-Ampel</h2><span class="unit-badge">ppm</span></div>
      <div class="chart-wrap tall"><canvas id="indCO2Chart"></canvas></div>
      <div class="chart-hint"><span style="color:var(--green)">■ &lt;800 Frische Luft</span> &nbsp; <span style="color:var(--orange)">■ 800–1400 Lüften empfohlen</span> &nbsp; <span style="color:var(--red)">■ &gt;1400 Schlechte Luft</span></div>`;
    container.appendChild(card);
    const tsF=inCO2H[0]?.ts, tsL=inCO2H[inCO2H.length-1]?.ts;
    const thrLine2=(y,col,lbl)=>({label:lbl,data:tsF?[{x:tsF,y},{x:tsL,y}]:[],borderColor:col,borderWidth:1.5,borderDash:[8,4],pointRadius:0,fill:false});
    indoorCharts.co2=new Chart(document.getElementById('indCO2Chart').getContext('2d'),{
      type:'line', data:{datasets:[
        {label:'CO₂',data:inCO2H.map(d=>({x:d.ts,y:d.value})),
         segment:{borderColor:ctx=>ctx.p1?.parsed?.y<800?'#3fb950':ctx.p1?.parsed?.y<1400?'#d29922':'#f85149'},
         backgroundColor:'transparent',borderWidth:2,pointRadius:0,tension:0.3,fill:false},
        thrLine2(800,'#3fb95066','✅ Gut – Grenze (800 ppm)'),
        thrLine2(1000,'#d2992266','⚠️ Lüften (1000 ppm)'),
        thrLine2(1400,'#f8514966','❌ Schlechte Luft (1400 ppm)'),
      ]},
      options:{responsive:true,maintainAspectRatio:false,interaction:{mode:'index',intersect:false},
        scales:{
          x:{type:'time',time:{tooltipFormat:'dd.MM.yyyy HH:mm',displayFormats:{hour:'dd.MM HH:mm',day:'dd.MM.yy',month:'MMM yy'}},grid:{color:'#30363d'},ticks:{color:'#8b949e',maxTicksLimit:8}},
          y:{min:350,grid:{color:'#30363d'},ticks:{color:'#8b949e',callback:v=>v.toLocaleString('de-DE')+' ppm'}},
        },
        plugins:{
          legend:{labels:{color:'#e6edf3',boxWidth:12,padding:10}},
          tooltip:{backgroundColor:'#161b22',borderColor:'#30363d',borderWidth:1,titleColor:'#e6edf3',bodyColor:'#8b949e',
            callbacks:{label:ctx=>{if(ctx.datasetIndex!==0) return null; const v=ctx.parsed.y; return ` CO₂: ${v?.toLocaleString('de-DE')} ppm  ${v<800?'✅ Frisch':v<1000?'🟢 Gut':v<1400?'⚠️ Lüften':'❌ Schlecht'}`;}}},
          zoom:{zoom:{wheel:{enabled:true},pinch:{enabled:true},mode:'x'},pan:{enabled:true,mode:'x'}},
        },
      },
    });
  }

  /* Chart 3: Innen vs. Außen */
  if((inTempSen&&outTempSen)||(inHumSen&&outHumSen)){
    const card=document.createElement('div'); card.className='chart-card';
    card.innerHTML=`<div class="chart-header"><h2>🌤️ Innen vs. Außen – Temperatur &amp; relative Feuchte</h2><span class="unit-badge">°C / %</span></div>
      <div class="chart-wrap tall"><canvas id="indInOutChart"></canvas></div>
      <div class="chart-hint">Temperaturdifferenz Innen–Außen zeigt Heizaufwand und Wärmedämmung. Feuchteunterschied deutet auf Lüftungsverhalten hin. Doppelachse: °C links, % rechts.</div>`;
    container.appendChild(card);
    const ds3=[];
    if(inTempSen)  ds3.push({label:`T Innen (${indoorModule})`,  data:inTempH.map(d=>({x:d.ts,y:d.value})),yAxisID:'yTIO',borderColor:'#4dc9f6',backgroundColor:'transparent',borderWidth:2,pointRadius:0,tension:0.3});
    if(outTempSen) ds3.push({label:`T Außen (${outdoorModuleInd})`,data:outTempH.map(d=>({x:d.ts,y:d.value})),yAxisID:'yTIO',borderColor:'#f59e0b',backgroundColor:'transparent',borderWidth:1.5,pointRadius:0,tension:0.3,borderDash:[5,3]});
    if(inHumSen)   ds3.push({label:`RH Innen (${indoorModule})`, data:inHumH.map(d=>({x:d.ts,y:d.value})),yAxisID:'yRHIO',borderColor:'#3fb950',backgroundColor:'transparent',borderWidth:2,pointRadius:0,tension:0.3});
    if(outHumSen)  ds3.push({label:`RH Außen (${outdoorModuleInd})`,data:outHumH.map(d=>({x:d.ts,y:d.value})),yAxisID:'yRHIO',borderColor:'#8b949e',backgroundColor:'transparent',borderWidth:1.5,pointRadius:0,tension:0.3,borderDash:[5,3]});
    indoorCharts.inout=new Chart(document.getElementById('indInOutChart').getContext('2d'),{
      type:'line', data:{datasets:ds3},
      options:{responsive:true,maintainAspectRatio:false,interaction:{mode:'index',intersect:false},
        scales:{
          x:{type:'time',time:{tooltipFormat:'dd.MM.yyyy HH:mm',displayFormats:{hour:'dd.MM HH:mm',day:'dd.MM.yy',month:'MMM yy'}},grid:{color:'#30363d'},ticks:{color:'#8b949e',maxTicksLimit:8}},
          yTIO:{type:'linear',position:'left',grid:{color:'#30363d'},ticks:{color:'#4dc9f6',callback:v=>v.toLocaleString('de-DE',{maximumFractionDigits:1})+'°C'},title:{display:true,text:'Temperatur °C',color:'#4dc9f6'}},
          yRHIO:{type:'linear',position:'right',min:0,max:100,grid:{display:false},ticks:{color:'#3fb950',callback:v=>v+'%'},title:{display:true,text:'Feuchte %',color:'#3fb950'}},
        },
        plugins:{legend:{labels:{color:'#e6edf3',boxWidth:12,padding:12}},
          tooltip:{backgroundColor:'#161b22',borderColor:'#30363d',borderWidth:1,titleColor:'#e6edf3',bodyColor:'#8b949e'},
          zoom:{zoom:{wheel:{enabled:true},pinch:{enabled:true},mode:'x'},pan:{enabled:true,mode:'x'}},
        },
      },
    });
  }

  /* Chart 4: Absolute Luftfeuchtigkeit */
  if(inTempSen&&inHumSen){
    const card=document.createElement('div'); card.className='chart-card';
    card.innerHTML=`<div class="chart-header"><h2>🌊 Absolute Luftfeuchtigkeit mit Komfortband (7–12 g/m³)</h2><span class="unit-badge">g/m³</span></div>
      <div class="chart-wrap tall"><canvas id="indAHChart"></canvas></div>
      <div class="chart-hint">Absolute Feuchte ist temperaturunabhängig und für Gesundheit/Komfort entscheidend. Ideal: 7–12 g/m³ (grünes Band). Unter 4 g/m³ → Schleimhäute trocken. Outdoor-Vergleich zeigt, ob Lüften hilft.</div>`;
    container.appendChild(card);
    const ahPts=derived.filter(d=>d.ah!==null).map(d=>({x:d.ts,y:d.ah}));
    const ahOutPts=derived.filter(d=>d.ahOut!==null).map(d=>({x:d.ts,y:d.ahOut}));
    const tsFA=ahPts[0]?.x, tsLA=ahPts[ahPts.length-1]?.x;
    const ds4=[
      {label:'AH Innen',data:ahPts,borderColor:'#3fb950',backgroundColor:'transparent',borderWidth:2,pointRadius:0,tension:0.3,fill:false,order:1},
      {label:'Komfortband oben (12)',data:tsFA?[{x:tsFA,y:12},{x:tsLA,y:12}]:[],borderColor:'#3fb95044',borderWidth:1,pointRadius:0,fill:'+1',backgroundColor:'#3fb95015',borderDash:[4,4],order:3},
      {label:'Komfortband unten (7)', data:tsFA?[{x:tsFA,y:7},{x:tsLA,y:7}]:[],borderColor:'#3fb95044',borderWidth:1,pointRadius:0,fill:false,backgroundColor:'transparent',borderDash:[4,4],order:4},
    ];
    if(ahOutPts.length) ds4.splice(1,0,{label:'AH Außen',data:ahOutPts,borderColor:'#f59e0b',backgroundColor:'transparent',borderWidth:1.5,pointRadius:0,tension:0.3,fill:false,order:2,borderDash:[5,3]});
    indoorCharts.ah=new Chart(document.getElementById('indAHChart').getContext('2d'),{
      type:'line', data:{datasets:ds4},
      options:{...commonChartOpts('g/m³',true),
        plugins:{...commonChartOpts('g/m³',true).plugins,
          legend:{labels:{color:'#e6edf3',boxWidth:12,padding:10,filter:it=>!it.text.includes('Komfortband')}},
        },
      },
    });
  }

  /* Chart 5: Wohlfühl-Score täglich */
  if(derived.some(d=>d.cs!==null)){
    const card=document.createElement('div'); card.className='chart-card';
    card.innerHTML=`<div class="chart-header"><h2>🏠 Täglicher Wohlfühl-Score</h2><span class="unit-badge">0–100</span></div>
      <div class="chart-wrap tall"><canvas id="indScoreChart"></canvas></div>
      <div class="chart-hint">Zusammengesetzt: Temperatur 30 % (18–24°C optimal) + Feuchte 30 % (40–60% optimal) + CO₂ 40 % (&lt;1000 ppm optimal). Grün ≥70, Gelb ≥40, Rot &lt;40.</div>`;
    container.appendChild(card);
    const dayMap5={};
    derived.forEach(d=>{
      if(d.cs===null) return;
      const dt=new Date(d.ts);
      const k=`${dt.getFullYear()}-${dt.getMonth()}-${dt.getDate()}`;
      if(!dayMap5[k]) dayMap5[k]={ts:new Date(dt.getFullYear(),dt.getMonth(),dt.getDate()).getTime(),vals:[]};
      dayMap5[k].vals.push(d.cs);
    });
    const spts=Object.values(dayMap5).sort((a,b)=>a.ts-b.ts).map(d=>({x:d.ts,y:Math.round(d.vals.reduce((a,b)=>a+b,0)/d.vals.length)}));
    const scol=spts.map(d=>d.y>=70?'#3fb950':d.y>=40?'#d29922':'#f85149');
    indoorCharts.score=new Chart(document.getElementById('indScoreChart').getContext('2d'),{
      type:'bar', data:{datasets:[{label:'Wohlfühl-Score',data:spts,backgroundColor:scol,borderColor:scol,borderWidth:0,borderRadius:3}]},
      options:{responsive:true,maintainAspectRatio:false,
        scales:{
          x:{type:'time',time:{tooltipFormat:'dd.MM.yyyy',displayFormats:{day:'dd.MM',month:'MMM yy'}},grid:{color:'#30363d'},ticks:{color:'#8b949e',maxTicksLimit:10}},
          y:{min:0,max:100,grid:{color:'#30363d'},ticks:{color:'#8b949e',callback:v=>v+'/100'},title:{display:true,text:'Komfort-Score',color:'#8b949e'}},
        },
        plugins:{legend:{display:false},tooltip:{backgroundColor:'#161b22',borderColor:'#30363d',borderWidth:1,titleColor:'#e6edf3',bodyColor:'#8b949e',
          callbacks:{label:ctx=>` Score: ${ctx.parsed.y}/100 — ${ctx.parsed.y>=70?'Behaglich':ctx.parsed.y>=40?'Mäßig':'Unbehaglich'}`}}},
      },
    });
  }

  /* Chart 6: Tages-Profil 24h */
  if((inTempSen||inHumSen||inCO2Sen)&&derived.length){
    const card=document.createElement('div'); card.className='chart-card';
    card.innerHTML=`<div class="chart-header"><h2>🕐 Typisches Tages-Profil (Ø je Stunde, alle Tage)</h2><span class="unit-badge">24h</span></div>
      <div class="chart-wrap tall"><canvas id="indHourChart"></canvas></div>
      <div class="chart-hint">Durchschnittlicher Tagesverlauf über alle Tage im Zeitraum. CO₂-Anstieg = Raumnutzung. Temperatur-Morgenanstieg = Heizsystem. Feuchte-Abfall = Lüftungsverhalten.</div>`;
    container.appendChild(card);
    const hB={t:Array.from({length:24},()=>[]),rh:Array.from({length:24},()=>[]),co2:Array.from({length:24},()=>[])};
    derived.forEach(d=>{const h=new Date(d.ts).getHours(); if(d.t!==null)hB.t[h].push(d.t); if(d.rh!==null)hB.rh[h].push(d.rh); if(d.co2!==null)hB.co2[h].push(d.co2);});
    const mn24=arr=>arr.map(vs=>vs.length?Math.round(vs.reduce((a,b)=>a+b,0)/vs.length*10)/10:null);
    const lbH=Array.from({length:24},(_,i)=>i.toString().padStart(2,'0')+':00');
    const ds6=[]; const scH6={x:{grid:{color:'#30363d'},ticks:{color:'#8b949e'}}};
    if(inTempSen){ds6.push({label:'Temperatur (°C)',data:mn24(hB.t),borderColor:'#4dc9f6',borderWidth:2.5,pointRadius:3,tension:0.4,fill:false,yAxisID:'yT6'}); scH6.yT6={type:'linear',position:'left',grid:{color:'#30363d44'},ticks:{color:'#4dc9f6',callback:v=>v+'°C'},title:{display:true,text:'°C',color:'#4dc9f6'}};}
    if(inHumSen) {ds6.push({label:'Feuchte (%)',   data:mn24(hB.rh),borderColor:'#3fb950',borderWidth:2.5,pointRadius:3,tension:0.4,fill:false,yAxisID:'yRH6'}); scH6.yRH6={type:'linear',position:'right',min:0,max:100,grid:{display:false},ticks:{color:'#3fb950',callback:v=>v+'%'},title:{display:true,text:'%',color:'#3fb950'}};}
    if(inCO2Sen) {ds6.push({label:'CO₂ (ppm)',    data:mn24(hB.co2),borderColor:'#bc8cff',borderWidth:2.5,pointRadius:3,tension:0.4,fill:false,yAxisID:'yCO26'}); scH6.yCO26={type:'linear',position:'right',grid:{display:false},ticks:{color:'#bc8cff',callback:v=>v.toLocaleString('de-DE')+' ppm'},title:{display:true,text:'CO₂ ppm',color:'#bc8cff'},...(inHumSen?{display:false}:{})};}
    indoorCharts.hour=new Chart(document.getElementById('indHourChart').getContext('2d'),{
      type:'line', data:{labels:lbH,datasets:ds6},
      options:{responsive:true,maintainAspectRatio:false,interaction:{mode:'index',intersect:false},scales:scH6,
        plugins:{legend:{labels:{color:'#e6edf3',boxWidth:12,padding:12}},tooltip:{backgroundColor:'#161b22',borderColor:'#30363d',borderWidth:1,titleColor:'#e6edf3',bodyColor:'#8b949e'}},
      },
    });
  }

  /* Chart 7: Komfort-Heatmap Wochentag × Stunde */
  if(derived.some(d=>d.cs!==null)){
    const card=document.createElement('div'); card.className='chart-card';
    card.innerHTML=`<div class="chart-header"><h2>📅 Komfort-Muster: Wochentag × Uhrzeit</h2><span class="unit-badge">Ø Score</span></div>
      <div id="indComfortHM" style="overflow-x:auto;margin-top:10px;"></div>
      <div class="chart-hint">Ø Wohlfühl-Score je Wochentag und Stunde. Grün = ideal, Gelb = mäßig, Rot = unbehaglich. Zeigt typische Schlaf-, Arbeits- und Freizeitmuster.</div>`;
    container.appendChild(card);
    const grid7=Array.from({length:7},()=>Array.from({length:24},()=>[])); 
    derived.forEach(d=>{if(d.cs===null) return; const dt=new Date(d.ts); grid7[(dt.getDay()+6)%7][dt.getHours()].push(d.cs);});
    const mn7=grid7.map(r=>r.map(vs=>vs.length?Math.round(vs.reduce((a,b)=>a+b,0)/vs.length):null));
    const dow7=['Mo','Di','Mi','Do','Fr','Sa','So'];
    const hrs7=Array.from({length:24},(_,i)=>i.toString().padStart(2,'0'));
    const sCol7=s=>s===null?'#1a1f27':s>=70?`rgba(63,185,80,${0.35+s/200})`:(s>=40?`rgba(210,153,34,${0.35+s/200})`:`rgba(248,81,73,${0.35+s/200})`);
    let tbl7=`<table class="heatmap" style="min-width:640px"><tr><th></th>${hrs7.map(h=>`<th style="min-width:25px;font-size:10px;">${h}</th>`).join('')}</tr>`;
    dow7.forEach((d,di)=>{
      tbl7+=`<tr><th style="text-align:right;padding-right:6px;color:var(--muted);white-space:nowrap">${d}</th>`;
      hrs7.forEach((_,hi)=>{const v=mn7[di][hi]; tbl7+=`<td class="cell" style="background:${sCol7(v)};min-width:25px;" title="${d} ${hi}:00 – ${v!==null?v+'/100':'–'}">${v!==null?v:''}</td>`;});
      tbl7+='</tr>';
    });
    tbl7+='</table>';
    document.getElementById('indComfortHM').innerHTML=tbl7;
    const legDiv7=document.createElement('div'); legDiv7.className='heatmap-legend';
    legDiv7.innerHTML=`<span>Unbehaglich</span><div style="width:180px;height:10px;border-radius:3px;background:linear-gradient(to right,#f85149,#d29922,#3fb950)"></div><span>Ideal</span>`;
    card.appendChild(legDiv7);
  }

  /* Chart 8: Gefühlte Temperatur (Heat Index) */
  if(inTempSen&&inHumSen&&derived.some(d=>d.hi!==null&&d.t!==null&&Math.abs(d.hi-d.t)>0.3)){
    const card=document.createElement('div'); card.className='chart-card';
    card.innerHTML=`<div class="chart-header"><h2>🌡️ Gefühlte Temperatur (Heat Index) vs. gemessene Temperatur</h2><span class="unit-badge">°C</span></div>
      <div class="chart-wrap"><canvas id="indHIChart"></canvas></div>
      <div class="chart-hint">Heat Index (Rothfusz-Formel) berücksichtigt Luftfeuchtigkeit bei der Wärmewahrnehmung. Relevant ab ~26°C und &gt;40% relativer Feuchte.</div>`;
    container.appendChild(card);
    indoorCharts.hi=new Chart(document.getElementById('indHIChart').getContext('2d'),{
      type:'line', data:{datasets:[
        {label:'Gefühlte Temp. (Heat Index)',data:derived.filter(d=>d.hi!==null).map(d=>({x:d.ts,y:d.hi})),borderColor:'#f85149',backgroundColor:'#f8514912',borderWidth:2,pointRadius:0,tension:0.3,fill:false},
        {label:'Gemessene Temperatur',data:derived.filter(d=>d.t!==null).map(d=>({x:d.ts,y:d.t})),borderColor:'#4dc9f6',backgroundColor:'transparent',borderWidth:1.5,pointRadius:0,tension:0.3,fill:false,borderDash:[5,3]},
      ]},
      options:commonChartOpts('°C',true),
    });
  }
}

/* ── Apply ── */
function renderCurrentView() {
  if (currentView === 'charts')    renderCharts(lastFiltered);
  else if (currentView === 'records')  renderRecords(lastFilteredRaw);
  else if (currentView === 'trends')   renderTrends(lastFilteredRaw);
  else if (currentView === 'longterm') renderLongterm(lastFilteredRaw);
  else if (currentView === 'compare')  renderCompare(lastFilteredRaw);
  else if (currentView === 'climate')  renderClimate(lastFilteredRaw);
  else if (currentView === 'extremes') renderExtremes(lastFilteredRaw);
  else if (currentView === 'indoor')   renderIndoor(lastFilteredRaw);
}

function applyFilters(options = {}) {
  const from = document.getElementById('dateFrom').value;
  const to   = document.getElementById('dateTo').value;
  const tsFrom = from ? new Date(from).getTime() : 0;
  const tsTo   = to   ? new Date(to).getTime() + 86399999 : Infinity;

  const rawFiltered = FLAT.filter(d =>
    d.ts >= tsFrom && d.ts <= tsTo &&
    selModules.has(d.module) &&
    selSensors.has(d.sensor)
  );

  lastFilteredRaw = rawFiltered;
  lastFiltered = aggregateData(rawFiltered, aggMode);
  if (!options.deferRender) renderCurrentView();

  document.getElementById('fileInfo').textContent =
    `${lastFiltered.length.toLocaleString('de-DE')} Punkte (${aggMode}) · ${rawFiltered.length.toLocaleString('de-DE')} Roh · ${from} bis ${to}`;
}

function resetFilters() {
  const bounds = getDateBounds();
  document.getElementById('dateFrom').value = bounds.min;
  document.getElementById('dateTo').value   = bounds.max;
  applyDefaultSelection();
  buildChips('moduleChips', PAYLOAD.modules, selModules);
  buildChips('sensorChips', PAYLOAD.sensors, selSensors);
  applyFilters();
}

/* ── Init ── */
window.addEventListener('load', () => {
  const bounds = getDateBounds();
  document.getElementById('dateFrom').value = bounds.min;
  document.getElementById('dateTo').value   = bounds.max;
  applyDefaultSelection();
  buildChips('moduleChips', PAYLOAD.modules, selModules);
  buildChips('sensorChips', PAYLOAD.sensors, selSensors);
  if (PAYLOAD.locked_enc) {
    document.getElementById('indoorLockPanel').style.display = 'block';
  }
  document.getElementById('fileInfo').textContent =
    `${FLAT.length.toLocaleString('de-DE')} Datenpunkte geladen · ${bounds.min} bis ${bounds.max}`;
  applyFilters({ deferRender: true });
  requestAnimationFrame(() => {
    requestAnimationFrame(() => renderCurrentView());
  });
});
</script>
</body>
</html>
"""


def generate_html(payload: dict) -> str:
    json_str = json.dumps(payload, ensure_ascii=False, separators=(',', ':'))
    return HTML_TEMPLATE.replace('__PAYLOAD__', json_str)


def open_in_browser(html: str, title: str = 'netatmo_viewer') -> str:
    fd, path = tempfile.mkstemp(suffix='.html', prefix=title + '_')
    with os.fdopen(fd, 'w', encoding='utf-8') as f:
        f.write(html)
    webbrowser.open(f'file:///{path.replace(os.sep, "/")}')
    return path


def default_pages_index_path() -> str:
    """Zielpfad für GitHub Pages: docs/index.html im Projektordner."""
    folder = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(folder, 'docs', 'index.html')


def export_html(html: str, path: str | None = None) -> str:
    """Schreibt die Viewer-HTML als veröffentlichbare statische App."""
    target = path or default_pages_index_path()
    os.makedirs(os.path.dirname(target), exist_ok=True)
    with open(target, 'w', encoding='utf-8') as f:
        f.write(html)

    nojekyll_path = os.path.join(os.path.dirname(target), '.nojekyll')
    with open(nojekyll_path, 'a', encoding='utf-8'):
        pass
    return target


# ─────────────────────────────────────────────
#  tkinter GUI
# ─────────────────────────────────────────────

class ViewerApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        root.title('Netatmo Daten-Viewer')
        root.geometry('560x560')
        root.resizable(False, False)

        style = ttk.Style()
        style.theme_use('clam')

        self.filepath = tk.StringVar()
        self.status = tk.StringVar(value='Bereit.')
        self._data = []
        self._source_files = []  # Liste der bereits eingelesenen Dateipfade
        self._direct_import_running = False

        self._build_ui()
        self._auto_load_cache()

    def _build_ui(self):
        pad = {'padx': 16, 'pady': 8}

        # ── Datei ──
        file_frame = ttk.LabelFrame(
            self.root, text='Excel / CSV Datei', padding=10)
        file_frame.pack(fill=tk.X, **pad)
        file_frame.columnconfigure(0, weight=1)
        ttk.Entry(file_frame, textvariable=self.filepath).grid(
            row=0, column=0, columnspan=3, sticky='ew')
        ttk.Button(file_frame, text='Durchsuchen…', command=self._browse).grid(
            row=0, column=3, padx=(6, 0))

        ttk.Button(file_frame, text='Laden (ersetzen)',
                   command=self._load_replace).grid(
            row=1, column=0, sticky='ew', pady=(8, 0))
        ttk.Button(file_frame, text='➕ Delta hinzufügen',
                   command=self._load_delta).grid(
            row=1, column=1, sticky='ew', padx=(6, 0), pady=(8, 0))
        ttk.Button(file_frame, text='🗑 Cache löschen',
                   command=self._clear_cache).grid(
            row=1, column=2, sticky='ew', padx=(6, 0), pady=(8, 0))
        btn_cache_to_sqlite = ttk.Button(
            file_frame, text='Cache in SQLite kopieren',
            command=self._copy_cache_to_archive)
        btn_cache_to_sqlite.grid(
            row=2, column=0, columnspan=2, sticky='ew', pady=(6, 0))
        btn_open_archive = ttk.Button(
            file_frame, text='SQLite-Archiv öffnen',
            command=self._open_archive)
        btn_open_archive.grid(
            row=2, column=2, sticky='ew', padx=(6, 0), pady=(6, 0))
        file_frame.columnconfigure(0, weight=1)
        file_frame.columnconfigure(1, weight=1)
        file_frame.columnconfigure(2, weight=1)

        # ── Direktimport ──
        direct_frame = ttk.LabelFrame(
            self.root, text='Direkt von Netatmo', padding=10)
        direct_frame.pack(fill=tk.X, **pad)
        direct_frame.columnconfigure(0, weight=1)
        ttk.Label(
            direct_frame,
            text='Lädt ohne Excel-Zwischenschritt und fügt neue Punkte dem Buffer hinzu.',
            foreground='gray').grid(row=0, column=0, sticky='w')
        self.btn_direct = ttk.Button(
            direct_frame, text='Netatmo direkt laden (Delta)',
            command=self._start_direct_import)
        self.btn_direct.grid(row=1, column=0, sticky='ew', pady=(8, 0))
        self._autopush_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            direct_frame,
            text='Nach Import automatisch → GitHub Pages aktualisieren (git push)',
            variable=self._autopush_var).grid(row=2, column=0, sticky='w', pady=(6, 0))

        # ── Info ──
        info_frame = ttk.LabelFrame(
            self.root, text='Geladene Daten', padding=10)
        info_frame.pack(fill=tk.X, **pad)
        self.info_text = tk.Text(info_frame, height=8, state=tk.DISABLED,
                                 background='#f5f5f5', relief='flat', wrap=tk.WORD)
        self.info_text.pack(fill=tk.X)

        # ── Öffnen ──
        action_frame = ttk.Frame(self.root)
        action_frame.pack(pady=(0, 8))
        self.btn_open = ttk.Button(
            action_frame, text='Im Browser öffnen  ➜', command=self._open,
            state=tk.DISABLED)
        self.btn_open.pack(side=tk.LEFT, padx=(0, 8))
        self.btn_open_docs = ttk.Button(
            action_frame, text='Im Browser (GitHub snapshot)',
            command=self._open_docs_snapshot, state=tk.NORMAL)
        self.btn_open_docs.pack(side=tk.LEFT, padx=(8, 0))
        self.btn_export_pages = ttk.Button(
            action_frame, text='Für GitHub Pages exportieren',
            command=self._export_for_github_pages, state=tk.DISABLED)
        self.btn_export_pages.pack(side=tk.LEFT)

        # ── Status ──
        ttk.Label(self.root, textvariable=self.status, foreground='gray',
                  font=('Segoe UI', 9)).pack()

    # ── Persistenz ──
    def _auto_load_cache(self):
        data, meta = load_cache()
        if data:
            self._data = data
            self._source_files = (meta or {}).get(
                'sources', []) if meta else []
            saved_at = (meta or {}).get('saved_at', '?') if meta else '?'
            self._refresh_info(
                cache_note=f'aus Cache (gespeichert: {saved_at})')
            self.btn_open.config(state=tk.NORMAL)
            self.btn_export_pages.config(state=tk.NORMAL)
            self.status.set(
                f'{len(self._data):,} Punkte aus Cache geladen.')
            return

        data, meta = load_archive()
        if data:
            self._data = data
            self._source_files = (meta or {}).get(
                'sources', []) if meta else []
            self._refresh_info(cache_note='aus SQLite-Archiv')
            self.btn_open.config(state=tk.NORMAL)
            self.btn_export_pages.config(state=tk.NORMAL)
            self._save_to_cache()
            self.status.set(
                f'{len(self._data):,} Punkte aus SQLite-Archiv geladen.')
            return

        self.status.set(
            'Kein Cache/Archiv gefunden – bitte eine Datei laden.')

    def _save_to_cache(self):
        meta = {
            'saved_at': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'sources': self._source_files,
            'count':   len(self._data),
        }
        try:
            save_cache(self._data, meta)
        except OSError as e:
            messagebox.showwarning(
                'Cache nicht gespeichert', f'Konnte Cache nicht schreiben:\n{e}')

    def _save_to_archive(self, data: list, source: str = '') -> tuple:
        try:
            return save_archive(data, source)
        except (OSError, sqlite3.Error, ValueError, TypeError, KeyError) as e:
            messagebox.showwarning(
                'Archiv nicht gespeichert',
                f'Konnte SQLite-Archiv nicht schreiben:\n{e}')
            return archive_path(), 0, len(data)

    def _copy_cache_to_archive(self):
        data, meta = load_cache()
        if not data:
            messagebox.showinfo(
                'Cache kopieren',
                f'Kein Cache gefunden:\n{cache_path()}')
            return

        sources = (meta or {}).get('sources', []) if meta else []
        source_label = 'Cache-Migration aus cache.pkl.gz'
        if sources:
            shown = ', '.join(os.path.basename(p) for p in sources[:5])
            if len(sources) > 5:
                shown += f' (+{len(sources) - 5} weitere)'
            source_label = f'Cache-Migration: {shown}'

        archive_file, added, dupes = self._save_to_archive(data, source_label)
        self._data = data
        self._source_files = sources
        self._refresh_info(
            cache_note=f'Cache → SQLite: +{added:,} neu / {dupes:,} bekannt')
        self.btn_open.config(state=tk.NORMAL)
        self.btn_export_pages.config(state=tk.NORMAL)
        self.status.set(
            f'Cache in SQLite kopiert: +{added:,} neu / {dupes:,} bekannt.')
        messagebox.showinfo(
            'Cache kopiert',
            f'{len(data):,} Cache-Punkte verarbeitet.\n'
            f'Neu im Archiv: {added:,}\n'
            f'Bereits vorhanden: {dupes:,}\n\n'
            f'{archive_file}')

    def _clear_cache(self):
        if not messagebox.askyesno(
                'Cache löschen',
                'Cache-Datei wirklich löschen? '
                'Die aktuell geladenen Daten bleiben im Speicher.'):
            return
        if clear_cache():
            self.status.set('Cache gelöscht.')
        else:
            self.status.set('Kein Cache vorhanden.')

    def _open_archive(self):
        path = archive_path()
        if not os.path.isfile(path):
            data, _meta = load_cache()
            if data and messagebox.askyesno(
                    'SQLite-Archiv',
                    'Noch keine Archivdatei vorhanden. Soll der vorhandene '
                    f'Cache jetzt nach SQLite kopiert werden?\n\n{path}'):
                self._copy_cache_to_archive()
            else:
                messagebox.showinfo(
                    'SQLite-Archiv',
                    'Noch keine Archivdatei vorhanden. Sie wird beim nächsten '
                    f'Import angelegt:\n{path}')
                return
            if not os.path.isfile(path):
                return
        try:
            os.startfile(path)
        except OSError as e:
            messagebox.showerror(
                'SQLite-Archiv öffnen',
                'Konnte die Archivdatei nicht öffnen. Installieren Sie z.B. '
                'DB Browser for SQLite oder öffnen Sie die Datei manuell:\n'
                f'{path}\n\n{e}')

    # ── Datei-Handling ──
    def _browse(self):
        filetypes = [('Excel / CSV', '*.xlsx *.xls *.csv'),
                     ('Alle Dateien', '*.*')]
        path = filedialog.askopenfilename(filetypes=filetypes)
        if path:
            self.filepath.set(path)

    def _read_file_or_warn(self, path: str):
        if not path or not os.path.isfile(path):
            messagebox.showerror(
                'Fehler', 'Bitte eine gültige Datei auswählen.')
            return None
        self.status.set('Lese Datei…')
        self.root.update()
        try:
            return load_data(path)
        except (ValueError, ImportError, OSError, KeyError) as e:
            messagebox.showerror('Fehler beim Laden', str(e))
            self.status.set('Fehler beim Laden.')
            return None

    def _load_replace(self):
        path = self.filepath.get().strip()
        new_data = self._read_file_or_warn(path)
        if new_data is None:
            return
        self._data = new_data
        self._source_files = [path]
        archive_file, arch_added, arch_dupes = self._save_to_archive(
            new_data, path)
        self._refresh_info(cache_note='neu eingelesen')
        self.btn_open.config(state=tk.NORMAL)
        self.btn_export_pages.config(state=tk.NORMAL)
        self._save_to_cache()
        self.status.set(
            f'{len(self._data):,} Punkte geladen. '
            f'Archiv: +{arch_added:,} neu / {arch_dupes:,} bekannt '
            f'({os.path.basename(archive_file)}).')

    def _load_delta(self):
        path = self.filepath.get().strip()
        new_data = self._read_file_or_warn(path)
        if new_data is None:
            return
        before = len(self._data)
        self._data, added, dupes = merge_data(self._data, new_data)
        if path not in self._source_files:
            self._source_files.append(path)
        archive_file, arch_added, arch_dupes = self._save_to_archive(
            new_data, path)
        self._refresh_info(
            cache_note=f'Δ: +{added:,} neu / {dupes:,} Duplikate ignoriert')
        self.btn_open.config(state=tk.NORMAL)
        self.btn_export_pages.config(state=tk.NORMAL)
        self._save_to_cache()
        self.status.set(
            f'Delta-Import: {before:,} → {len(self._data):,} '
            f'(+{added:,} neu, {dupes:,} Duplikate übersprungen). '
            f'Archiv: +{arch_added:,} neu / {arch_dupes:,} bekannt '
            f'({os.path.basename(archive_file)}).')

    def _default_direct_start_date(self) -> datetime.datetime:
        if not self._data:
            return datetime.datetime(datetime.datetime.now().year, 1, 1)
        last_ts = max(d['ts'] for d in self._data)
        last_dt = datetime.datetime.fromtimestamp(last_ts / 1000)
        return datetime.datetime.combine(last_dt.date(), datetime.time())

    def _start_direct_import(self):
        if self._direct_import_running:
            return
        try:
            from netatmo import LoginWindow
        except ImportError as e:
            messagebox.showerror(
                'Netatmo Direktimport',
                f'netatmo.py konnte nicht geladen werden:\n{e}')
            return

        login_win = tk.Toplevel(self.root)
        login_win.transient(self.root)
        login_win.grab_set()

        def on_success(creds):
            login_win.grab_release()
            login_win.destroy()
            self._ask_direct_import_range(creds)

        LoginWindow(login_win, on_success=on_success)

    def _ask_direct_import_range(self, creds: dict):
        dialog = tk.Toplevel(self.root)
        dialog.title('Netatmo Direktimport')
        dialog.geometry('360x190')
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()

        frame = ttk.Frame(dialog, padding=16)
        frame.pack(fill=tk.BOTH, expand=True)
        frame.columnconfigure(1, weight=1)

        default_start = self._default_direct_start_date().strftime('%Y-%m-%d')
        default_end = datetime.datetime.now().strftime('%Y-%m-%d')
        start_var = tk.StringVar(value=default_start)
        end_var = tk.StringVar(value=default_end)

        ttk.Label(frame, text='Startdatum (JJJJ-MM-TT):').grid(
            row=0, column=0, sticky='w', pady=5, padx=(0, 8))
        ttk.Entry(frame, textvariable=start_var).grid(
            row=0, column=1, sticky='ew', pady=5)
        ttk.Label(frame, text='Enddatum (JJJJ-MM-TT):').grid(
            row=1, column=0, sticky='w', pady=5, padx=(0, 8))
        ttk.Entry(frame, textvariable=end_var).grid(
            row=1, column=1, sticky='ew', pady=5)
        ttk.Label(
            frame,
            text='Es werden alle verfügbaren Module und Sensoren geladen.',
            foreground='gray', wraplength=320).grid(
            row=2, column=0, columnspan=2, sticky='w', pady=(4, 10))

        btns = ttk.Frame(frame)
        btns.grid(row=3, column=0, columnspan=2, sticky='e')

        def cancel():
            dialog.grab_release()
            dialog.destroy()

        def start():
            try:
                start_date = datetime.datetime.strptime(
                    start_var.get().strip(), '%Y-%m-%d')
                end_date = datetime.datetime.strptime(
                    end_var.get().strip(), '%Y-%m-%d')
            except ValueError:
                messagebox.showerror(
                    'Netatmo Direktimport', 'Ungültiges Datumsformat.')
                return
            if end_date <= start_date:
                messagebox.showerror(
                    'Netatmo Direktimport', 'Enddatum muss nach dem Startdatum liegen.')
                return
            # Enddatum ist einschließlich: bis Ende des gewählten Tages laden,
            # nicht nur bis Mitternacht am Tagesbeginn (sonst fehlen heutige Daten).
            end_date_inclusive = min(
                end_date + datetime.timedelta(days=1),
                datetime.datetime.now())
            dialog.grab_release()
            dialog.destroy()
            self._run_direct_import(creds, start_date, end_date_inclusive)

        ttk.Button(btns, text='Abbrechen', command=cancel).pack(
            side=tk.RIGHT, padx=(6, 0))
        ttk.Button(btns, text='Laden', command=start).pack(side=tk.RIGHT)

    def _run_direct_import(self, creds: dict, start_date: datetime.datetime,
                           end_date: datetime.datetime):
        self._direct_import_running = True
        self.btn_direct.config(state=tk.DISABLED)
        self.status.set('Netatmo Direktimport startet…')

        thread = threading.Thread(
            target=self._direct_import_worker,
            args=(creds, start_date, end_date),
            daemon=True)
        thread.start()

    def _direct_import_worker(self, creds: dict, start_date: datetime.datetime,
                              end_date: datetime.datetime):
        try:
            from netatmo import NetatmoDataDownloader

            downloader = NetatmoDataDownloader(
                creds['client_id'], creds['client_secret'], creds['access_token'])
            stations_data = downloader.get_stations_data()
            if not stations_data:
                self.root.after(0, lambda: self._finish_direct_import(
                    None, 'Konnte Stationsdaten nicht laden.'))
                return

            devices = stations_data.get('body', {}).get('devices', [])
            if not devices:
                self.root.after(0, lambda: self._finish_direct_import(
                    None, 'Keine Netatmo-Geräte gefunden.'))
                return

            device_id = devices[0].get('_id')
            modules = downloader.extract_modules(devices)
            total_requests = sum(len(m.get('data_types', []))
                                 for m in modules.values())
            completed = 0
            rows = []

            for module_id, module_info in modules.items():
                module_name = module_info.get('name', module_id)
                for sensor_type in module_info.get('data_types', []):
                    completed += 1
                    self.root.after(0, lambda c=completed, t=total_requests,
                                    m=module_name, s=sensor_type:
                                    self.status.set(
                                        f'Lade Netatmo {c}/{t}: {m} / {s}…'))
                    rows.extend(downloader.get_sensor_data(
                        device_id, module_id, module_name, sensor_type,
                        start_date, end_date))

            data = self._convert_netatmo_rows(rows)
            self.root.after(0, lambda: self._finish_direct_import(
                data,
                f'Netatmo API {start_date:%Y-%m-%d} bis {end_date:%Y-%m-%d}'))
        except (AttributeError, ImportError, KeyError, OSError, RuntimeError, TypeError, ValueError) as e:
            self.root.after(0, lambda err=e: self._finish_direct_import(
                None, f'Unerwarteter Fehler: {err}'))

    def _convert_netatmo_rows(self, rows: list) -> list:
        result = []
        for row in rows:
            try:
                dt = row['Sortierung']
                value = float(row['Wert'])
                result.append({
                    'ts': int(dt.timestamp() * 1000),
                    'date': dt.date().isoformat(),
                    'module': str(row.get('Module', '') or '').strip(),
                    'sensor': str(row.get('Messwert', '') or '').strip(),
                    'value': value,
                    'unit': str(row.get('Einheit', '') or '').strip(),
                })
            except (ValueError, TypeError, KeyError, AttributeError):
                continue
        return result

    def _finish_direct_import(self, new_data, source_or_error: str):
        self._direct_import_running = False
        self.btn_direct.config(state=tk.NORMAL)
        if new_data is None:
            messagebox.showerror('Netatmo Direktimport', source_or_error)
            self.status.set('Netatmo Direktimport fehlgeschlagen.')
            return
        if not new_data:
            messagebox.showwarning(
                'Netatmo Direktimport', 'Keine Daten im gewählten Zeitraum gefunden.')
            self.status.set('Netatmo Direktimport ohne Daten beendet.')
            return

        before = len(self._data)
        self._data, added, dupes = merge_data(self._data, new_data)
        if source_or_error not in self._source_files:
            self._source_files.append(source_or_error)
        archive_file, arch_added, arch_dupes = self._save_to_archive(
            new_data, source_or_error)
        self._refresh_info(
            cache_note=f'Netatmo Δ: +{added:,} neu / {dupes:,} Duplikate ignoriert')
        self.btn_open.config(state=tk.NORMAL)
        self.btn_export_pages.config(state=tk.NORMAL)
        self._save_to_cache()
        self.status.set(
            f'Netatmo Direktimport: {before:,} → {len(self._data):,} '
            f'(+{added:,} neu, {dupes:,} Duplikate übersprungen). '
            f'Archiv: +{arch_added:,} neu / {arch_dupes:,} bekannt '
            f'({os.path.basename(archive_file)}).')
        if self._autopush_var.get():
            threading.Thread(
                target=self._auto_push_github,
                args=(list(self._data),),
                daemon=True).start()

    def _auto_push_github(self, data: list) -> None:
        import subprocess

        def _run(cmd, **kw):
            """Run git command, raise with stderr on failure."""
            r = subprocess.run(cmd, cwd=repo_root,
                               capture_output=True, text=True, **kw)
            if r.returncode != 0:
                detail = (r.stderr or r.stdout or '').strip()
                raise RuntimeError(f"{' '.join(cmd)}\n{detail}")
            return r

        repo_root = os.path.dirname(os.path.abspath(__file__))
        docs_dir = os.path.join(repo_root, 'docs')
        os.makedirs(docs_dir, exist_ok=True)

        self.root.after(0, lambda: self.status.set(
            'GitHub Pages: exportiere HTML + data.json…'))
        try:
            data_file = os.path.join(docs_dir, 'data.json')
            with open(data_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, separators=(',', ':'))

            payload = prepare_chart_payload(data)
            html = generate_html(payload)
            export_html(html)
            open(os.path.join(docs_dir, '.nojekyll'), 'a').close()

            # Ermittle den aktuellen Branch (funktioniert mit main, master, gh-pages)
            branch = _run(['git', 'branch', '--show-current']).stdout.strip()

            today = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
            self.root.after(0, lambda: self.status.set(
                f'GitHub Pages: git push → {branch}…'))

            # data.json nicht in git (zu groß); nur HTML und .nojekyll
            _run(['git', 'add', 'docs/index.html', 'docs/.nojekyll'])
            diff = subprocess.run(
                ['git', 'diff', '--cached', '--quiet'], cwd=repo_root)
            if diff.returncode != 0:
                _run(['git', 'commit', '-m', f'auto: Netatmo Update {today}'])
                _run(['git', 'push', 'origin', branch])
                self.root.after(0, lambda t=today: self.status.set(
                    f'✅ GitHub Pages aktualisiert ({t}) – '
                    'live in ~1 Min.: https://falcon237.github.io/WeatherApp/'))
            else:
                self.root.after(0, lambda: self.status.set(
                    'GitHub Pages: keine Änderungen – kein Push nötig.'))
        except Exception as exc:
            err = str(exc)
            self.root.after(0, lambda e=err: self.status.set(
                f'❌ GitHub Push fehlgeschlagen: {e}'))

    def _refresh_info(self, cache_note: str = ''):
        if not self._data:
            text = 'Keine Daten geladen.'
        else:
            modules = sorted({d['module'] for d in self._data})
            sensors = sorted({d['sensor'] for d in self._data})
            dates = [d['date'] for d in self._data]
            min_d, max_d = min(dates), max(dates)
            sources = ('\nQuellen:      '
                       + '\n              '.join(
                           os.path.basename(p) for p in self._source_files)
                       ) if self._source_files else ''
            note = f'  [{cache_note}]' if cache_note else ''
            text = (
                f"Datenpunkte:  {len(self._data):,}{note}\n"
                f"Zeitraum:     {min_d}  →  {max_d}\n"
                f"Module ({len(modules)}):   {', '.join(modules)}\n"
                f"Sensoren ({len(sensors)}): {', '.join(sensors)}"
                f"\nArchiv:      {archive_path()}"
                f"{sources}"
            )
        self.info_text.config(state=tk.NORMAL)
        self.info_text.delete('1.0', tk.END)
        self.info_text.insert(tk.END, text)
        self.info_text.config(state=tk.DISABLED)

    def _open(self):
        if not self._data:
            messagebox.showwarning(
                'Keine Daten', 'Bitte zuerst eine Datei laden.')
            return
        self.status.set('Erstelle HTML…')
        self.root.update()
        try:
            payload = prepare_chart_payload(self._data)
            html = generate_html(payload)
            path = open_in_browser(html)
            self.status.set(
                f'Browser geöffnet. Temp-Datei: {os.path.basename(path)}')
        except (OSError, ValueError, KeyError) as e:
            messagebox.showerror('Fehler', str(e))
            self.status.set('Fehler beim Öffnen.')

    def _open_docs_snapshot(self):
        """Öffnet lokal vorhandenes docs/index.html (GitHub Pages Snapshot)."""
        repo_root = os.path.dirname(os.path.abspath(__file__))
        docs_index = os.path.join(repo_root, 'docs', 'index.html')
        if not os.path.isfile(docs_index):
            messagebox.showerror(
                'Nicht gefunden', f'Datei nicht gefunden:\n{docs_index}')
            return
        try:
            webbrowser.open(f'file:///{docs_index.replace(os.sep, "/")}')
            self.status.set(
                f'Browser geöffnet: {os.path.basename(docs_index)}')
        except OSError as e:
            messagebox.showerror('Fehler', str(e))
            self.status.set('Fehler beim Öffnen.')

    def _export_for_github_pages(self):
        if not self._data:
            messagebox.showwarning(
                'Keine Daten', 'Bitte zuerst eine Datei laden.')
            return
        self.status.set('Exportiere docs/index.html…')
        self.root.update()
        try:
            payload = prepare_chart_payload(self._data)
            html = generate_html(payload)
            path = export_html(html)
            self.status.set(f'GitHub-Pages-App exportiert: {path}')
            messagebox.showinfo(
                'GitHub Pages Export',
                'Die veröffentlichbare App wurde erstellt:\n\n'
                f'{path}\n\n'
                'In GitHub Pages danach als Quelle "main" und Ordner '
                '"/docs" auswählen. Der Link öffnet direkt diese App.')
        except (OSError, ValueError, KeyError) as e:
            messagebox.showerror('Fehler beim Export', str(e))
            self.status.set('Fehler beim Export.')


def main():
    root = tk.Tk()
    ViewerApp(root)
    root.mainloop()


if __name__ == '__main__':
    main()
